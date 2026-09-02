import logging
import logging.handlers
import aioodbc
import json
import datetime
import calendar
import time
import random
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, UploadFile, File, Query, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Any
from decimal import Decimal
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from passlib.context import CryptContext
from jose import JWTError, jwt
import asyncio
from fastapi.concurrency import run_in_threadpool
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from fastapi import Request 
from dotenv import load_dotenv

load_dotenv()

# --- Logging Configuration ---
# Previously NSSM captured stdout/stderr to files (AppStdout/AppStderr).
# Now the app manages its own log files, independent of how it's launched.
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

LOG_FORMAT = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
log_formatter = logging.Formatter(LOG_FORMAT)

root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)

# General log: INFO and above, rotates daily at midnight, keeps 14 days
info_file_handler = logging.handlers.TimedRotatingFileHandler(
    filename=os.path.join(LOG_DIR, "app.log"),
    when="midnight",
    backupCount=14,
    encoding="utf-8",
)
info_file_handler.setLevel(logging.INFO)
info_file_handler.setFormatter(log_formatter)
root_logger.addHandler(info_file_handler)

# Dedicated error log: ERROR and above only, rotates at 5MB, keeps 10 backups
error_file_handler = logging.handlers.RotatingFileHandler(
    filename=os.path.join(LOG_DIR, "error.log"),
    maxBytes=5 * 1024 * 1024,
    backupCount=10,
    encoding="utf-8",
)
error_file_handler.setLevel(logging.ERROR)
error_file_handler.setFormatter(log_formatter)
root_logger.addHandler(error_file_handler)

# Console output too, useful when running interactively (e.g. `python main.py`)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(log_formatter)
root_logger.addHandler(console_handler)

# Route uvicorn's own loggers into the same handlers as above.
# This must run at import time (not just inside `if __name__ == "__main__"`),
# because when the app is launched as `uvicorn main:app ...` (e.g. from NSSM),
# this module is imported rather than executed as a script, so the
# `__main__` guard never runs. Uvicorn otherwise installs its own handlers
# on these loggers with propagate=False, which is why uvicorn.access /
# uvicorn.error lines were missing from app.log in production.
for uv_logger_name in ("uvicorn", "uvicorn.error", "uvicorn.access"):
    uv_logger = logging.getLogger(uv_logger_name)
    uv_logger.handlers = []
    uv_logger.propagate = True

logger = logging.getLogger(__name__)

# --- Auth Configuration ---
SECRET_KEY = "CHANGE_THIS_TO_A_SUPER_SECRET_KEY"  # IMPORTANT: Change this!
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 43200 # 30 days expiration

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")


scheduler = AsyncIOScheduler()

# --- Initialization ---

async def startup_db():
    """Ensure the LogCalculator SQL Server database has the required tables."""
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        # Gate table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Gate' AND xtype='U')
            CREATE TABLE Gate (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                gate_name  NVARCHAR(255),
                from_loc   NVARCHAR(255),
                to_loc     NVARCHAR(255),
                uom        NVARCHAR(100),
                unit       INT,
                cost       DECIMAL(18,6),
                created_at NVARCHAR(30),
                created_by INT,
                edited_at  NVARCHAR(30),
                edited_by  INT
            )
        """)

        # Item_Pricing table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Item_Pricing' AND xtype='U')
            CREATE TABLE Item_Pricing (
                id                  INT PRIMARY KEY,
                gate_id             INT,
                bu                  NVARCHAR(100),
                item_id             NVARCHAR(255),
                item_name           NVARCHAR(500),
                principal           NVARCHAR(255),
                brand               NVARCHAR(255),
                transportation_cost NVARCHAR(255),
                created_at          NVARCHAR(30),
                created_by          INT,
                edited_at           NVARCHAR(30),
                edited_by           INT,
                FOREIGN KEY (gate_id) REFERENCES Gate(id)
            )
        """)

        # Calculation_History table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Calculation_History' AND xtype='U')
            CREATE TABLE Calculation_History (
                id                 BIGINT PRIMARY KEY,
                created_at         NVARCHAR(30),
                gate_name          NVARCHAR(255),
                from_loc           NVARCHAR(255),
                to_loc             NVARCHAR(255),
                doc_nums           NVARCHAR(MAX),
                total_weight       DECIMAL(18,6),
                manual_total_cost  DECIMAL(18,6),
                additional_charges DECIMAL(18,6),
                final_total_cost   DECIMAL(18,6),
                channel            NVARCHAR(100),
                status             NVARCHAR(50)  DEFAULT 'saved',
                created_by         INT,
                submitted_by       INT,
                submitted_at       NVARCHAR(30),
                claimed_by         INT,
                claimed_at         NVARCHAR(30)
            )
        """)
        
        
        # Add destination to Gate table if it doesn't exist
        await cursor.execute("""
            IF COL_LENGTH('Gate', 'destination') IS NULL
            BEGIN
                ALTER TABLE Gate ADD destination NVARCHAR(255);
            END
        """)

        # Add historical gate snapshot columns if they don't exist
        await cursor.execute("""
            IF COL_LENGTH('Calculation_History', 'gate_cost') IS NULL
            BEGIN
                ALTER TABLE Calculation_History ADD gate_cost DECIMAL(18,6);
                ALTER TABLE Calculation_History ADD gate_uom NVARCHAR(100);
                ALTER TABLE Calculation_History ADD gate_unit INT;
            END
        """)

        # Add POSM total cost column if it doesn't exist
        await cursor.execute("""
            IF COL_LENGTH('Calculation_History', 'posm_total_cost') IS NULL
            BEGIN
                ALTER TABLE Calculation_History ADD posm_total_cost DECIMAL(18,6);
            END
        """)

        # Add total_weight column if it doesn't exist
        await cursor.execute("""
            IF COL_LENGTH('Calculation_History', 'total_weight') IS NULL
            BEGIN
                ALTER TABLE Calculation_History ADD total_weight DECIMAL(18,6);
            END
        """)

        # Add warehouse code and edited value columns to Calculation_Products if they don't exist
        await cursor.execute("""
            IF COL_LENGTH('Calculation_Products', 'FromWhsCode') IS NULL
            BEGIN
                ALTER TABLE Calculation_Products ADD FromWhsCode NVARCHAR(255);
                ALTER TABLE Calculation_Products ADD ToWhsCode NVARCHAR(255);
            END
            IF COL_LENGTH('Calculation_Products', 'edited_ctns') IS NULL
            BEGIN
                ALTER TABLE Calculation_Products ADD edited_ctns DECIMAL(18,6);
                ALTER TABLE Calculation_Products ADD edited_weight DECIMAL(18,6);
            END
        """)

        # Calculation_Products table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Calculation_Products' AND xtype='U')
            CREATE TABLE Calculation_Products (
                id                 INT IDENTITY(1,1) PRIMARY KEY,
                calc_id            BIGINT NOT NULL,
                code               NVARCHAR(255),
                name               NVARCHAR(500),
                weight             DECIMAL(18,6),
                doc_date           NVARCHAR(30),
                sin_no             NVARCHAR(255),
                principal          NVARCHAR(255),
                brand              NVARCHAR(255),
                ctns               DECIMAL(18,6),
                bu                 NVARCHAR(100),
                b_code             NVARCHAR(255),
                b_name             NVARCHAR(500),
                b_dept             NVARCHAR(255),
                b_principal        NVARCHAR(255),
                b_desc             NVARCHAR(500),
                s_dept             NVARCHAR(255),
                s_principal        NVARCHAR(255),
                calculation_type   NVARCHAR(100),
                system_rate        DECIMAL(18,6),
                unit_cost          DECIMAL(18,6),
                total_cost         DECIMAL(18,6),
                standard_unit_cost DECIMAL(18,6),
                FOREIGN KEY (calc_id) REFERENCES Calculation_History(id)
            )
        """)

        # Calculation_POSM_Products table (POSM Calculation line items, separate from Calculation_Products)
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Calculation_POSM_Products' AND xtype='U')
            CREATE TABLE Calculation_POSM_Products (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                calc_id     BIGINT NOT NULL,
                department  NVARCHAR(255),
                item_name   NVARCHAR(500),
                uom         NVARCHAR(100),
                quantity    DECIMAL(18,6),
                unit_cost   DECIMAL(18,6),
                total_cost  DECIMAL(18,6),
                b_code      NVARCHAR(255),
                b_name      NVARCHAR(500),
                b_dept      NVARCHAR(255),
                b_principal NVARCHAR(255),
                b_desc      NVARCHAR(500),
                s_dept      NVARCHAR(255),
                s_principal NVARCHAR(255),
                FOREIGN KEY (calc_id) REFERENCES Calculation_History(id)
            )
        """)

        # Rate_Cart table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Rate_Cart' AND xtype='U')
            CREATE TABLE Rate_Cart (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                location   NVARCHAR(255) UNIQUE,
                cost       DECIMAL(18,6),
                created_at NVARCHAR(30),
                created_by INT,
                edited_at  NVARCHAR(30),
                edited_by  INT
            )
        """)


        # Daily_Driver_Override table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Daily_Driver_Override' AND xtype='U')
            CREATE TABLE Daily_Driver_Override (
                id                INT IDENTITY(1,1) PRIMARY KEY,
                target_date       NVARCHAR(10) NOT NULL,
                driver_name       NVARCHAR(255) NOT NULL,
                override_amount   DECIMAL(18,6) NULL,
                override_ctns     DECIMAL(18,6) NULL,
                created_at        NVARCHAR(30),
                created_by        INT,
                edited_at         NVARCHAR(30),
                edited_by         INT,
                UNIQUE (target_date, driver_name)
            )
        """)
        
        # Add override_ctns column if it doesn't exist
        await cursor.execute("""
            IF COL_LENGTH('Daily_Driver_Override', 'override_ctns') IS NULL
            BEGIN
                ALTER TABLE Daily_Driver_Override ADD override_ctns DECIMAL(18,6) NULL;
            END
        """)
        
        # Add new override and weight columns to Daily_Item_Report if they don't exist
        await cursor.execute("""
            IF COL_LENGTH('Daily_Item_Report', 'override_rate_cart_cost') IS NULL
            BEGIN
                ALTER TABLE Daily_Item_Report ADD override_rate_cart_cost DECIMAL(18,6);
                ALTER TABLE Daily_Item_Report ADD override_driver_total_ctns DECIMAL(18,6);
            END
            IF COL_LENGTH('Daily_Item_Report', 'weight') IS NULL
            BEGIN
                ALTER TABLE Daily_Item_Report ADD weight DECIMAL(18,6);
                ALTER TABLE Daily_Item_Report ADD driver_total_weight DECIMAL(18,6);
            END
            IF COL_LENGTH('Daily_Item_Report', 'volumetric_weight') IS NULL
            BEGIN
                ALTER TABLE Daily_Item_Report ADD volumetric_weight DECIMAL(18,6);
                ALTER TABLE Daily_Item_Report ADD driver_total_volumetric_weight DECIMAL(18,6);
            END
        """)

        # Add new override and weight columns to Daily_Township_Report if they don't exist
        await cursor.execute("""
            IF COL_LENGTH('Daily_Township_Report', 'override_rate_cart_cost') IS NULL
            BEGIN
                ALTER TABLE Daily_Township_Report ADD override_rate_cart_cost DECIMAL(18,6);
                ALTER TABLE Daily_Township_Report ADD override_driver_total_ctns DECIMAL(18,6);
            END
            IF COL_LENGTH('Daily_Township_Report', 'weight') IS NULL
            BEGIN
                ALTER TABLE Daily_Township_Report ADD weight DECIMAL(18,6);
                ALTER TABLE Daily_Township_Report ADD driver_total_weight DECIMAL(18,6);
            END
            IF COL_LENGTH('Daily_Township_Report', 'volumetric_weight') IS NULL
            BEGIN
                ALTER TABLE Daily_Township_Report ADD volumetric_weight DECIMAL(18,6);
                ALTER TABLE Daily_Township_Report ADD driver_total_volumetric_weight DECIMAL(18,6);
            END
        """)

        # Daily_Item_Report table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Daily_Item_Report' AND xtype='U')
            CREATE TABLE Daily_Item_Report (
                id                INT IDENTITY(1,1) PRIMARY KEY,
                target_date       NVARCHAR(10) NOT NULL,
                bu                NVARCHAR(100),
                branch            NVARCHAR(255),
                driver_name       NVARCHAR(255),
                item_code         NVARCHAR(255),
                item_name         NVARCHAR(500),
                principal         NVARCHAR(255),
                brand             NVARCHAR(255),
                ctns              DECIMAL(18,6),
                allocated_cost    DECIMAL(18,6),
                cost_per_carton   DECIMAL(18,6),
                driver_total_ctns DECIMAL(18,6),
                rate_cart_cost    DECIMAL(18,6),
                sales_amount      DECIMAL(18,6),
                created_at        NVARCHAR(30),
                created_by        INT,
                edited_at         NVARCHAR(30),
                edited_by         INT
            )
        """)

        # Daily_Township_Report table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Daily_Township_Report' AND xtype='U')
            CREATE TABLE Daily_Township_Report (
                id                  INT IDENTITY(1,1) PRIMARY KEY,
                target_date         NVARCHAR(10) NOT NULL,
                branch              NVARCHAR(255),
                driver_name         NVARCHAR(255),
                township            NVARCHAR(255),
                customer_code       NVARCHAR(255),
                contact_person      NVARCHAR(500),
                ctns                DECIMAL(18,6),
                driver_total_ctns   DECIMAL(18,6),
                rate_cart_cost      DECIMAL(18,6),
                cost_per_carton     DECIMAL(18,6),
                allocated_cost      DECIMAL(18,6),
                total_drop_points   DECIMAL(18,6),
                cost_per_drop_point DECIMAL(18,6),
                sales_amount        DECIMAL(18,6),
                created_at          NVARCHAR(30),
                created_by          INT,
                edited_at           NVARCHAR(30),
                edited_by           INT
            )
        """)

        # User_Activity_Log table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='User_Activity_Log' AND xtype='U')
            CREATE TABLE User_Activity_Log (
                id        INT IDENTITY(1,1) PRIMARY KEY,
                user_id   INT,   
                action    NVARCHAR(255),
                details   NVARCHAR(MAX),
                timestamp NVARCHAR(30)
            )
        """)

        # --- User Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Users' AND xtype='U')
            CREATE TABLE Users (
                id              INT IDENTITY(1,1) PRIMARY KEY,
                username        NVARCHAR(255) UNIQUE,
                hashed_password NVARCHAR(500),
                role            NVARCHAR(100),
                created_at      NVARCHAR(30),
                created_by      INT,
                edited_at       NVARCHAR(30),
                edited_by       INT
            )
        """)

        # --- Roles Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Roles' AND xtype='U')
            CREATE TABLE Roles (
                name        NVARCHAR(100) PRIMARY KEY,
                permissions NVARCHAR(MAX),
                created_at  NVARCHAR(30),
                created_by  INT,
                edited_at   NVARCHAR(30),
                edited_by   INT
            )
        """)

        # --- System Settings Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='System_Settings' AND xtype='U')
            CREATE TABLE System_Settings (
                setting_key NVARCHAR(255) PRIMARY KEY,
                setting_value NVARCHAR(255),
                updated_at NVARCHAR(30),
                updated_by INT
            )
        """)
        
        await cursor.execute("SELECT setting_value FROM System_Settings WHERE setting_key = 'volumetric_divisor'")
        if not await cursor.fetchone():
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await cursor.execute("INSERT INTO System_Settings (setting_key, setting_value, updated_at) VALUES (?, ?, ?)", ('volumetric_divisor', '5000', now_str))

        # --- Reference Tables ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Locations' AND xtype='U')
            CREATE TABLE Locations (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                name       NVARCHAR(255) UNIQUE,
                created_at NVARCHAR(30),
                created_by INT
            )
        """)

        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Rate_Cart_Locations' AND xtype='U')
            CREATE TABLE Rate_Cart_Locations (
                id   INT IDENTITY(1,1) PRIMARY KEY,
                name NVARCHAR(255) UNIQUE
            )
        """)

        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='UOMs' AND xtype='U')
            CREATE TABLE UOMs (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                name       NVARCHAR(100) UNIQUE,
                created_at NVARCHAR(30),
                created_by INT
            )
        """)

        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Channels' AND xtype='U')
            CREATE TABLE Channels (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                name       NVARCHAR(255) UNIQUE,
                created_at NVARCHAR(30),
                created_by INT
            )
        """)

        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Departments' AND xtype='U')
            CREATE TABLE Departments (
                id         INT IDENTITY(1,1) PRIMARY KEY,
                name       NVARCHAR(255) UNIQUE,
                created_at NVARCHAR(30),
                created_by INT
            )
        """)

        # --- Branch Code Mapping Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Branch_Code' AND xtype='U')
            CREATE TABLE Branch_Code (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                log_pric    NVARCHAR(255) UNIQUE,
                code        NVARCHAR(100),
                name        NVARCHAR(500),
                dept        NVARCHAR(255),
                principal   NVARCHAR(255),
                description NVARCHAR(500),
                created_at  NVARCHAR(30),
                created_by  INT,
                edited_at   NVARCHAR(30),
                edited_by   INT
            )
        """)

        # Branch Code Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Branch_Code_Change_Log' AND xtype='U')
            CREATE TABLE Branch_Code_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                log_pric    NVARCHAR(255),
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX)
            )
        """)

        # SD Code Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='SD_Code_Change_Log' AND xtype='U')
            CREATE TABLE SD_Code_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                log_pric    NVARCHAR(255),
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX)
            )
        """)

        # --- SD Code Mapping Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='SD_Code' AND xtype='U')
            CREATE TABLE SD_Code (
                id        INT IDENTITY(1,1) PRIMARY KEY,
                channel   NVARCHAR(100),
                code      NVARCHAR(100),
                name      NVARCHAR(500),
                dept      NVARCHAR(255),
                principal NVARCHAR(255),
                log_pric  NVARCHAR(255) UNIQUE,
                created_at NVARCHAR(30),
                created_by INT,
                edited_at  NVARCHAR(30),
                edited_by  INT
            )
        """)

        # Gate Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Gate_Change_Log' AND xtype='U')
            CREATE TABLE Gate_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                gate_id     INT,
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX),
                FOREIGN KEY (gate_id) REFERENCES Gate(id)
            )
        """)

        # Item Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Item_Change_Log' AND xtype='U')
            CREATE TABLE Item_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                pricing_id  INT,
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX),
                FOREIGN KEY (pricing_id) REFERENCES Item_Pricing(id)
            )
        """)

        # Rate Cart Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Rate_Cart_Change_Log' AND xtype='U')
            CREATE TABLE Rate_Cart_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                location    NVARCHAR(255),
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX),
                FOREIGN KEY (location) REFERENCES Rate_Cart(location)
            )
        """)

        # --- Location Mapping Table ---
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Location_Mapping' AND xtype='U')
            CREATE TABLE Location_Mapping (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                to_location NVARCHAR(255) UNIQUE,
                branch_code NVARCHAR(100),
                created_at  NVARCHAR(30),
                created_by  INT
            )
        """)
        
        # --- Add tracking columns to existing tables if they don't exist ---
        tables_to_update = ['Gate', 'Item_Pricing', 'Rate_Cart', 'Daily_Driver_Override', 'Branch_Code', 'SD_Code', 'Users', 'Roles', 'Daily_Item_Report', 'Daily_Township_Report']
        for table in tables_to_update:
            await cursor.execute(f"""
                IF COL_LENGTH('{table}', 'created_at') IS NULL
                BEGIN
                    ALTER TABLE {table} ADD created_at NVARCHAR(30), created_by INT, edited_at NVARCHAR(30), edited_by INT;
                END
            """)
            
        tables_to_update_created_only = ['Locations', 'Rate_Cart_Locations', 'UOMs', 'Channels', 'Departments', 'Location_Mapping']
        for table in tables_to_update_created_only:
            await cursor.execute(f"""
                IF COL_LENGTH('{table}', 'created_at') IS NULL
                BEGIN
                    ALTER TABLE {table} ADD created_at NVARCHAR(30), created_by INT;
                END
            """)
            await cursor.execute(f"""
                IF COL_LENGTH('{table}', 'edited_at') IS NULL
                BEGIN
                    ALTER TABLE {table} ADD edited_at NVARCHAR(30), edited_by INT;
                END
            """)
        
        # Driver Override Change Log Table
        await cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Driver_Override_Change_Log' AND xtype='U')
            CREATE TABLE Driver_Override_Change_Log (
                id          INT IDENTITY(1,1) PRIMARY KEY,
                target_date NVARCHAR(10),
                driver_name NVARCHAR(255),
                changed_by  INT,  
                change_date NVARCHAR(30),
                field_name  NVARCHAR(255),
                old_value   NVARCHAR(MAX),
                new_value   NVARCHAR(MAX)
            )
        """)
        
        # Pre-seed with existing defaults to save you time
        await cursor.execute("SELECT COUNT(*) FROM Location_Mapping")
        if (await cursor.fetchone())[0] == 0:
            default_mappings = [
                ("Yangon", "YGN"), ("Mandalay", "MDY"), ("Naypyitaw", "NPT"),
                ("Magaway", "MGW"), ("Taunggyi", "TGI"), ("Taunggu", "TGU"),
                ("Pathein", "PTN"), ("Mawlamyine", "MLM"), ("Bago", "BGO")
            ]
            for to_loc, br_code in default_mappings:
                await cursor.execute(
                    "IF NOT EXISTS (SELECT 1 FROM Location_Mapping WHERE to_location=?) INSERT INTO Location_Mapping (to_location, branch_code) VALUES (?,?)",
                    (to_loc, to_loc, br_code)
                )

        # Create default users
        await cursor.execute("SELECT * FROM Users WHERE username = 'account'")
        if not await cursor.fetchone():
            default_pw = pwd_context.hash("account123") 
            await cursor.execute("INSERT INTO Users (username, hashed_password, role) VALUES (?, ?, ?)", ('account', default_pw, 'account'))
            
        await cursor.execute("SELECT * FROM Users WHERE username = 'logistic'")
        if not await cursor.fetchone():
            log_pw = pwd_context.hash("log123")
            await cursor.execute("INSERT INTO Users (username, hashed_password, role) VALUES (?, ?, ?)", ('logistic', log_pw, 'logistic'))
            
        await cursor.execute("SELECT * FROM Users WHERE username = 'admin'")
        if not await cursor.fetchone():
            admin_pw = pwd_context.hash("admin123")
            await cursor.execute("INSERT INTO Users (username, hashed_password, role) VALUES (?, ?, ?)", ('admin', admin_pw, 'admin'))

        # Seed references
        default_locs = [
            ('Yangon',), ('Mandalay',), ('Naypyitaw',), ('Magaway',), 
            ('Taunggyi',), ('Taunggu',), ('Pathein',), ('Mawlamyine',),
            ('Taungtwingyi',), ('Meikhtila',), ('Dawei',), ('Myingyan',),
            ('Yae',), ('Chauk',), ('Aunglan',), ('Danuphyu',),
            ('Ngathaingchaung',), ('Yamethin',), ('Phyue',), ('Kyauktagar',),
            ('Nyaunglaypin',), ('Nyaungoo',), ('Kyaukpadaung',), ('Myeik',),
            ('Twantay',)
        ]

        default_locsrc = [
            ('YGN',), ('MDY',), ('BGO',), ('MGW',), 
            ('TGI',), ('TGU',), ('PTN',), ('MLM',),
            ('NPT',)
        ]

        await cursor.execute("SELECT COUNT(*) FROM Locations")
        if (await cursor.fetchone())[0] == 0:
            for (loc_name,) in default_locs:
                await cursor.execute("IF NOT EXISTS (SELECT 1 FROM Locations WHERE name=?) INSERT INTO Locations (name) VALUES (?)", (loc_name, loc_name))

        await cursor.execute("SELECT COUNT(*) FROM Rate_Cart_Locations")
        if (await cursor.fetchone())[0] == 0:
            for (rc_name,) in default_locsrc:
                await cursor.execute("IF NOT EXISTS (SELECT 1 FROM Rate_Cart_Locations WHERE name=?) INSERT INTO Rate_Cart_Locations (name) VALUES (?)", (rc_name, rc_name))

        await cursor.execute("SELECT COUNT(*) FROM UOMs")
        if (await cursor.fetchone())[0] == 0:
            default_uoms = [('Kg',), ('Ton',)]
            for (uom_name,) in default_uoms:
                await cursor.execute("IF NOT EXISTS (SELECT 1 FROM UOMs WHERE name=?) INSERT INTO UOMs (name) VALUES (?)", (uom_name, uom_name))

        await cursor.execute("SELECT COUNT(*) FROM Channels")
        if (await cursor.fetchone())[0] == 0:
            default_channels = [('SD',), ('Branch',), ('Telecom Branch',), ('Telecom SD',), ('Outlet',)]
            for (ch_name,) in default_channels:
                await cursor.execute("IF NOT EXISTS (SELECT 1 FROM Channels WHERE name=?) INSERT INTO Channels (name) VALUES (?)", (ch_name, ch_name))
        
        await conn.commit()
        await conn.close()
        logger.info("Logistic DB initialized successfully")

    except Exception as e:
        logger.error(f"Error initializing database: {str(e)}")
        raise  # or handle explicitly — don't let this hide scheduler failures

    # Scheduler setup outside the DB migration try block
    scheduler.add_job(daily_job_generator, 'cron', hour=23, minute=55, misfire_grace_time=600, coalesce=True)
    scheduler.add_job(cleanup_old_activity_logs, 'cron', hour=1, minute=0)
    scheduler.start()
    logger.info("Schedulers started.")

def _route_uvicorn_logs_to_file():
    """Ensure uvicorn's loggers feed into our file handlers.

    Uvicorn applies its own logging config (handlers + propagate=False)
    when the server starts. Depending on whether the process was launched
    via `python main.py` or `uvicorn main:app ...` (as NSSM does), that
    can happen after our module-level setup runs. Re-applying here, inside
    the lifespan startup, guarantees it takes effect last either way.
    """
    for uv_logger_name in ("uvicorn", "uvicorn.error", "uvicorn.access"):
        uv_logger = logging.getLogger(uv_logger_name)
        uv_logger.handlers = []
        uv_logger.propagate = True

@asynccontextmanager
async def lifespan(app: FastAPI):
    _route_uvicorn_logs_to_file()
    await startup_db()
    yield

app = FastAPI(lifespan=lifespan)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
async def health_check():
    """Lightweight liveness endpoint for the external watchdog.
    Deliberately does not touch the database - it only needs to prove
    the HTTP listener is alive and accepting connections."""
    return {"status": "ok"}

# --- Database Connections ---

async def get_dwbi_connection():
    """Create and return a SQL Server connection to DWBI (Read-Only Source)"""
    
    # Securely fetch credentials from the .env file
    db_user = os.getenv("DWBI_USER")
    db_password = os.getenv("DWBI_PASSWORD")
    
    # Optional but recommended: Safety check to ensure variables loaded correctly
    if not db_user or not db_password:
        logger.error("Missing DWBI database credentials in .env file!")
        raise ValueError("Database credentials are not configured properly.")

    conn_str = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=phm\\reportingsvr;'
        'DATABASE=DWBI;'
        f'UID={db_user};'
        f'PWD={db_password};'
    )
    return await aioodbc.connect(dsn=conn_str, autocommit=False)

async def get_logistic_connection():
    """Create and return a SQL Server connection to LogCalculator (Read/Write Source)"""
    db_user = os.getenv("DWBI_USER")
    db_password = os.getenv("DWBI_PASSWORD")

    if not db_user or not db_password:
        logger.error("Missing DWBI database credentials in .env file!")
        raise ValueError("Database credentials are not configured properly.")

    conn_str = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=phm\\reportingsvr;'
        'DATABASE=LogCalculator;'
        f'UID={db_user};'
        f'PWD={db_password};'
    )
    return await aioodbc.connect(dsn=conn_str, autocommit=False)

# --- Helper: Activity Logger ---
async def log_user_activity(user_id: int, action: str, details: str = ""):
    """Inserts a new record into the User_Activity_Log table.

    IMPORTANT: takes a trusted user_id (e.g. current_user['id'], which is
    validated against the DB on every request) rather than a username.
    Usernames are mutable and can be reassigned/reused; JWTs live for 30
    days and embed the username at login time, so re-resolving a username
    back to an id at log time can silently attribute the action to the
    WRONG user if that username was edited/reassigned in the meantime.
    """
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        await cursor.execute(
            "INSERT INTO User_Activity_Log (user_id, action, details, timestamp) VALUES (?, ?, ?, ?)",
            (user_id, action, details, now_str)
        )
        await conn.commit()
        await conn.close()
    except Exception as e:
        logger.error(f"Failed to log user activity: {str(e)}")

async def cleanup_old_activity_logs():
    """Automated job to delete activity logs older than 30 days."""
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        # Calculate the threshold date (30 days ago)
        threshold_date = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
        
        await cursor.execute("DELETE FROM User_Activity_Log WHERE timestamp < ?", (threshold_date,))
        deleted_count = cursor.rowcount
        
        await conn.commit()
        await conn.close()
        
        if deleted_count > 0:
            logger.info(f"Cleaned up {deleted_count} system activity logs older than {threshold_date}.")
    except Exception as e:
        logger.error(f"Failed to clean up old activity logs: {str(e)}")

# --- Pydantic Models ---

class BranchCodeData(BaseModel):
    original_log_pric: Optional[str] = None
    log_pric: str
    code: Optional[str] = ""
    name: Optional[str] = ""
    dept: Optional[str] = ""
    principal: Optional[str] = ""
    description: Optional[str] = ""

class SDCodeData(BaseModel):
    original_log_pric: Optional[str] = None
    channel: Optional[str] = ""
    code: Optional[str] = ""
    name: Optional[str] = ""
    dept: Optional[str] = ""
    principal: Optional[str] = ""
    log_pric: str

class LocationMappingItem(BaseModel):
    to_location: str
    branch_code: str

class BranchCodeLogItem(BaseModel):
    id: int
    log_pric: str
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

class SDCodeLogItem(BaseModel):
    id: int
    log_pric: str
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

class DriverOverrideData(BaseModel):
    target_date: str
    driver_name: str
    override_amount: Optional[Decimal] = None
    override_ctns: Optional[Decimal] = None

class RateCartData(BaseModel):
    location: str
    cost: Decimal

class RateCartLogItem(BaseModel):
    id: int
    location: str
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

class GateData(BaseModel):
    gate_id: Optional[int] = None
    gate_name: str
    from_loc: str  
    to_loc: str
    destination: Optional[str] = None
    uom: Optional[str] = None       
    unit: Optional[int] = None      
    cost: Optional[Decimal] = None 
    original_gate_name: Optional[str] = None

class GateLogItem(BaseModel):
    id: int
    gate_id: int
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

class ItemLogItem(BaseModel):
    id: int
    pricing_id: int
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

class ItemPricingData(BaseModel):
    pricing_id: Optional[int] = None
    gate_id: int
    bu: Optional[str] = ""
    item_code: str
    item_name: str
    principal: Optional[str] = ""
    brand: Optional[str] = ""
    transportation_cost: str
    original_item_code: Optional[str] = None

class POSMItem(BaseModel):
    department: Optional[str] = ""
    item_name: str
    uom: Optional[str] = ""
    quantity: Decimal

class CalculationSaveRequest(BaseModel):
    id: Optional[int] = None
    gate_name: str
    from_loc: str
    to_loc: str
    doc_nums: List[str]
    total_weight: Optional[Decimal] = None
    manual_total_cost: Optional[Decimal] = None
    additional_charges: Optional[Decimal] = Decimal("0.0")
    final_total_cost: Decimal
    channel: Optional[str] = ""
    status: Optional[str] = "saved"
    calculated_products: List[Any] = []
    # --- NEW FIELDS ---
    gate_cost: Optional[Decimal] = None
    gate_uom: Optional[str] = None
    gate_unit: Optional[int] = None
    # --- POSM CALCULATION (optional) ---
    posm_items: List[Any] = []
    posm_total_cost: Optional[Decimal] = None
    posm_products: List[Any] = []

class ReferenceItem(BaseModel):
    name: str

class ReferenceEditItem(BaseModel):
    id: int
    new_name: str

class SettingUpdate(BaseModel):
    value: str

class LocationMappingEditItem(BaseModel):
    original_to_location: str
    new_to_location: str
    new_branch_code: str

class UserCreate(BaseModel):
    username: str
    password: str
    role: str

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None

class UserResponse(BaseModel):
    id: int
    username: str
    role: str

class RoleCreate(BaseModel):
    name: str
    permissions: List[str]

class RoleUpdate(BaseModel):
    permissions: List[str]

class RoleResponse(BaseModel):
    name: str
    permissions: List[str]

class Token(BaseModel):
    access_token: str
    token_type: str
    role: str
    username: str
    permissions: List[str]
    id: int
    
class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

class ActivityLogResponse(BaseModel):
    id: int
    username: Optional[str] = None
    action: str
    details: str
    timestamp: str

class ActivityLogPaginatedResponse(BaseModel):
    total: int
    logs: List[ActivityLogResponse]

# --- Auth Helpers ---
    
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_id: int = payload.get("id")
        
        if username is None:
            raise credentials_exception
            
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT u.username, u.role, r.permissions 
            FROM Users u
            LEFT JOIN Roles r ON u.role = r.name
            WHERE u.id = ?
        """, (user_id,))
        user_record = await cursor.fetchone()
        await conn.close()

        if not user_record:
            raise credentials_exception 
            
        fresh_username = user_record[0]
        fresh_role = user_record[1]
        fresh_permissions = json.loads(user_record[2]) if user_record[2] else []

        return {"id": user_id, "username": fresh_username, "role": fresh_role, "permissions": fresh_permissions}
        
    except JWTError:
        raise credentials_exception


@app.get("/users/me")
async def get_user_me(current_user: dict = Depends(get_current_user)):
    """Returns the fresh user session data directly from the DB"""
    return current_user

def require_permission(perm: str):
    async def permission_checker(current_user: dict = Depends(get_current_user)):
        if perm not in current_user.get("permissions", []):
            raise HTTPException(status_code=403, detail=f"Requires '{perm}' permission")
        return current_user
    return permission_checker

# --- Login & Token ---

@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    conn = await get_logistic_connection()
    cursor = await conn.cursor()
    await cursor.execute("""
        SELECT u.id, u.username, u.hashed_password, u.role, r.permissions 
        FROM Users u
        LEFT JOIN Roles r ON u.role = r.name
        WHERE u.username = ?
    """, (form_data.username,))
    user = await cursor.fetchone()
    await conn.close()
    
    if not user or not verify_password(form_data.password, user[2]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    permissions = json.loads(user[4]) if user[4] else []
    access_token = create_access_token(data={"sub": user[1], "id": user[0], "role": user[3], "permissions": permissions})
    
    await log_user_activity(user[0], "LOGIN", "User authenticated successfully")
    
    return {"access_token": access_token, "token_type": "bearer", "role": user[3], "username": user[1], "permissions": permissions, "id": user[0]}

# --- System Activity Log Endpoint ---

@app.get("/admin/activity-logs", response_model=ActivityLogPaginatedResponse)
async def get_activity_logs(
    limit: int = Query(50, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    timestamp: Optional[str] = Query(None, description="Filter by timestamp"),
    username: Optional[str] = Query(None, description="Filter by username"),
    action: Optional[str] = Query(None, description="Filter by action type"),
    details: Optional[str] = Query(None, description="Filter by details"),
    user: dict = Depends(require_permission("view_activity_logs"))
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        query = """
            SELECT l.id, COALESCE(u.username, '(deleted user)') AS username, l.action, l.details, l.timestamp 
            FROM User_Activity_Log l
            LEFT JOIN Users u ON l.user_id = u.id
            WHERE 1=1
        """
        count_query = """
            SELECT COUNT(*) 
            FROM User_Activity_Log l
            LEFT JOIN Users u ON l.user_id = u.id
            WHERE 1=1
        """
        params = []
        
        if timestamp:
            query += " AND l.timestamp LIKE ?"
            count_query += " AND l.timestamp LIKE ?"
            params.append(f"%{timestamp}%")
        if username:
            query += " AND u.username LIKE ?"
            count_query += " AND u.username LIKE ?"
            params.append(f"%{username}%")
        if action:
            query += " AND l.action LIKE ?"
            count_query += " AND l.action LIKE ?"
            params.append(f"%{action}%")
        if details:
            query += " AND l.details LIKE ?"
            count_query += " AND l.details LIKE ?"
            params.append(f"%{details}%")
            
        query += " ORDER BY l.timestamp DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
        
        await cursor.execute(count_query, params)
        total_count = (await cursor.fetchone())[0]
        
        params.extend([offset, limit])
        await cursor.execute(query, params)
        rows = await cursor.fetchall()
        
        await conn.close()
        
        logs = [{"id": r[0], "username": r[1], "action": r[2], "details": r[3], "timestamp": r[4]} for r in rows]
        return {"total": total_count, "logs": logs}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching activity logs: {str(e)}")

# --- Role Management Endpoints ---

@app.get("/roles", response_model=List[RoleResponse])
async def get_all_roles(user: dict = Depends(require_permission("view_roles"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name, permissions FROM Roles ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"name": row[0], "permissions": json.loads(row[1]) if row[1] else []} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching roles: {str(e)}")

@app.post("/roles")
async def create_role(role_data: RoleCreate, user: dict = Depends(require_permission("add_role"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM Roles WHERE name = ?", (role_data.name,))
        if await cursor.fetchone():
            await conn.close()
            raise HTTPException(status_code=400, detail="Role already exists")
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("INSERT INTO Roles (name, permissions, created_at, created_by) VALUES (?, ?, ?, ?)", 
                      (role_data.name, json.dumps(role_data.permissions), now_str, user['id']))
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "CREATE_ROLE", f"Created role: {role_data.name}")
        return {"message": "Role created successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating role: {str(e)}")

@app.put("/roles/{role_name}")
async def update_role(role_name: str, role_data: RoleUpdate, user: dict = Depends(require_permission("edit_role"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Roles SET permissions = ?, edited_at = ?, edited_by = ? WHERE name = ?", 
                      (json.dumps(role_data.permissions), now_str, user['id'], role_name))
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "UPDATE_ROLE", f"Updated permissions for role: {role_name}")
        return {"message": "Role updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating role: {str(e)}")

@app.delete("/roles/{role_name}")
async def delete_role(role_name: str, user: dict = Depends(require_permission("delete_role"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT COUNT(*) FROM Users WHERE role = ?", (role_name,))
        if (await cursor.fetchone())[0] > 0:
            await conn.close()
            raise HTTPException(status_code=400, detail="Cannot delete role currently assigned to users")
        await cursor.execute("DELETE FROM Roles WHERE name = ?", (role_name,))
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "DELETE_ROLE", f"Deleted role: {role_name}")
        return {"message": "Role deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting role: {str(e)}")

# --- User Management Endpoints ---

@app.get("/users", response_model=List[UserResponse])
async def get_all_users(user: dict = Depends(require_permission("view_users"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, username, role FROM Users ORDER BY username")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "username": row[1], "role": row[2]} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching users: {str(e)}")
    

@app.post("/users")
async def create_user(user_data: UserCreate, user: dict = Depends(require_permission("add_user"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT username FROM Users WHERE username = ?", (user_data.username,))
        if await cursor.fetchone():
            await conn.close()
            raise HTTPException(status_code=400, detail="Username already exists")
        hashed_pw = pwd_context.hash(user_data.password)
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("INSERT INTO Users (username, hashed_password, role, created_at, created_by) VALUES (?, ?, ?, ?, ?)", 
                      (user_data.username, hashed_pw, user_data.role, now_str, user['id']))
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "CREATE_USER", f"Created user: {user_data.username} with role: {user_data.role}")
        return {"message": "User created successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating user: {str(e)}")

@app.put("/users/{user_id}")
async def update_user(user_id: int, user_data: UserUpdate, user: dict = Depends(require_permission("edit_user"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT username FROM Users WHERE id = ?", (user_id,))
        target_user = await cursor.fetchone()
        if not target_user:
            await conn.close()
            raise HTTPException(status_code=404, detail="User not found")
            
        changes = []
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if user_data.username and user_data.username != target_user[0]:
            await cursor.execute("SELECT id FROM Users WHERE username = ?", (user_data.username,))
            if await cursor.fetchone():
                await conn.close()
                raise HTTPException(status_code=400, detail="Username already exists")
            await cursor.execute("UPDATE Users SET username = ?, edited_at = ?, edited_by = ? WHERE id = ?", (user_data.username, now_str, user['id'], user_id))
            changes.append(f"username to {user_data.username}")

        if user_data.password:
            hashed_pw = pwd_context.hash(user_data.password)
            await cursor.execute("UPDATE Users SET hashed_password = ?, edited_at = ?, edited_by = ? WHERE id = ?", (hashed_pw, now_str, user['id'], user_id))
            changes.append("password")
        if user_data.role:
            await cursor.execute("UPDATE Users SET role = ?, edited_at = ?, edited_by = ? WHERE id = ?", (user_data.role, now_str, user['id'], user_id))
            changes.append(f"role to {user_data.role}")
            
        await conn.commit()
        await conn.close()
        
        if changes:
            await log_user_activity(user['id'], "UPDATE_USER", f"Updated user ID {user_id} ({', '.join(changes)})")
        return {"message": "User updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating user: {str(e)}")

@app.delete("/users/{user_id}")
async def delete_user(user_id: int, user: dict = Depends(require_permission("delete_user"))):
    if user_id == user.get("id"):
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        await cursor.execute("SELECT username FROM Users WHERE id = ?", (user_id,))
        target_user = await cursor.fetchone()
        
        await cursor.execute("DELETE FROM Users WHERE id = ?", (user_id,))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="User not found")
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "DELETE_USER", f"Deleted user: {target_user[0]}")
        return {"message": "User deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting user: {str(e)}")
    
@app.put("/users/me/password")
async def change_password(data: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        await cursor.execute("SELECT hashed_password FROM Users WHERE id = ?", (current_user["id"],))
        row = await cursor.fetchone()
        
        if not row:
            await conn.close()
            raise HTTPException(status_code=404, detail="User not found")
            
        current_hashed_password = row[0]
        
        if not verify_password(data.old_password, current_hashed_password):
            await conn.close()
            raise HTTPException(status_code=400, detail="Incorrect old password")
            
        new_hashed_password = pwd_context.hash(data.new_password)
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Users SET hashed_password = ?, edited_at = ?, edited_by = ? WHERE id = ?", (new_hashed_password, now_str, current_user['id'], current_user['id']))
        
        await conn.commit()
        await conn.close()
        
        await log_user_activity(current_user['id'], "CHANGE_PASSWORD", "User changed their own password")
        return {"message": "Password changed successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error changing password: {str(e)}")

# --- Helper Functions for Calculation ---
async def determine_calculation_type_sql(gate_id):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT cost, uom FROM Gate WHERE id = ?", (gate_id,))
        row = await cursor.fetchone()

        await conn.close()

        if row and row[0] is not None:
            try:
                price = Decimal(str(row[0]))
                uom = str(row[1]).strip().lower() if row[1] else ""
                if price > 0:
                    return "per_trip_pricing" if uom == "trip" else "gate_pricing"
            except (ValueError, TypeError):
                pass
        return "direct_pricing"
    except Exception as e:
        logger.error(f"Error determining calc type: {str(e)}")
        return "unknown"

def get_rounded_ctns(val):
    if not val:
        return Decimal("0")
    try:
        f_val = Decimal(str(val))
        if f_val <= 0:
            return Decimal("0")
        # Ensure the returned value is explicitly a Decimal
        return Decimal(str(max(1, round(f_val))))
    except (ValueError, TypeError):
        return Decimal("0")
    
    
async def _perform_calculation_logic(gate_name, doc_nums, from_loc=None, to_loc=None, manual_total_cost=None, additional_charges=Decimal("0.0"), posm_items=None, posm_total_cost=None, edited_items_json=None):
    add_charges = Decimal(str(additional_charges)) if additional_charges is not None else Decimal("0.0")

    pg_nums = [str(d).replace("PG - ", "").replace("PG-", "") for d in doc_nums if not str(d).startswith("PDG")]
    pdg_nums = [str(d).replace("PDG - ", "").replace("PDG-", "") for d in doc_nums if str(d).startswith("PDG")]

    pick_rows = []

    try:
        conn_dwbi = await get_dwbi_connection()
        cursor_dwbi = await conn_dwbi.cursor()

        if pg_nums:
            placeholders = ','.join('?' * len(pg_nums))
            query_pg = f"""
                SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(ItemWeight), MAX(DocDate), DocNum, MAX(Principal), SUM(BatchQtyByCtn), MAX(Brand), 'PG', MAX(FromWhsCod), MAX(WhsCode)
                FROM PG_Transfer_Details 
                WHERE DocNum IN ({placeholders}) 
                GROUP BY DocNum, ItemCode
                ORDER BY DocNum, ItemCode
            """
            await cursor_dwbi.execute(query_pg, pg_nums)
            pick_rows.extend(await cursor_dwbi.fetchall())

        if pdg_nums:
            placeholders = ','.join('?' * len(pdg_nums))
            query_pdg = f"""
                SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(LineTotalWeight), MAX(DocDate), DocNum, MAX(Principal), SUM(QtyCtn), MAX(Brand), 'PDG', MAX(FromWhsCod), MAX(WhsCode)
                FROM PDG_Transfer_Details 
                WHERE DocNum IN ({placeholders}) 
                GROUP BY DocNum, ItemCode
                ORDER BY DocNum, ItemCode
            """
            await cursor_dwbi.execute(query_pdg, pdg_nums)
            pick_rows.extend(await cursor_dwbi.fetchall())

        await conn_dwbi.close()
    except Exception as e:
        raise Exception(f"Error fetching transfer details: {str(e)}")
    
    if not pick_rows:
        raise Exception("No products found for the selected Doc Nums")
        
    # --- NEW: Process overrides and calculate remaining amounts ---
    conn_log = await get_logistic_connection()
    cursor_log = await conn_log.cursor()
    
    sin_nos = []
    for row in pick_rows:
        prefix = row[9] # 'PG' or 'PDG'
        docnum = row[5] # DocNum
        if docnum:
            sin_nos.append(f"{prefix} - {docnum}")
            
    sin_nos = list(set(sin_nos))
    used_map = {}
    
    if sin_nos:
        placeholders = ','.join('?' * len(sin_nos))
        await cursor_log.execute(f"""
            SELECT cp.sin_no, cp.code, SUM(COALESCE(cp.edited_ctns, cp.ctns)), SUM(COALESCE(cp.edited_weight, cp.weight))
            FROM Calculation_Products cp
            JOIN Calculation_History ch ON cp.calc_id = ch.id
            WHERE ch.status IN ('submitted', 'claimed')
            AND cp.sin_no IN ({placeholders})
            GROUP BY cp.sin_no, cp.code
        """, sin_nos)
        used_amounts_rows = await cursor_log.fetchall()
        for r in used_amounts_rows:
            sin_no = r[0]
            code = r[1]
            used_ctns = Decimal(str(r[2])) if r[2] is not None else Decimal("0.0")
            used_weight = Decimal(str(r[3])) if r[3] is not None else Decimal("0.0")
            used_map[(sin_no, code)] = {"used_ctns": used_ctns, "used_weight": used_weight}
            
    await conn_log.close()

    edited_items = json.loads(edited_items_json) if edited_items_json else {}

    item_totals = {}
    for row in pick_rows:
        code = row[0]
        prefix = row[9]
        docnum = row[5]
        sin_no = f"{prefix} - {docnum}" if docnum else ""
        
        orig_ctns = get_rounded_ctns(row[7])
        orig_weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
        
        used = used_map.get((sin_no, code), {"used_ctns": Decimal("0.0"), "used_weight": Decimal("0.0")})
        remaining_ctns = orig_ctns - used["used_ctns"]
        remaining_weight = orig_weight - used["used_weight"]
        
        if remaining_ctns < Decimal("0.01"): remaining_ctns = Decimal("0")
        if remaining_weight < Decimal("0.01"): remaining_weight = Decimal("0")
        
        if code not in item_totals:
            item_totals[code] = {'ctns': Decimal("0.0"), 'weight': Decimal("0.0")}
            
        item_totals[code]['ctns'] += remaining_ctns
        item_totals[code]['weight'] += remaining_weight

    adjusted_pick_rows = []
    for row in pick_rows:
        row_list = list(row)
        code = row_list[0]
        prefix = row_list[9]
        docnum = row_list[5]
        sin_no = f"{prefix} - {docnum}" if docnum else ""
        
        orig_ctns = get_rounded_ctns(row_list[7])
        orig_weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
        
        used = used_map.get((sin_no, code), {"used_ctns": Decimal("0.0"), "used_weight": Decimal("0.0")})
        default_ctns = orig_ctns - used["used_ctns"]
        default_weight = orig_weight - used["used_weight"]
        
        if default_ctns < Decimal("0.01"): default_ctns = Decimal("0")
        if default_weight < Decimal("0.01"): default_weight = Decimal("0")

        if code in edited_items:
            edited_total_ctns = Decimal(str(edited_items[code].get('ctns', 0)))
            rem_total_ctns = item_totals.get(code, {}).get('ctns', Decimal("0.0"))
            if rem_total_ctns > Decimal("0"):
                ratio = edited_total_ctns / rem_total_ctns
                final_ctns = default_ctns * ratio
                final_weight = default_weight * ratio
            else:
                final_ctns = Decimal("0")
                final_weight = Decimal("0")
        else:
            final_ctns = default_ctns
            final_weight = default_weight
            
        if final_ctns != orig_ctns or final_weight != orig_weight:
            original_ctns_val = orig_ctns
            original_weight_val = orig_weight
        else:
            original_ctns_val = None
            original_weight_val = None
            
        row_list[7] = final_ctns
        row_list[3] = final_weight
        
        row_list.append(original_ctns_val) # index 12
        row_list.append(original_weight_val) # index 13
        row_list.append(used["used_ctns"]) # index 14
        row_list.append(used["used_weight"]) # index 15
        row_list.append(orig_ctns) # index 16
        row_list.append(orig_weight) # index 17
        
        adjusted_pick_rows.append(tuple(row_list))
        
    pick_rows = adjusted_pick_rows
    # --- END NEW ---
    
    
    try:
        conn_log = await get_logistic_connection()
        cursor_log = await conn_log.cursor()
        
        if from_loc and to_loc:
            await cursor_log.execute("SELECT id, from_loc, to_loc, cost, unit, uom FROM Gate WHERE gate_name = ? AND from_loc = ? AND to_loc = ?", (gate_name, from_loc, to_loc))
        else:
            await cursor_log.execute("SELECT id, from_loc, to_loc, cost, unit, uom FROM Gate WHERE gate_name = ?", (gate_name,))
            
        gate_row = await cursor_log.fetchone()
        
        await cursor_log.execute("SELECT log_pric, code, name, dept, principal, description FROM Branch_Code")
        branch_code_map = {row[0].strip().lower(): {
            "Code": row[1], "Name": row[2], "Dept": row[3], "Principal": row[4], "Description": row[5]
        } for row in await cursor_log.fetchall() if row[3]}

        await cursor_log.execute("SELECT dept, principal, log_pric FROM SD_Code")
        sd_code_map = {row[2].strip().lower(): {
            "Dept": row[0], "Principal": row[1]
        } for row in await cursor_log.fetchall() if row[2]}

    except Exception as e:
        raise Exception(f"Error fetching local configs: {str(e)}")
    
    if not gate_row:
        if 'conn_log' in locals(): await conn_log.close()
        raise Exception(f"Gate {gate_name} not found")
        
    gate_id = gate_row[0]
    matched_from_loc = gate_row[1]
    matched_to_loc = gate_row[2]
    cost = Decimal(str(gate_row[3] or 0))
    gate_unit = Decimal(str(gate_row[4] or 1.0))
    gate_uom = str(gate_row[5]).strip().lower() if gate_row[5] else ""
    
    await cursor_log.execute("SELECT item_id, transportation_cost FROM Item_Pricing WHERE gate_id = ?", (gate_id,))
    pricing_rows = await cursor_log.fetchall()
    await conn_log.close()
    
    item_pricing = {}
    for row in pricing_rows:
        i_code = row[0]
        t_cost = str(row[1]).strip() if row[1] else ""
        if not t_cost or t_cost.lower() == 'nan' or t_cost.lower() == 'none' or t_cost == '':
            item_pricing[i_code] = {'type': 'ton', 'value': None}
        else:
            try:
                val = Decimal(t_cost)
                item_pricing[i_code] = {'type': 'direct', 'value': val}
            except:
                item_pricing[i_code] = {'type': 'unknown', 'value': None}

    if cost > Decimal("0") and gate_uom == "trip": calc_type = "per_trip_pricing"
    elif cost > Decimal("0"): calc_type = "gate_pricing"
    else: calc_type = "direct_pricing"

    calculated_products = []
    total_cost = Decimal("0.0")
    estimated_total_cost = Decimal("0.0")

    if calc_type == "gate_pricing":
        ton_items = []
        direct_items = []
        ton_cost_total = Decimal("0.0")
        
        for row in pick_rows:
            doc_date_val = row[4]
            if isinstance(doc_date_val, datetime.datetime):
                doc_date_str = doc_date_val.strftime("%Y-%m-%d")
            elif isinstance(doc_date_val, str) and len(doc_date_val) >= 10:
                doc_date_str = doc_date_val[:10]
            else:
                doc_date_str = str(doc_date_val) if doc_date_val else ""

            principal_val = row[6] or ""
            brand_val = row[8] or ""
            bc_info = branch_code_map.get(principal_val.strip().lower(), {})
            sd_info = sd_code_map.get(principal_val.strip().lower(), {})

            item_data = {
                "code": row[0] if row[0] else "",
                "name": row[1] if row[1] else "",
                "uom": row[2] if row[2] else "",
                "weight": row[3],
                "doc_date": doc_date_str,         
                "sin_no": f"{row[9]} - {str(row[5])}" if row[5] else "",           
                "principal": principal_val,
                "brand": brand_val,
                "ctns": row[7],
                "original_ctns": row[12],
                "original_weight": row[13],
                "used_ctns": row[14] if len(row) > 14 else Decimal("0.0"),
                "used_weight": row[15] if len(row) > 15 else Decimal("0.0"),
                "original_total_ctns": row[16] if len(row) > 16 else row[7],
                "original_total_weight": row[17] if len(row) > 17 else row[3],
                "bu": row[9],
                "b_code": bc_info.get("Code", ""),
                "b_name": bc_info.get("Name", ""),
                "b_dept": bc_info.get("Dept", ""),
                "b_principal": bc_info.get("Principal", ""),
                "b_desc": bc_info.get("Description", ""),
                "s_dept": sd_info.get("Dept", ""),
                "s_principal": sd_info.get("Principal", ""),
                "FromWhsCode": row[10] if len(row) > 10 and row[10] else "",
                "ToWhsCode": row[11] if len(row) > 11 and row[11] else ""
            }
            
            p_info = item_pricing.get(item_data['code'], {})
            p_type = p_info.get('type', 'ton')
            p_val = p_info.get('value', Decimal("0.0"))
            
            if p_type == 'direct':
                estimated_total_cost += (item_data['ctns'] * p_val)
                item_data['standard_unit_cost'] = p_val
                direct_items.append(item_data)
            else:
                effective_rate = cost / gate_unit if gate_unit > Decimal("0") else cost
                cost_item = item_data['weight'] * effective_rate
                estimated_total_cost += cost_item
                ton_cost_total += cost_item
                item_data['total_cost'] = cost_item
                ton_items.append(item_data)

        direct_unit_cost = Decimal("0.0")
        if manual_total_cost is not None:
            manual_total_cost_dec = Decimal(str(manual_total_cost))
            
            # --- NEW VALIDATION: Prevent negative values ---
            if manual_total_cost_dec < ton_cost_total:
                raise Exception(f"Total Cost (Manual Override) cannot be less than the total weight-based cost ({ton_cost_total:,.2f} MMK).")
            # -----------------------------------------------
            
            remainder = manual_total_cost_dec - ton_cost_total
            total_direct_ctns = sum(item['ctns'] for item in direct_items)
            if total_direct_ctns > Decimal("0"): direct_unit_cost = remainder / total_direct_ctns
            total_cost = manual_total_cost_dec
        else:
            total_cost = estimated_total_cost

        for item in ton_items:
            avg_unit_cost = item['total_cost'] / item['ctns'] if item['ctns'] > Decimal("0") else Decimal("0.0")
            calculated_products.append({
                **item, "calculation_type": "weight", "system_rate": None,
                "unit_cost": avg_unit_cost, "total_cost": item['total_cost'] 
            })
        
        for item in direct_items:
            final_unit_cost = direct_unit_cost if manual_total_cost is not None else item['standard_unit_cost']
            final_item_cost = item['ctns'] * final_unit_cost
            calculated_products.append({
                **item, "calculation_type": "override" if manual_total_cost else "direct",
                "system_rate": item['standard_unit_cost'], "unit_cost": final_unit_cost, "total_cost": final_item_cost
            })

    elif calc_type == "per_trip_pricing":
        # Per Trip gates have a single flat cost for the whole trip (no per-item pricing
        # configured). Allocate that flat cost across items proportionally by ctn quantity.
        trip_total = Decimal(str(manual_total_cost)) if manual_total_cost is not None else cost

        trip_items = []
        total_ctns_all = Decimal("0.0")
        for row in pick_rows:
            doc_date_val = row[4]
            if isinstance(doc_date_val, datetime.datetime):
                doc_date_str = doc_date_val.strftime("%Y-%m-%d")
            elif isinstance(doc_date_val, str) and len(doc_date_val) >= 10:
                doc_date_str = doc_date_val[:10]
            else:
                doc_date_str = str(doc_date_val) if doc_date_val else ""

            principal_val = row[6] or ""
            brand_val = row[8] or ""
            bc_info = branch_code_map.get(principal_val.strip().lower(), {})
            sd_info = sd_code_map.get(principal_val.strip().lower(), {})

            item_data = {
                "code": row[0] if row[0] else "",
                "name": row[1] if row[1] else "",
                "uom": row[2] if row[2] else "",
                "weight": row[3],
                "doc_date": doc_date_str,
                "sin_no": f"{row[9]} - {str(row[5])}" if row[5] else "",
                "principal": principal_val,
                "brand": brand_val,
                "ctns": row[7],
                "original_ctns": row[12],
                "original_weight": row[13],
                "used_ctns": row[14] if len(row) > 14 else Decimal("0.0"),
                "used_weight": row[15] if len(row) > 15 else Decimal("0.0"),
                "original_total_ctns": row[16] if len(row) > 16 else row[7],
                "original_total_weight": row[17] if len(row) > 17 else row[3],
                "bu": row[9],
                "b_code": bc_info.get("Code", ""),
                "b_name": bc_info.get("Name", ""),
                "b_dept": bc_info.get("Dept", ""),
                "b_principal": bc_info.get("Principal", ""),
                "b_desc": bc_info.get("Description", ""),
                "s_dept": sd_info.get("Dept", ""),
                "s_principal": sd_info.get("Principal", ""),
                "FromWhsCode": row[10] if len(row) > 10 and row[10] else "",
                "ToWhsCode": row[11] if len(row) > 11 and row[11] else ""
            }
            total_ctns_all += item_data['ctns']
            trip_items.append(item_data)

        item_count = len(trip_items)
        for item in trip_items:
            if total_ctns_all > Decimal("0"):
                proportion = item['ctns'] / total_ctns_all
            else:
                proportion = (Decimal("1") / item_count) if item_count > 0 else Decimal("0")
            item_total_cost = trip_total * proportion
            estimated_total_cost += item_total_cost
            unit_cost = item_total_cost / item['ctns'] if item['ctns'] > Decimal("0") else Decimal("0.0")
            calculated_products.append({
                **item, "calculation_type": "per_trip", "system_rate": None,
                "unit_cost": unit_cost, "total_cost": item_total_cost
            })

        total_cost = trip_total

    elif calc_type == "direct_pricing":
        for row in pick_rows:
            doc_date_val = row[4]
            if isinstance(doc_date_val, datetime.datetime):
                doc_date_str = doc_date_val.strftime("%Y-%m-%d")
            elif isinstance(doc_date_val, str) and len(doc_date_val) >= 10:
                doc_date_str = doc_date_val[:10]
            else:
                doc_date_str = str(doc_date_val) if doc_date_val else ""

            item_code = row[0] if row[0] else ""
            pricing_info = item_pricing.get(item_code, {})
            
            weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
            ctns = get_rounded_ctns(row[7])
            principal_val = row[6] or ""
            brand_val = row[8] or ""
            bc_info = branch_code_map.get(principal_val.strip().lower(), {})
            sd_info = sd_code_map.get(principal_val.strip().lower(), {})

            unit_cost = pricing_info.get('value', Decimal("0.0")) or Decimal("0.0")
            item_cost = ctns * unit_cost
            
            total_cost += item_cost
            estimated_total_cost += item_cost
            
            calculated_products.append({
                "code": item_code, "name": row[1] if row[1] else "", "ctns": ctns,
                "uom": row[2] if row[2] else "", "weight": weight, "doc_date": doc_date_str,       
                "sin_no": f"{row[9]} - {str(row[5])}" if row[5] else "", "principal": principal_val, "brand": brand_val,
                "original_ctns": row[12], "original_weight": row[13],
                "used_ctns": row[14] if len(row) > 14 else Decimal("0.0"),
                "used_weight": row[15] if len(row) > 15 else Decimal("0.0"),
                "original_total_ctns": row[16] if len(row) > 16 else ctns,
                "original_total_weight": row[17] if len(row) > 17 else weight,
                "bu": row[9], "b_code": bc_info.get("Code", ""), "b_name": bc_info.get("Name", ""), 
                "b_dept": bc_info.get("Dept", ""), "b_principal": bc_info.get("Principal", ""), 
                "b_desc": bc_info.get("Description", ""), "s_dept": sd_info.get("Dept", ""), 
                "s_principal": sd_info.get("Principal", ""), "calculation_type": "direct", 
                "system_rate": unit_cost if unit_cost > Decimal("0") else None,
                "unit_cost": unit_cost, "total_cost": item_cost,
                "FromWhsCode": row[10] if len(row) > 10 and row[10] else "",
                "ToWhsCode": row[11] if len(row) > 11 and row[11] else ""
            })

    if add_charges != Decimal("0") and calculated_products:
        total_ctns = sum(p.get('ctns', Decimal("0")) for p in calculated_products)
        for p in calculated_products:
            proportion = p.get('ctns', Decimal("0")) / total_ctns if total_ctns > 0 else Decimal("0")
            extra_cost = add_charges * proportion
            p['total_cost'] += extra_cost
            if p['ctns'] > Decimal("0"): 
                p['unit_cost'] = p['total_cost'] / p['ctns']

    calculated_products.sort(key=lambda x: (x.get('sin_no', ''), x['code']))
    total_cost += add_charges
    estimated_total_cost += add_charges

    # --- POSM Calculation (optional add-on) ---
    # POSM items don't have a Principal like normal transfer products, so they are matched
    # against a dedicated "POSM" entry (log_pric = 'POSM') in Branch_Code / SD_Code mapping.
    posm_products = []
    posm_cost_dec = Decimal("0.0")
    if posm_items:
        posm_cost_dec = Decimal(str(posm_total_cost)) if posm_total_cost is not None else Decimal("0.0")
        total_qty = sum(Decimal(str(pi.get('quantity', 0) or 0)) for pi in posm_items)
        posm_bc_info = branch_code_map.get("posm", {})
        posm_sd_info = sd_code_map.get("posm", {})

        for pi in posm_items:
            qty = Decimal(str(pi.get('quantity', 0) or 0))
            unit_cost = (posm_cost_dec / total_qty) if total_qty > Decimal("0") else Decimal("0.0")
            item_total_cost = qty * unit_cost
            posm_products.append({
                "department": pi.get('department', ''),
                "item_name": pi.get('item_name', ''),
                "uom": pi.get('uom', ''),
                "quantity": qty,
                "unit_cost": unit_cost,
                "total_cost": item_total_cost,
                "principal": "POSM",
                "b_code": posm_bc_info.get("Code", ""),
                "b_name": posm_bc_info.get("Name", ""),
                "b_dept": posm_bc_info.get("Dept", ""),
                "b_principal": posm_bc_info.get("Principal", ""),
                "b_desc": posm_bc_info.get("Description", ""),
                "s_dept": posm_sd_info.get("Dept", ""),
                "s_principal": posm_sd_info.get("Principal", ""),
            })

        total_cost += posm_cost_dec
        estimated_total_cost += posm_cost_dec

    return {
        "calculation_type": calc_type, "gate_name": gate_name, "from_loc": matched_from_loc,
        "to_loc": matched_to_loc, "cost": cost, "additional_charges": add_charges,
        "calculated_products": calculated_products, "total_cost": total_cost,
        "estimated_total_cost": estimated_total_cost,
        "posm_products": posm_products, "posm_total_cost": posm_cost_dec
    }

# --- Rate Cart Endpoints ---

@app.get("/account/rate-cuts")
async def get_rate_carts(user: dict = Depends(require_permission("view_rate_carts"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT location, cost FROM Rate_Cart ORDER BY location")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"location": row[0], "cost": row[1]} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading rate carts: {str(e)}")

@app.post("/account/rate-cuts")
async def save_rate_cart(data: RateCartData, user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        await cursor.execute("SELECT cost FROM Rate_Cart WHERE location = ?", (data.location,))
        existing = await cursor.fetchone()
        
        change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        username = user['username']

        if existing:
            if "edit_rate_cart" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'edit_rate_cart' permission")
            
            old_cost = Decimal(str(existing[0]))
            if old_cost != data.cost:
                await cursor.execute("""
                    INSERT INTO Rate_Cart_Change_Log (location, changed_by, change_date, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (data.location, user['id'], change_date, 'Cost', str(old_cost), str(data.cost)))

            await cursor.execute("UPDATE Rate_Cart SET cost = ?, edited_at = ?, edited_by = ? WHERE location = ?", (data.cost, change_date, user['id'], data.location))
            await log_user_activity(user['id'], "UPDATE_RATE_CART", f"Updated rate cart for {data.location}")
        else:
            if "add_rate_cart" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'add_rate_cart' permission")
            await cursor.execute("INSERT INTO Rate_Cart (location, cost, created_at, created_by) VALUES (?, ?, ?, ?)", (data.location, data.cost, change_date, user['id']))
            await log_user_activity(user['id'], "ADD_RATE_CART", f"Added rate cart for {data.location}")
            
        await conn.commit()
        await conn.close()
        return {"message": "Rate cart saved successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error saving rate cart: {str(e)}")

@app.delete("/account/rate-cuts/{location}")
async def delete_rate_cart(location: str, user: dict = Depends(require_permission("delete_rate_cart"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("DELETE FROM Rate_Cart_Change_Log WHERE location = ?", (location,))
        await cursor.execute("DELETE FROM Rate_Cart WHERE location = ?", (location,))
        
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="Rate cart not found")
        
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "DELETE_RATE_CART", f"Deleted rate cart for {location}")
        return {"message": "Rate cart deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error deleting rate cart: {str(e)}")

@app.get("/account/rate-cuts/{location}/logs", response_model=List[RateCartLogItem])
async def get_rate_cart_logs(location: str, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.location, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM Rate_Cart_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.location = ? ORDER BY l.change_date DESC
        """, (location,))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "location": r[1], "changed_by": r[2], "change_date": r[3], "field_name": r[4], "old_value": r[5], "new_value": r[6]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching rate cart logs: {str(e)}")


# --- Daily Report Logic & Automation ---

async def get_rate_carts_for_date(target_date: str) -> dict:
    conn = await get_logistic_connection()
    cursor = await conn.cursor()
    
    await cursor.execute("SELECT location, cost FROM Rate_Cart")
    current_costs = {row[0].strip().upper(): Decimal(str(row[1])) for row in await cursor.fetchall()}
    
    query_threshold = target_date + " 23:59:59"
    await cursor.execute("""
        SELECT location, change_date, old_value 
        FROM Rate_Cart_Change_Log 
        WHERE change_date > ? 
        ORDER BY change_date DESC
    """, (query_threshold,))
    logs = await cursor.fetchall()
    await conn.close()

    historical_costs = current_costs.copy()
    for row in logs:
        loc = row[0].strip().upper()
        old_val_str = row[2]
        old_val = Decimal(str(old_val_str)) if old_val_str else Decimal("0.0")
        historical_costs[loc] = old_val

    return historical_costs

async def _get_daily_report_data(target_date: str):
    rate_carts = await get_rate_carts_for_date(target_date)

    conn_log = await get_logistic_connection()
    cursor_log = await conn_log.cursor()
    await cursor_log.execute("SELECT driver_name, override_amount, override_ctns FROM Daily_Driver_Override WHERE target_date = ?", (target_date,))
    overrides = {
        row[0].strip(): {
            'amount': Decimal(str(row[1])) if row[1] is not None else None,
            'ctns': Decimal(str(row[2])) if row[2] is not None else None
        } for row in await cursor_log.fetchall()
    }
    
    await cursor_log.execute("SELECT setting_value FROM System_Settings WHERE setting_key = 'volumetric_divisor'")
    div_row = await cursor_log.fetchone()
    volumetric_divisor = Decimal(div_row[0]) if div_row and div_row[0] else Decimal("5000.0")
    if volumetric_divisor <= Decimal("0"):
        volumetric_divisor = Decimal("5000.0")

    await conn_log.close()

    conn_dwbi = await get_dwbi_connection()
    cursor_dwbi = await conn_dwbi.cursor()
    query = """
        SELECT v.Branch, v.ItemCode, MAX(v.ItemName), MAX(v.Principal), MAX(v.Brand), v.[Driver Name], SUM(v.ctnQty), v.CustomerCode, MAX(v.ContactPerson), v.Township, SUM(v.SalesAmount), MAX(v.BU), MAX(i.BWeight1), MAX(i.BVolume)
        FROM VersaFleetDetail_TC v
        LEFT JOIN ItemMaster i ON v.ItemCode = i.ItemCode
        WHERE CONVERT(DATE, v.[Task Date]) = ? 
          AND v.[Task Status] = 'successful'
          AND v.ItemCode IS NOT NULL 
          AND LTRIM(RTRIM(v.ItemCode)) <> ''
        GROUP BY v.Branch, v.[Driver Name], v.ItemCode, v.CustomerCode, v.Township
    """
    await cursor_dwbi.execute(query, (target_date,))
    rows = await cursor_dwbi.fetchall()
    await conn_dwbi.close()

    granular_data = []
    driver_totals = {}
    driver_weight_totals = {}
    driver_volumetric_weight_totals = {}
    driver_customers = {} 

    for row in rows:
        branch = row[0].strip().upper() if row[0] else "UNKNOWN"
        item_code = row[1].strip() if row[1] else ""
        item_name = row[2].strip() if row[2] else ""
        principal = row[3].strip() if row[3] else ""
        brand = row[4].strip() if row[4] else ""
        driver_name = row[5].strip() if row[5] else ""
        ctns = Decimal(str(row[6] or 0))
        customer_code = row[7].strip() if row[7] else "UNKNOWN"
        
        contact_person_raw = row[8].strip() if row[8] else ""
        if " - " in contact_person_raw:
            contact_person = contact_person_raw.split(" - ", 1)[-1].strip()
        else:
            contact_person = contact_person_raw
        
        township = row[9].strip() if row[9] else "UNKNOWN"
        sales_amount = Decimal(str(row[10] or 0))
        bu = row[11].strip() if (len(row) > 11 and row[11]) else ""
        
        # Calculate Weight and Volumetric Weight
        bweight1 = Decimal(str(row[12] or 0)) if (len(row) > 12 and row[12] is not None) else Decimal("0.0")
        bvolume = Decimal(str(row[13] or 0)) if (len(row) > 13 and row[13] is not None) else Decimal("0.0")
        
        weight = ctns * bweight1
        
        # BVolume is in mm³. Convert to cm³ by dividing by 1000, then divide by volumetric_divisor for volumetric weight
        bvolume_cm3 = bvolume / Decimal("1000.0")
        volumetric_weight = (bvolume_cm3 / volumetric_divisor) * ctns

        granular_data.append({
            "branch": branch, "item_code": item_code, "item_name": item_name,
            "principal": principal, "brand": brand, "driver_name": driver_name,
            "ctns": ctns, "weight": weight, "volumetric_weight": volumetric_weight, "customer_code": customer_code, "contact_person": contact_person,
            "township": township, "sales_amount": sales_amount, "bu": bu
        })
        
        driver_key = (branch, driver_name)
        driver_totals[driver_key] = driver_totals.get(driver_key, Decimal("0.0")) + ctns
        driver_weight_totals[driver_key] = driver_weight_totals.get(driver_key, Decimal("0.0")) + weight
        driver_volumetric_weight_totals[driver_key] = driver_volumetric_weight_totals.get(driver_key, Decimal("0.0")) + volumetric_weight
        
        if driver_key not in driver_customers:
            driver_customers[driver_key] = set()
        driver_customers[driver_key].add(customer_code)

    item_report_dict = {}
    township_report_dict = {}

    for g in granular_data:
        b, d = g["branch"], g["driver_name"]
        d_total = driver_totals.get((b, d), Decimal("0.0"))
        d_weight_total = driver_weight_totals.get((b, d), Decimal("0.0"))
        d_volumetric_weight_total = driver_volumetric_weight_totals.get((b, d), Decimal("0.0"))
        b_cost = rate_carts.get(b, Decimal("0.0"))
        
        driver_override = overrides.get(d, {'amount': None, 'ctns': None})
        driver_extra = driver_override['amount']
        driver_extra_ctns = driver_override['ctns']
        
        # Use 0.0 for the actual math calculations so it doesn't crash
        math_extra_amount = driver_extra if driver_extra is not None else Decimal("0.0")
        math_extra_ctns = driver_extra_ctns if driver_extra_ctns is not None else Decimal("0.0")
        
        effective_rate_cart_cost = b_cost + math_extra_amount
        effective_driver_total_ctns = d_total + math_extra_ctns
        
        store_override_cost = effective_rate_cart_cost if driver_extra is not None else None
        store_override_ctns = effective_driver_total_ctns if driver_extra_ctns is not None else None
        
        cost_per_ctn = (effective_rate_cart_cost / effective_driver_total_ctns) if effective_driver_total_ctns > Decimal("0.0") else Decimal("0.0")
        allocated_cost = g["ctns"] * cost_per_ctn

        d_total_customers = Decimal(str(len(driver_customers.get((b, d), set()))))
        cost_per_drop_point = (effective_rate_cart_cost / d_total_customers) if d_total_customers > Decimal("0.0") else Decimal("0.0")

        i_key = (b, d, g["item_code"])
        if i_key not in item_report_dict:
            item_report_dict[i_key] = {
                "target_date": target_date, 
                "bu": g["bu"], "branch": b, "driver_name": d, "item_code": g["item_code"],
                "item_name": g["item_name"], "principal": g["principal"], "brand": g["brand"],
                "ctns": Decimal("0.0"), "weight": Decimal("0.0"), "volumetric_weight": Decimal("0.0"), "allocated_cost": Decimal("0.0"), "cost_per_carton": cost_per_ctn,
                "driver_total_ctns": d_total, "driver_total_weight": d_weight_total, "driver_total_volumetric_weight": d_volumetric_weight_total, "rate_cart_cost": b_cost, 
                "override_driver_total_ctns": store_override_ctns, "override_rate_cart_cost": store_override_cost, 
                "override_amount": driver_extra, "sales_amount": Decimal("0.0")
            }
        item_report_dict[i_key]["ctns"] += g["ctns"]
        item_report_dict[i_key]["weight"] += g["weight"]
        item_report_dict[i_key]["volumetric_weight"] += g["volumetric_weight"]
        item_report_dict[i_key]["allocated_cost"] += allocated_cost
        item_report_dict[i_key]["sales_amount"] += g["sales_amount"]

        t_key = (g["branch"], g["township"], g["customer_code"], g["driver_name"])
        if t_key not in township_report_dict:
            township_report_dict[t_key] = {
                "target_date": target_date, 
                "branch": b, "driver_name": g["driver_name"], "township": g["township"], 
                "customer_code": g["customer_code"], "contact_person": g["contact_person"], 
                "ctns": Decimal("0.0"), "weight": Decimal("0.0"), "volumetric_weight": Decimal("0.0"), "driver_total_ctns": d_total, "driver_total_weight": d_weight_total, "driver_total_volumetric_weight": d_volumetric_weight_total, "rate_cart_cost": b_cost,
                "override_driver_total_ctns": store_override_ctns, "override_rate_cart_cost": store_override_cost,
                "override_amount": driver_extra,
                "cost_per_carton": cost_per_ctn, "allocated_cost": Decimal("0.0"),
                "total_drop_points": d_total_customers, "cost_per_drop_point": cost_per_drop_point,
                "sales_amount": Decimal("0.0")
            }
        township_report_dict[t_key]["ctns"] += g["ctns"]
        township_report_dict[t_key]["weight"] += g["weight"]
        township_report_dict[t_key]["volumetric_weight"] += g["volumetric_weight"]
        township_report_dict[t_key]["allocated_cost"] += allocated_cost
        township_report_dict[t_key]["sales_amount"] += g["sales_amount"]

    item_report_list = list(item_report_dict.values())
    item_report_list.sort(key=lambda x: (x["branch"], x["driver_name"], x["item_code"]))

    township_report_list = list(township_report_dict.values())
    township_report_list.sort(key=lambda x: (x["branch"], x["driver_name"], x["township"], x["customer_code"]))

    return {
        "item_report": item_report_list,
        "township_report": township_report_list
    }

async def generate_and_save_daily_report(target_date: str, user_id: int = 30):
    try:
        data = await _get_daily_report_data(target_date)
        if not data["item_report"] and not data["township_report"]:
            return

        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Check if a report for this date already exists to track edit status
        await cursor.execute("SELECT TOP 1 created_at, created_by FROM Daily_Item_Report WHERE target_date = ?", (target_date,))
        existing_report = await cursor.fetchone()
        
        if existing_report:
            orig_created_at = existing_report[0] if existing_report[0] else now_str
            orig_created_by = existing_report[1] if existing_report[1] else user_id
            edited_at = now_str
            edited_by = user_id 
        else:
            orig_created_at = now_str
            orig_created_by = user_id
            edited_at = None
            edited_by = None

        await cursor.execute("DELETE FROM Daily_Item_Report WHERE target_date = ?", (target_date,))
        for it in data["item_report"]:
            await cursor.execute("""
                INSERT INTO Daily_Item_Report
                (target_date, bu, branch, driver_name, item_code, item_name, principal, brand,
                 ctns, weight, volumetric_weight, allocated_cost, cost_per_carton, driver_total_ctns, driver_total_weight, driver_total_volumetric_weight, rate_cart_cost, override_driver_total_ctns, override_rate_cart_cost, sales_amount, created_at, created_by, edited_at, edited_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                target_date, it.get("bu",""), it.get("branch",""), it.get("driver_name",""),
                it.get("item_code",""), it.get("item_name",""), it.get("principal",""), it.get("brand",""),
                it.get("ctns", Decimal("0.0")), it.get("weight", Decimal("0.0")), it.get("volumetric_weight", Decimal("0.0")), it.get("allocated_cost", Decimal("0.0")), it.get("cost_per_carton", Decimal("0.0")),
                it.get("driver_total_ctns", Decimal("0.0")), it.get("driver_total_weight", Decimal("0.0")), it.get("driver_total_volumetric_weight", Decimal("0.0")), it.get("rate_cart_cost", Decimal("0.0")), 
                it.get("override_driver_total_ctns"), it.get("override_rate_cart_cost"), 
                it.get("sales_amount", Decimal("0.0")), orig_created_at, orig_created_by, edited_at, edited_by
            ))

        await cursor.execute("DELETE FROM Daily_Township_Report WHERE target_date = ?", (target_date,))
        for tw in data["township_report"]:
            await cursor.execute("""
                INSERT INTO Daily_Township_Report
                (target_date, branch, driver_name, township, customer_code, contact_person,
                 ctns, weight, volumetric_weight, driver_total_ctns, driver_total_weight, driver_total_volumetric_weight, rate_cart_cost, override_driver_total_ctns, override_rate_cart_cost, cost_per_carton, allocated_cost,
                 total_drop_points, cost_per_drop_point, sales_amount, created_at, created_by, edited_at, edited_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                target_date, tw.get("branch",""), tw.get("driver_name",""), tw.get("township",""),
                tw.get("customer_code",""), tw.get("contact_person",""),
                tw.get("ctns", Decimal("0.0")), tw.get("weight", Decimal("0.0")), tw.get("volumetric_weight", Decimal("0.0")), tw.get("driver_total_ctns", Decimal("0.0")), tw.get("driver_total_weight", Decimal("0.0")), tw.get("driver_total_volumetric_weight", Decimal("0.0")), tw.get("rate_cart_cost", Decimal("0.0")),
                tw.get("override_driver_total_ctns"), tw.get("override_rate_cart_cost"),
                tw.get("cost_per_carton", Decimal("0.0")), tw.get("allocated_cost", Decimal("0.0")),
                tw.get("total_drop_points", Decimal("0.0")), tw.get("cost_per_drop_point", Decimal("0.0")), tw.get("sales_amount", Decimal("0.0")), orig_created_at, orig_created_by, edited_at, edited_by
            ))

        await conn.commit()
        await conn.close()
        logger.info(f"Successfully generated and saved report for {target_date}")
        return data
    except Exception as e:
        logger.error(f"Failed to generate and save daily report for {target_date}: {str(e)}")
        return None

async def daily_job_generator():
    target_date = datetime.datetime.now().strftime("%Y-%m-%d")
    logger.info(f"Running automated EOD report generation for {target_date}")
    await generate_and_save_daily_report(target_date)

async def get_or_generate_daily_report(target_date: str):
    conn = await get_logistic_connection()
    cursor = await conn.cursor()
    await cursor.execute("SELECT TOP 1 target_date FROM Daily_Item_Report WHERE target_date = ?", (target_date,))
    row = await cursor.fetchone()

    if row:
        await cursor.execute("""
            SELECT bu, branch, driver_name, item_code, item_name, principal, brand,
                   ctns, weight, volumetric_weight, allocated_cost, cost_per_carton, driver_total_ctns, driver_total_weight, driver_total_volumetric_weight, rate_cart_cost, override_driver_total_ctns, override_rate_cart_cost, sales_amount
            FROM Daily_Item_Report WHERE target_date = ?
        """, (target_date,))
        ir_cols = ["bu","branch","driver_name","item_code","item_name","principal","brand",
                   "ctns","weight","volumetric_weight","allocated_cost","cost_per_carton","driver_total_ctns","driver_total_weight","driver_total_volumetric_weight","rate_cart_cost", "override_driver_total_ctns", "override_rate_cart_cost", "sales_amount"]
        item_report = [dict(zip(ir_cols, r)) for r in await cursor.fetchall()]
        for it in item_report:
            it["target_date"] = target_date
            for k in ["ctns","weight","volumetric_weight","allocated_cost","cost_per_carton","driver_total_ctns","driver_total_weight","driver_total_volumetric_weight","rate_cart_cost", "sales_amount"]:
                it[k] = Decimal(str(it[k])) if it[k] is not None else Decimal("0.0")
            for k in ["override_driver_total_ctns", "override_rate_cart_cost"]:
                it[k] = Decimal(str(it[k])) if it[k] is not None else None

        await cursor.execute("""
            SELECT branch, driver_name, township, customer_code, contact_person,
                   ctns, weight, volumetric_weight, driver_total_ctns, driver_total_weight, driver_total_volumetric_weight, rate_cart_cost, override_driver_total_ctns, override_rate_cart_cost, cost_per_carton, allocated_cost,
                   total_drop_points, cost_per_drop_point, sales_amount
            FROM Daily_Township_Report WHERE target_date = ?
        """, (target_date,))
        tr_cols = ["branch","driver_name","township","customer_code","contact_person",
                   "ctns","weight","volumetric_weight","driver_total_ctns","driver_total_weight","driver_total_volumetric_weight","rate_cart_cost", "override_driver_total_ctns", "override_rate_cart_cost", "cost_per_carton","allocated_cost",
                   "total_drop_points","cost_per_drop_point","sales_amount"]
        township_report = [dict(zip(tr_cols, r)) for r in await cursor.fetchall()]
        for tw in township_report:
            tw["target_date"] = target_date
            for k in ["ctns","weight","volumetric_weight","driver_total_ctns","driver_total_weight","driver_total_volumetric_weight","rate_cart_cost", "cost_per_carton","allocated_cost","total_drop_points","cost_per_drop_point","sales_amount"]:
                tw[k] = Decimal(str(tw[k])) if tw[k] is not None else Decimal("0.0")
            for k in ["override_driver_total_ctns", "override_rate_cart_cost"]:
                tw[k] = Decimal(str(tw[k])) if tw[k] is not None else None

        await cursor.execute("SELECT driver_name, override_amount, override_ctns FROM Daily_Driver_Override WHERE target_date = ?", (target_date,))
        overrides = {
            r[0].strip(): {
                'amount': Decimal(str(r[1])) if r[1] is not None else None,
                'ctns': Decimal(str(r[2])) if r[2] is not None else None
            } for r in await cursor.fetchall()
        }
        for it in item_report:
            ov = overrides.get((it.get("driver_name") or "").strip(), {'amount': None, 'ctns': None})
            it["override_amount"] = ov['amount']
            it["override_ctns"] = ov['ctns']
        for tw in township_report:
            ov = overrides.get((tw.get("driver_name") or "").strip(), {'amount': None, 'ctns': None})
            tw["override_amount"] = ov['amount']
            tw["override_ctns"] = ov['ctns']

        await conn.close()
        return {"item_report": item_report, "township_report": township_report}
    else:
        await conn.close()
        return await _get_daily_report_data(target_date)

def _aggregate_reports(daily_datas: List[dict]):
    item_report_dict = {}
    township_report_dict = {}

    for data in daily_datas:
        for item in data.get("item_report", []):
            t_date = item.get("target_date", "")
            i_key = (t_date, item.get("branch", ""), item.get("driver_name", ""), item.get("item_code", ""))
            if i_key not in item_report_dict:
                item_report_dict[i_key] = {
                    "target_date": t_date,
                    "bu": item.get("bu", ""), "branch": item.get("branch", ""), "driver_name": item.get("driver_name", ""),
                    "item_code": item.get("item_code", ""), "item_name": item.get("item_name", ""), 
                    "principal": item.get("principal", ""), "brand": item.get("brand", ""),
                    "ctns": Decimal("0.0"), "allocated_cost": Decimal("0.0"), "driver_total_ctns": Decimal("0.0"), 
                    "override_driver_total_ctns": Decimal(str(item.get("override_driver_total_ctns"))) if item.get("override_driver_total_ctns") is not None else None, 
                    "rate_cart_cost": Decimal(str(item.get("rate_cart_cost", Decimal("0.0")))), 
                    "override_rate_cart_cost": Decimal(str(item.get("override_rate_cart_cost"))) if item.get("override_rate_cart_cost") is not None else None, 
                    "override_amount": Decimal(str(item.get("override_amount"))) if item.get("override_amount") is not None else None,
                    "override_ctns": Decimal(str(item.get("override_ctns"))) if item.get("override_ctns") is not None else None,
                    "sales_amount": Decimal("0.0")
                }
            else:
                item_report_dict[i_key]["rate_cart_cost"] = Decimal(str(item.get("rate_cart_cost", item_report_dict[i_key]["rate_cart_cost"])))
                item_report_dict[i_key]["override_rate_cart_cost"] = Decimal(str(item.get("override_rate_cart_cost"))) if item.get("override_rate_cart_cost") is not None else None
                item_report_dict[i_key]["override_amount"] = Decimal(str(item.get("override_amount"))) if item.get("override_amount") is not None else item_report_dict[i_key]["override_amount"]
                item_report_dict[i_key]["override_ctns"] = Decimal(str(item.get("override_ctns"))) if item.get("override_ctns") is not None else item_report_dict[i_key]["override_ctns"]

            item_report_dict[i_key]["ctns"] += Decimal(str(item.get("ctns", Decimal("0.0"))))
            item_report_dict[i_key]["allocated_cost"] += Decimal(str(item.get("allocated_cost", Decimal("0.0"))))
            item_report_dict[i_key]["sales_amount"] += Decimal(str(item.get("sales_amount", Decimal("0.0"))))
            item_report_dict[i_key]["driver_total_ctns"] += Decimal(str(item.get("driver_total_ctns", Decimal("0.0"))))

        for tw in data.get("township_report", []):
            t_date = tw.get("target_date", "")
            t_key = (t_date, tw.get("branch", ""), tw.get("township", ""), tw.get("customer_code", ""), tw.get("driver_name", ""))
            if t_key not in township_report_dict:
                township_report_dict[t_key] = {
                    "target_date": t_date,
                    "branch": tw.get("branch", ""), "driver_name": tw.get("driver_name", ""), 
                    "township": tw.get("township", ""), "customer_code": tw.get("customer_code", ""), 
                    "contact_person": tw.get("contact_person", ""), "ctns": Decimal("0.0"), "allocated_cost": Decimal("0.0"), 
                    "driver_total_ctns": Decimal("0.0"), 
                    "override_driver_total_ctns": Decimal(str(tw.get("override_driver_total_ctns"))) if tw.get("override_driver_total_ctns") is not None else None, 
                    "rate_cart_cost": Decimal(str(tw.get("rate_cart_cost", Decimal("0.0")))), 
                    "override_rate_cart_cost": Decimal(str(tw.get("override_rate_cart_cost"))) if tw.get("override_rate_cart_cost") is not None else None, 
                    "override_amount": Decimal(str(tw.get("override_amount"))) if tw.get("override_amount") is not None else None,
                    "override_ctns": Decimal(str(tw.get("override_ctns"))) if tw.get("override_ctns") is not None else None,
                    "total_drop_points": Decimal("0.0"), "sales_amount": Decimal("0.0")
                }
            else:
                township_report_dict[t_key]["rate_cart_cost"] = Decimal(str(tw.get("rate_cart_cost", township_report_dict[t_key]["rate_cart_cost"])))
                township_report_dict[t_key]["override_rate_cart_cost"] = Decimal(str(tw.get("override_rate_cart_cost"))) if tw.get("override_rate_cart_cost") is not None else None
                township_report_dict[t_key]["override_amount"] = Decimal(str(tw.get("override_amount"))) if tw.get("override_amount") is not None else township_report_dict[t_key]["override_amount"]
                township_report_dict[t_key]["override_ctns"] = Decimal(str(tw.get("override_ctns"))) if tw.get("override_ctns") is not None else township_report_dict[t_key]["override_ctns"]

            township_report_dict[t_key]["ctns"] += Decimal(str(tw.get("ctns", Decimal("0.0"))))
            township_report_dict[t_key]["allocated_cost"] += Decimal(str(tw.get("allocated_cost", Decimal("0.0"))))
            township_report_dict[t_key]["sales_amount"] += Decimal(str(tw.get("sales_amount", Decimal("0.0"))))
            township_report_dict[t_key]["driver_total_ctns"] += Decimal(str(tw.get("driver_total_ctns", Decimal("0.0"))))
            township_report_dict[t_key]["total_drop_points"] += Decimal(str(tw.get("total_drop_points", Decimal("0.0"))))

    item_report_list = list(item_report_dict.values())
    for item in item_report_list:
        item["cost_per_carton"] = item["allocated_cost"] / item["ctns"] if item["ctns"] > Decimal("0.0") else Decimal("0.0")
    item_report_list.sort(key=lambda x: (x.get("target_date", ""), x["branch"], x["driver_name"], x["item_code"]))

    township_report_list = list(township_report_dict.values())
    for tw in township_report_list:
        tw["cost_per_carton"] = tw["allocated_cost"] / tw["ctns"] if tw["ctns"] > Decimal("0.0") else Decimal("0.0")
        tw["cost_per_drop_point"] = tw["rate_cart_cost"] / tw["total_drop_points"] if tw["total_drop_points"] > Decimal("0.0") else Decimal("0.0")
    township_report_list.sort(key=lambda x: (x.get("target_date", ""), x["branch"], x["driver_name"], x["township"], x["customer_code"]))

    return {
        "item_report": item_report_list,
        "township_report": township_report_list
    }

@app.post("/account/daily-override")
async def save_daily_override(data: DriverOverrideData, user: dict = Depends(require_permission("view_daily_report"))):
    try:
        if data.override_amount is None and data.override_ctns is None:
            raise HTTPException(status_code=400, detail="Either Override Amount or Override Ctns must be provided.")

        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        await cursor.execute("SELECT override_amount, override_ctns FROM Daily_Driver_Override WHERE target_date = ? AND driver_name = ?", (data.target_date, data.driver_name))
        existing = await cursor.fetchone()
        
        username = user['username']
        user_id = user['id']
        change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if existing:
            old_amount = existing[0]
            old_ctns = existing[1]
            
            old_cost_str = str(old_amount) if old_amount is not None else ""
            new_cost_str = str(data.override_amount) if data.override_amount is not None else ""
            
            if old_cost_str != new_cost_str:
                await cursor.execute("""
                    INSERT INTO Driver_Override_Change_Log (target_date, driver_name, changed_by, change_date, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (data.target_date, data.driver_name, user_id, change_date, 'Override Amount', old_cost_str, new_cost_str))

            old_ctns_str = str(old_ctns) if old_ctns is not None else ""
            new_ctns_str = str(data.override_ctns) if data.override_ctns is not None else ""

            if old_ctns_str != new_ctns_str:
                await cursor.execute("""
                    INSERT INTO Driver_Override_Change_Log (target_date, driver_name, changed_by, change_date, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (data.target_date, data.driver_name, user_id, change_date, 'Override Ctns', old_ctns_str, new_ctns_str))
        else:
            new_cost_str = str(data.override_amount) if data.override_amount is not None else ""
            await cursor.execute("""
                INSERT INTO Driver_Override_Change_Log (target_date, driver_name, changed_by, change_date, field_name, old_value, new_value)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (data.target_date, data.driver_name, user_id, change_date, 'Record Created', None, new_cost_str))

        await cursor.execute("""
            MERGE Daily_Driver_Override AS target
            USING (SELECT ? AS target_date, ? AS driver_name, ? AS override_amount, ? AS override_ctns, ? AS edited_at, ? AS edited_by, ? AS created_at, ? AS created_by) AS source
            ON target.target_date = source.target_date AND target.driver_name = source.driver_name
            WHEN MATCHED THEN UPDATE SET override_amount = source.override_amount, override_ctns = source.override_ctns, edited_at = source.edited_at, edited_by = source.edited_by
            WHEN NOT MATCHED THEN INSERT (target_date, driver_name, override_amount, override_ctns, created_at, created_by) VALUES (source.target_date, source.driver_name, source.override_amount, source.override_ctns, source.created_at, source.created_by);
        """, (data.target_date, data.driver_name, data.override_amount, data.override_ctns, change_date, user_id, change_date, user_id))
        
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user_id, "ADD_EDIT_DRIVER_OVERRIDE", f"Set override {data.override_amount} for {data.driver_name} on {data.target_date}")
        
        # Trigger an automatic recalculation for this date
        await generate_and_save_daily_report(data.target_date, user_id=user['id'])
        
        return {"message": "Override saved and report recalculated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving override: {str(e)}")
        
@app.get("/account/daily-override/drivers")
async def get_drivers_for_date(target_date: str, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT DISTINCT driver_name FROM Daily_Item_Report WHERE target_date = ?", (target_date,))
        drivers = [row[0] for row in await cursor.fetchall() if row[0]]
        await conn.close()
        return {"drivers": drivers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching drivers: {str(e)}")

class OverrideLogItem(BaseModel):
    id: int
    target_date: str
    driver_name: str
    changed_by: str
    change_date: str
    field_name: str
    old_value: Optional[str]
    new_value: Optional[str]

@app.get("/account/daily-override/logs", response_model=List[OverrideLogItem])
async def get_daily_override_logs(
    target_date: str = Query(...), 
    driver_name: str = Query(...), 
    user: dict = Depends(get_current_user)
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.target_date, l.driver_name, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM Driver_Override_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.target_date = ? AND l.driver_name = ? 
            ORDER BY l.change_date DESC
        """, (target_date, driver_name))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "target_date": r[1], "driver_name": r[2], "changed_by": r[3], "change_date": r[4], "field_name": r[5], "old_value": r[6], "new_value": r[7]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Error fetching override logs: {str(e)}")

@app.get("/account/daily-override/list")
async def get_daily_overrides(target_date: str, user: dict = Depends(require_permission("view_daily_report"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, target_date, driver_name, override_amount, override_ctns FROM Daily_Driver_Override WHERE target_date = ?", (target_date,))
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": r[0], "target_date": r[1], "driver_name": r[2], "override_amount": r[3], "override_ctns": r[4]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching overrides: {str(e)}")

@app.delete("/account/daily-override/{override_id}")
async def delete_daily_override(override_id: int, target_date: str, user: dict = Depends(require_permission("view_daily_report"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        await cursor.execute("SELECT driver_name, override_amount FROM Daily_Driver_Override WHERE id = ?", (override_id,))
        record = await cursor.fetchone()
        if not record:
            await conn.close()
            raise HTTPException(status_code=404, detail="Override not found")
            
        driver_name = record[0]
        amount = record[1]
        username = user['username']
        change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        await cursor.execute("""
            INSERT INTO Driver_Override_Change_Log (target_date, driver_name, changed_by, change_date, field_name, old_value, new_value)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (target_date, driver_name, user['id'], change_date, 'Record Deleted', str(amount), None))

        await cursor.execute("DELETE FROM Daily_Driver_Override WHERE id = ?", (override_id,))
            
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "DELETE_DRIVER_OVERRIDE", f"Deleted override ID {override_id}")
        await generate_and_save_daily_report(target_date, user_id=user['id'])
        
        return {"message": "Override deleted and report recalculated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting override: {str(e)}")

@app.get("/account/daily-rate-cut-report")
async def get_daily_rate_cart_report(
    target_date: Optional[str] = None, 
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_permission("view_daily_report"))
):
    try:
        if start_date and end_date:
            try:
                start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
            
            if start_dt > end_dt:
                raise HTTPException(status_code=400, detail="start_date cannot be after end_date")

            daily_datas = []
            current_dt = start_dt
            while current_dt <= end_dt:
                dt_str = current_dt.strftime("%Y-%m-%d")
                daily_datas.append(await get_or_generate_daily_report(dt_str))
                current_dt += datetime.timedelta(days=1)
            
            aggregated = _aggregate_reports(daily_datas)
            return {
                "target_date": f"{start_date} to {end_date}",
                "report": aggregated["item_report"],
                "township_report": aggregated["township_report"]
            }
            
        else:
            if not target_date:
                target_date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                
            data = await get_or_generate_daily_report(target_date)
            return {
                "target_date": target_date, 
                "report": data["item_report"], 
                "township_report": data["township_report"]
            }
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing daily report: {str(e)}")

# --- Reference Management Endpoints ---

@app.get("/references/locations")
async def get_ref_locations():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, name FROM Locations ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "name": row[1]} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/locations")
async def add_ref_location(item: ReferenceItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        try:
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await cursor.execute("INSERT INTO Locations (name, created_at, created_by) VALUES (?, ?, ?)", (item.name, now_str, user['id']))
            await conn.commit()
        except Exception as e:
            if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
                await conn.close()
                raise HTTPException(status_code=400, detail="Location already exists")
            raise
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Added location: {item.name}")
        return {"message": "Added successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/locations/{location_id}")
async def delete_ref_location(location_id: int, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM Locations WHERE id = ?", (location_id,))
        row = await cursor.fetchone()
        if not row:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await cursor.execute("DELETE FROM Locations WHERE id = ?", (location_id,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted location: {row[0]}")
        return {"message": "Deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/references/rate-cart-locations")
async def get_ref_rate_cart_locations():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, name FROM Rate_Cart_Locations ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "name": row[1]} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/rate-cart-locations")
async def add_ref_rate_cart_location(item: ReferenceItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        try:
            await cursor.execute("INSERT INTO Rate_Cart_Locations (name) VALUES (?)", (item.name,))
            await conn.commit()
        except Exception as e:
            if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
                await conn.close()
                raise HTTPException(status_code=400, detail="Rate Cart Location already exists")
            raise
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Added rate cart location: {item.name}")
        return {"message": "Added successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/rate-cart-locations/{location_id}")
async def delete_ref_rate_cart_location(location_id: int, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM Rate_Cart_Locations WHERE id = ?", (location_id,))
        row = await cursor.fetchone()
        if not row:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await cursor.execute("DELETE FROM Rate_Cart_Locations WHERE id = ?", (location_id,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted rate cart location: {row[0]}")
        return {"message": "Deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/references/uoms")
async def get_ref_uoms():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, name FROM UOMs ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "name": row[1]} for row in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/uoms")
async def add_ref_uom(item: ReferenceItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        try:
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await cursor.execute("INSERT INTO UOMs (name, created_at, created_by) VALUES (?, ?, ?)", (item.name, now_str, user['id']))
            await conn.commit()
        except Exception as e:
            if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
                await conn.close()
                raise HTTPException(status_code=400, detail="UOM already exists")
            raise
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Added UOM: {item.name}")
        return {"message": "Added successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/uoms/{uom_id}")
async def delete_ref_uom(uom_id: int, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM UOMs WHERE id = ?", (uom_id,))
        row = await cursor.fetchone()
        if not row:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await cursor.execute("DELETE FROM UOMs WHERE id = ?", (uom_id,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted UOM: {row[0]}")
        return {"message": "Deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/references/channels")
async def get_ref_channels():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, name FROM Channels ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "name": row[1]} for row in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/channels")
async def add_ref_channel(item: ReferenceItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        try:
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await cursor.execute("INSERT INTO Channels (name, created_at, created_by) VALUES (?, ?, ?)", (item.name, now_str, user['id']))
            await conn.commit()
        except Exception as e:
            if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
                await conn.close()
                raise HTTPException(status_code=400, detail="Channel already exists")
            raise
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Added channel: {item.name}")
        return {"message": "Added successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/channels/{channel_id}")
async def delete_ref_channel(channel_id: int, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM Channels WHERE id = ?", (channel_id,))
        row = await cursor.fetchone()
        if not row:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await cursor.execute("DELETE FROM Channels WHERE id = ?", (channel_id,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted channel: {row[0]}")
        return {"message": "Deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/references/departments")
async def get_ref_departments():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, name FROM Departments ORDER BY name")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"id": row[0], "name": row[1]} for row in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/departments")
async def add_ref_department(item: ReferenceItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        try:
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            await cursor.execute("INSERT INTO Departments (name, created_at, created_by) VALUES (?, ?, ?)", (item.name, now_str, user['id']))
            await conn.commit()
        except Exception as e:
            if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
                await conn.close()
                raise HTTPException(status_code=400, detail="Department already exists")
            raise
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Added department: {item.name}")
        return {"message": "Added successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/departments/{department_id}")
async def delete_ref_department(department_id: int, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT name FROM Departments WHERE id = ?", (department_id,))
        row = await cursor.fetchone()
        if not row:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await cursor.execute("DELETE FROM Departments WHERE id = ?", (department_id,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted department: {row[0]}")
        return {"message": "Deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/references/locations")
async def edit_ref_location(item: ReferenceEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Locations SET name = ?, edited_at = ?, edited_by = ? WHERE id = ?", (item.new_name, now_str, user['id'], item.id))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited location (id {item.id}) to {item.new_name}")
        return {"message": "Edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="Location already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/references/rate-cart-locations")
async def edit_ref_rate_cart_location(item: ReferenceEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Rate_Cart_Locations SET name = ?, edited_at = ?, edited_by = ? WHERE id = ?", (item.new_name, now_str, user['id'], item.id))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited rate cart location (id {item.id}) to {item.new_name}")
        return {"message": "Edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="Rate Cart Location already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/references/uoms")
async def edit_ref_uom(item: ReferenceEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE UOMs SET name = ?, edited_at = ?, edited_by = ? WHERE id = ?", (item.new_name, now_str, user['id'], item.id))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited UOM (id {item.id}) to {item.new_name}")
        return {"message": "Edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="UOM already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/references/channels")
async def edit_ref_channel(item: ReferenceEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Channels SET name = ?, edited_at = ?, edited_by = ? WHERE id = ?", (item.new_name, now_str, user['id'], item.id))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited channel (id {item.id}) to {item.new_name}")
        return {"message": "Edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="Channel already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/references/departments")
async def edit_ref_department(item: ReferenceEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Departments SET name = ?, edited_at = ?, edited_by = ? WHERE id = ?", (item.new_name, now_str, user['id'], item.id))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited department (id {item.id}) to {item.new_name}")
        return {"message": "Edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="Department already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/references/settings/volumetric-divisor")
async def get_volumetric_divisor():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT setting_value FROM System_Settings WHERE setting_key = 'volumetric_divisor'")
        row = await cursor.fetchone()
        await conn.close()
        return {"value": row[0] if row else "5000"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/references/settings/volumetric-divisor")
async def update_volumetric_divisor(item: SettingUpdate, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE System_Settings SET setting_value = ?, updated_at = ?, updated_by = ? WHERE setting_key = 'volumetric_divisor'", (item.value, now_str, user['id']))
        if cursor.rowcount == 0:
            await cursor.execute("INSERT INTO System_Settings (setting_key, setting_value, updated_at, updated_by) VALUES (?, ?, ?, ?)", ('volumetric_divisor', item.value, now_str, user['id']))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Updated volumetric divisor to {item.value}")
        return {"message": "Updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
async def get_ref_location_mappings():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT to_location, branch_code FROM Location_Mapping ORDER BY to_location")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"to_location": row[0], "branch_code": row[1]} for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/references/location-mappings")
async def add_ref_location_mapping(item: LocationMappingItem, user: dict = Depends(require_permission("add_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("""
            MERGE Location_Mapping AS target
            USING (SELECT ? AS to_location, ? AS branch_code, ? AS created_at, ? AS created_by) AS source
            ON target.to_location = source.to_location
            WHEN MATCHED THEN UPDATE SET branch_code = source.branch_code
            WHEN NOT MATCHED THEN INSERT (to_location, branch_code, created_at, created_by) VALUES (source.to_location, source.branch_code, source.created_at, source.created_by);
        """, (item.to_location, item.branch_code, now_str, user['id']))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "ADD_REFERENCE", f"Mapped {item.to_location} to {item.branch_code}")
        return {"message": "Mapping saved successfully"}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
    
@app.put("/references/location-mappings")
async def edit_ref_location_mapping(item: LocationMappingEditItem, user: dict = Depends(require_permission("edit_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("""
            UPDATE Location_Mapping 
            SET to_location = ?, branch_code = ?, edited_at = ?, edited_by = ? 
            WHERE to_location = ?
        """, (item.new_to_location, item.new_branch_code, now_str, user['id'], item.original_to_location))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Mapping not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "EDIT_REFERENCE", f"Edited mapping from {item.original_to_location} to {item.new_to_location} ({item.new_branch_code})")
        return {"message": "Mapping edited successfully"}
    except Exception as e:
        if "UNIQUE" in str(e).upper() or "duplicate" in str(e).lower() or "Violation" in str(e):
            raise HTTPException(status_code=400, detail="Mapping for this location already exists")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/references/location-mappings/{to_location}")
async def delete_ref_location_mapping(to_location: str, user: dict = Depends(require_permission("delete_reference"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("DELETE FROM Location_Mapping WHERE to_location = ?", (to_location,))
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_REFERENCE", f"Deleted mapping for: {to_location}")
        return {"message": "Deleted successfully"}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

# --- Branch Code Management Endpoints ---

@app.get("/account/branch-codes")
async def get_branch_codes(user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT log_pric, code, name, dept, principal, description FROM Branch_Code")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"log_pric": r[0], "code": r[1], "name": r[2], "dept": r[3], "principal": r[4], "description": r[5]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading branch codes: {str(e)}")

@app.post("/account/branch-codes")
async def save_branch_code(data: BranchCodeData, user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        if data.original_log_pric:
            if "edit_branch_code" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'edit_branch_code' permission")
            
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']
            
            await cursor.execute("SELECT log_pric, code, name, dept, principal, description FROM Branch_Code WHERE log_pric = ?", (data.original_log_pric,))
            existing = await cursor.fetchone()
            if existing:
                changes = []
                fields = [('log_pric', data.log_pric), ('code', data.code), ('name', data.name), ('dept', data.dept), ('principal', data.principal), ('description', data.description)]
                for idx, (field_name, new_val) in enumerate(fields):
                    old_str = str(existing[idx]).strip() if existing[idx] else ""
                    new_str = str(new_val).strip() if new_val else ""
                    if old_str != new_str:
                        changes.append((data.log_pric, username, now_str, field_name, old_str, new_str))
                
                if data.original_log_pric != data.log_pric:
                    await cursor.execute("UPDATE Branch_Code_Change_Log SET log_pric = ? WHERE log_pric = ?", (data.log_pric, data.original_log_pric))
                
                if changes:
                    await cursor.executemany("INSERT INTO Branch_Code_Change_Log (log_pric, changed_by, change_date, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)", changes)

            await cursor.execute("""
                UPDATE Branch_Code 
                SET log_pric = ?, code = ?, name = ?, dept = ?, principal = ?, description = ?, edited_at = ?, edited_by = ?
                WHERE log_pric = ?
            """, (data.log_pric, data.code, data.name, data.dept, data.principal, data.description, now_str, username, data.original_log_pric))
            await log_user_activity(user['id'], "UPDATE_BRANCH_CODE", f"Updated Branch Code: {data.log_pric}")
        else:
            if "add_branch_code" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'add_branch_code' permission")
            
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']

            await cursor.execute("""
                INSERT INTO Branch_Code (log_pric, code, name, dept, principal, description, created_at, created_by) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (data.log_pric, data.code, data.name, data.dept, data.principal, data.description, now_str, username))
            await log_user_activity(user['id'], "ADD_BRANCH_CODE", f"Added Branch Code: {data.log_pric}")

        await conn.commit()
        await conn.close()
        return {"message": "Branch code saved successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving branch code: {str(e)}")

@app.get("/account/branch-codes/{log_pric}/logs", response_model=List[BranchCodeLogItem])
async def get_branch_code_logs(log_pric: str, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.log_pric, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM Branch_Code_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.log_pric = ? ORDER BY l.change_date DESC
        """, (log_pric,))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "log_pric": r[1], "changed_by": r[2], "change_date": r[3], "field_name": r[4], "old_value": r[5], "new_value": r[6]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching logs: {str(e)}")

@app.delete("/account/branch-codes/{log_pric}")
async def delete_branch_code(log_pric: str, user: dict = Depends(require_permission("delete_branch_code"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("DELETE FROM Branch_Code_Change_Log WHERE log_pric = ?", (log_pric,))
        await cursor.execute("DELETE FROM Branch_Code WHERE log_pric = ?", (log_pric,))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="Branch code not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_BRANCH_CODE", f"Deleted Branch Code: {log_pric}")
        return {"message": "Branch code deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting branch code: {str(e)}")

# --- SD Code Management Endpoints ---

@app.get("/account/sd-codes")
async def get_sd_codes(user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT channel, code, name, dept, principal, log_pric FROM SD_Code")
        rows = await cursor.fetchall()
        await conn.close()
        return [{"channel": r[0], "code": r[1], "name": r[2], "dept": r[3], "principal": r[4], "log_pric": r[5]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading SD codes: {str(e)}")

@app.post("/account/sd-codes")
async def save_sd_code(data: SDCodeData, user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        if data.original_log_pric:
            if "edit_sd_code" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'edit_sd_code' permission")
            
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']

            await cursor.execute("SELECT log_pric, channel, code, name, dept, principal FROM SD_Code WHERE log_pric = ?", (data.original_log_pric,))
            existing = await cursor.fetchone()
            if existing:
                changes = []
                fields = [('log_pric', data.log_pric), ('channel', data.channel), ('code', data.code), ('name', data.name), ('dept', data.dept), ('principal', data.principal)]
                for idx, (field_name, new_val) in enumerate(fields):
                    old_str = str(existing[idx]).strip() if existing[idx] else ""
                    new_str = str(new_val).strip() if new_val else ""
                    if old_str != new_str:
                        changes.append((data.log_pric, username, now_str, field_name, old_str, new_str))
                
                if data.original_log_pric != data.log_pric:
                    await cursor.execute("UPDATE SD_Code_Change_Log SET log_pric = ? WHERE log_pric = ?", (data.log_pric, data.original_log_pric))
                
                if changes:
                    await cursor.executemany("INSERT INTO SD_Code_Change_Log (log_pric, changed_by, change_date, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)", changes)

            await cursor.execute("""
                UPDATE SD_Code 
                SET channel = ?, code = ?, name = ?, dept = ?, principal = ?, log_pric = ?, edited_at = ?, edited_by = ?
                WHERE log_pric = ?
            """, (data.channel, data.code, data.name, data.dept, data.principal, data.log_pric, now_str, username, data.original_log_pric))
            await log_user_activity(user['id'], "UPDATE_SD_CODE", f"Updated SD Code: {data.log_pric}")
        else:
            if "add_sd_code" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'add_sd_code' permission")
            
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']

            await cursor.execute("""
                INSERT INTO SD_Code (channel, code, name, dept, principal, log_pric, created_at, created_by) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (data.channel, data.code, data.name, data.dept, data.principal, data.log_pric, now_str, username))
            await log_user_activity(user['id'], "ADD_SD_CODE", f"Added SD Code: {data.log_pric}")

        await conn.commit()
        await conn.close()
        return {"message": "SD code saved successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving SD code: {str(e)}")

@app.get("/account/sd-codes/{log_pric}/logs", response_model=List[SDCodeLogItem])
async def get_sd_code_logs(log_pric: str, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.log_pric, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM SD_Code_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.log_pric = ? ORDER BY l.change_date DESC
        """, (log_pric,))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "log_pric": r[1], "changed_by": r[2], "change_date": r[3], "field_name": r[4], "old_value": r[5], "new_value": r[6]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching logs: {str(e)}")

@app.delete("/account/sd-codes/{log_pric}")
async def delete_sd_code(log_pric: str, user: dict = Depends(require_permission("delete_sd_code"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("DELETE FROM SD_Code_Change_Log WHERE log_pric = ?", (log_pric,))
        await cursor.execute("DELETE FROM SD_Code WHERE log_pric = ?", (log_pric,))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="SD code not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_SD_CODE", f"Deleted SD Code: {log_pric}")
        return {"message": "SD code deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting SD code: {str(e)}")

# --- Calculation History Endpoints ---

@app.post("/history/save")
async def save_calculation(data: CalculationSaveRequest, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        doc_nums_json = json.dumps(data.doc_nums)
        created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        await cursor.execute("SELECT log_pric, code, name FROM SD_Code")
        sd_code_db_map = {str(row[0]).strip().lower(): {"code": row[1], "name": row[2]} for row in await cursor.fetchall() if row[0]}
        
        channel_name = data.channel or ""
        is_sd_channel = channel_name in ["SD", "Telecom SD"]
        is_branch_channel = channel_name in ["Branch", "Outlet", "Telecom Branch"]

        async def _upsert_products(calc_id, products):
            await cursor.execute("DELETE FROM Calculation_Products WHERE calc_id = ?", (calc_id,))
            for p in products:
                db_ctns = p.get("original_ctns") if p.get("original_ctns") is not None else p.get("ctns", 0)
                db_weight = p.get("original_weight") if p.get("original_weight") is not None else p.get("weight", 0)
                
                db_edited_ctns = p.get("ctns") if p.get("original_ctns") is not None else None
                db_edited_weight = p.get("weight") if p.get("original_weight") is not None else None

                b_code_val = p.get("b_code", "")
                b_name_val = p.get("b_name", "")
                b_desc_val = p.get("b_desc", "")
                b_dept_val = p.get("b_dept", "")
                b_principal_val = p.get("b_principal", "")
                s_dept_val = p.get("s_dept", "")
                s_principal_val = p.get("s_principal", "")

                if is_sd_channel:
                    principal_val = str(p.get("principal", "")).strip().lower()
                    sd_info = sd_code_db_map.get(principal_val, {})
                    b_code_val = sd_info.get("code", b_code_val)
                    b_name_val = sd_info.get("name", b_name_val)
                    b_desc_val = sd_info.get("name", b_desc_val)
                    b_dept_val = None
                    b_principal_val = None
                elif is_branch_channel:
                    s_dept_val = None
                    s_principal_val = None

                await cursor.execute("""
                    INSERT INTO Calculation_Products
                    (calc_id, code, name, weight, doc_date, sin_no, principal, brand, ctns, bu,
                     b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal,
                     calculation_type, system_rate, unit_cost, total_cost, standard_unit_cost,
                     FromWhsCode, ToWhsCode, edited_ctns, edited_weight)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    calc_id,
                    p.get("code", ""), p.get("name", ""),
                    Decimal(str(db_weight)), p.get("doc_date", ""), p.get("sin_no", ""),
                    p.get("principal", ""), p.get("brand", ""), Decimal(str(db_ctns)),
                    p.get("bu", ""), b_code_val, b_name_val,
                    b_dept_val, b_principal_val, b_desc_val,
                    s_dept_val, s_principal_val,
                    p.get("calculation_type", ""), p.get("system_rate"),
                    Decimal(str(p.get("unit_cost", 0))), Decimal(str(p.get("total_cost", 0))),
                    p.get("standard_unit_cost"),
                    p.get("FromWhsCode", ""), p.get("ToWhsCode", ""),
                    Decimal(str(db_edited_ctns)) if db_edited_ctns is not None else None, 
                    Decimal(str(db_edited_weight)) if db_edited_weight is not None else None
                ))

        async def _upsert_posm_products(calc_id, posm_products):
            await cursor.execute("DELETE FROM Calculation_POSM_Products WHERE calc_id = ?", (calc_id,))
            for p in posm_products:
                b_code_val = p.get("b_code", "")
                b_name_val = p.get("b_name", "")
                b_desc_val = p.get("b_desc", "")
                b_dept_val = p.get("b_dept", "")
                b_principal_val = p.get("b_principal", "")
                s_dept_val = p.get("s_dept", "")
                s_principal_val = p.get("s_principal", "")

                if is_sd_channel:
                    sd_info = sd_code_db_map.get("posm", {})
                    b_code_val = sd_info.get("code", b_code_val)
                    b_name_val = sd_info.get("name", b_name_val)
                    b_desc_val = sd_info.get("name", b_desc_val)
                    b_dept_val = None
                    b_principal_val = None
                elif is_branch_channel:
                    s_dept_val = None
                    s_principal_val = None

                await cursor.execute("""
                    INSERT INTO Calculation_POSM_Products
                    (calc_id, department, item_name, uom, quantity, unit_cost, total_cost,
                     b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    calc_id,
                    p.get("department", ""), p.get("item_name", ""), p.get("uom", ""),
                    Decimal(str(p.get("quantity", 0))), Decimal(str(p.get("unit_cost", 0))),
                    Decimal(str(p.get("total_cost", 0))),
                    b_code_val, b_name_val, b_dept_val,
                    b_principal_val, b_desc_val,
                    s_dept_val, s_principal_val
                ))

        if data.id:
            await cursor.execute("SELECT id FROM Calculation_History WHERE id = ?", (data.id,))
            if not await cursor.fetchone():
                await conn.close()
                raise HTTPException(status_code=404, detail="Record to update not found")
            await cursor.execute("""
                    UPDATE Calculation_History 
                    SET created_at = ?, gate_name = ?, from_loc = ?, to_loc = ?, doc_nums = ?, total_weight = ?, manual_total_cost = ?, additional_charges = ?, final_total_cost = ?, channel = ?, status = ?, gate_cost = ?, gate_uom = ?, gate_unit = ?, posm_total_cost = ?
                    WHERE id = ?
                """, (
                    created_at, data.gate_name, data.from_loc, data.to_loc, doc_nums_json, data.total_weight, data.manual_total_cost, data.additional_charges, data.final_total_cost, data.channel, data.status, data.gate_cost, data.gate_uom, data.gate_unit, data.posm_total_cost, data.id
                ))
            await _upsert_products(data.id, data.calculated_products)
            await _upsert_posm_products(data.id, data.posm_products)
            message = "Calculation updated successfully"
            await log_user_activity(user['id'], "UPDATE_CALCULATION", f"Updated saved calculation ID: {data.id}")
        else:
            while True:
                new_id = int(datetime.datetime.now().strftime("%y%m%d%H%M%S"))
                await cursor.execute("SELECT 1 FROM Calculation_History WHERE id = ?", (new_id,))
                if not await cursor.fetchone(): break
                await asyncio.sleep(1)

            await cursor.execute("""
                    INSERT INTO Calculation_History ([id], [created_at], [gate_name], [from_loc], [to_loc], [doc_nums], [total_weight], [manual_total_cost], [additional_charges], [final_total_cost], [channel], [status], [created_by], [gate_cost], [gate_uom], [gate_unit], [posm_total_cost])
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    new_id, created_at, data.gate_name, data.from_loc, data.to_loc, doc_nums_json, data.total_weight, data.manual_total_cost, data.additional_charges, data.final_total_cost, data.channel, data.status, user["id"], data.gate_cost, data.gate_uom, data.gate_unit, data.posm_total_cost
                ))
            await _upsert_products(new_id, data.calculated_products)
            await _upsert_posm_products(new_id, data.posm_products)
            message = "Calculation saved successfully"
            await log_user_activity(user['id'], "SAVE_CALCULATION", f"Saved new calculation ID: {new_id}")

        await conn.commit()
        await conn.close()
        return {"message": message}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error saving history: {str(e)}")
    

@app.put("/history/{record_id}/submit")
async def submit_history_item(record_id: int, user: dict = Depends(require_permission("submit_calculation"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        # --- NEW VALIDATION: Check for exceeding cartons/weight ---
        await cursor.execute("""
            SELECT code, name, 
                   COALESCE(edited_ctns, ctns) as used_ctns, 
                   COALESCE(edited_weight, weight) as used_weight,
                   ctns as original_ctns,
                   weight as original_weight,
                   sin_no
            FROM Calculation_Products
            WHERE calc_id = ?
        """, (record_id,))
        current_products = await cursor.fetchall()
        
        for prod in current_products:
            code, name, used_ctns, used_weight, original_ctns, original_weight, sin_no = prod
            
            # Sum used ctns and weight for this item across all OTHER submitted/claimed calculations
            await cursor.execute("""
                SELECT SUM(COALESCE(cp.edited_ctns, cp.ctns)), SUM(COALESCE(cp.edited_weight, cp.weight))
                FROM Calculation_Products cp
                JOIN Calculation_History ch ON cp.calc_id = ch.id
                WHERE ch.status IN ('submitted', 'claimed')
                AND ch.id != ?
                AND cp.code = ?
                AND cp.sin_no = ?
            """, (record_id, code, sin_no))
            
            sums = await cursor.fetchone()
            prev_ctns = Decimal(str(sums[0])) if sums and sums[0] is not None else Decimal("0.0")
            prev_weight = Decimal(str(sums[1])) if sums and sums[1] is not None else Decimal("0.0")
            
            curr_ctns = Decimal(str(used_ctns)) if used_ctns is not None else Decimal("0.0")
            curr_weight = Decimal(str(used_weight)) if used_weight is not None else Decimal("0.0")
            
            orig_ctns = Decimal(str(original_ctns)) if original_ctns is not None else Decimal("0.0")
            orig_weight = Decimal(str(original_weight)) if original_weight is not None else Decimal("0.0")
            
            # Allow a tiny margin for float precision errors
            if prev_ctns + curr_ctns > orig_ctns + Decimal("0.01"):
                await conn.close()
                raise HTTPException(
                    status_code=400, 
                    detail=f"Cannot submit: Item '{code}' in Doc Num '{sin_no}' exceeds original cartons. Used: {prev_ctns + curr_ctns}, Allowed: {orig_ctns}."
                )
                
            if prev_weight + curr_weight > orig_weight + Decimal("0.01"):
                await conn.close()
                raise HTTPException(
                    status_code=400, 
                    detail=f"Cannot submit: Item '{code}' in Doc Num '{sin_no}' exceeds original weight. Used: {prev_weight + curr_weight}, Allowed: {orig_weight}."
                )
        # ------------------------------------------------------

        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Calculation_History SET status = 'submitted', submitted_by = ?, submitted_at = ? WHERE id = ?", (user["id"], now_str, record_id))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="Record not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "SUBMIT_CALCULATION", f"Submitted calculation ID: {record_id}")
        return {"message": "Calculation submitted successfully"}
    except HTTPException: 
        raise # Ensures the 400 error passes cleanly to the frontend
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Error submitting record: {str(e)}")

@app.put("/history/{record_id}/claim")
async def claim_history_item(record_id: int, user: dict = Depends(require_permission("claim_calculation"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        await cursor.execute("UPDATE Calculation_History SET status = 'claimed', claimed_by = ?, claimed_at = ? WHERE id = ?", (user["id"], now_str, record_id))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="Record not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "CLAIM_CALCULATION", f"Claimed calculation ID: {record_id}")
        return {"message": "Calculation claimed successfully"}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error claiming record: {str(e)}")

@app.get("/history/{record_id}")
async def get_history_record(record_id: int, user: dict = Depends(require_permission("view_history"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT ch.*, 
                   u1.username as created_by_name, 
                   u2.username as submitted_by_name, 
                   u3.username as claimed_by_name 
            FROM Calculation_History ch
            LEFT JOIN Users u1 ON ch.created_by = u1.id
            LEFT JOIN Users u2 ON ch.submitted_by = u2.id
            LEFT JOIN Users u3 ON ch.claimed_by = u3.id
            WHERE ch.id = ?
        """, (record_id,))
        row = await cursor.fetchone()
        if not row:
            await conn.close()
            raise HTTPException(status_code=404, detail="Record not found")

        columns = [desc[0] for desc in cursor.description]
        record = dict(zip(columns, row))
        record['doc_nums'] = json.loads(record['doc_nums']) if record['doc_nums'] else []
        
        # Swap the integer IDs for the dynamic string usernames
        record["created_by"] = record.pop("created_by_name", None)
        record["submitted_by"] = record.pop("submitted_by_name", None)
        record["claimed_by"] = record.pop("claimed_by_name", None)

        await cursor.execute("""
            SELECT code, name, 
                   COALESCE(edited_weight, weight) AS weight, 
                   doc_date, sin_no, principal, brand, 
                   COALESCE(edited_ctns, ctns) AS ctns, 
                   bu, b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal,
                   calculation_type, system_rate, unit_cost, total_cost, standard_unit_cost,
                   FromWhsCode, ToWhsCode, 
                   CASE WHEN edited_ctns IS NOT NULL THEN ctns ELSE NULL END AS original_ctns, 
                   CASE WHEN edited_weight IS NOT NULL THEN weight ELSE NULL END AS original_weight
            FROM Calculation_Products WHERE calc_id = ?
        """, (record_id,))
        prod_rows = await cursor.fetchall()
        prod_cols = ["code","name","weight","doc_date","sin_no","principal","brand","ctns","bu",
                     "b_code","b_name","b_dept","b_principal","b_desc","s_dept","s_principal",
                     "calculation_type","system_rate","unit_cost","total_cost","standard_unit_cost",
                     "FromWhsCode", "ToWhsCode", "original_ctns", "original_weight"]
        record['calculated_products'] = [dict(zip(prod_cols, r)) for r in prod_rows]

        await cursor.execute("""
            SELECT department, item_name, uom, quantity, unit_cost, total_cost,
                   b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal
            FROM Calculation_POSM_Products WHERE calc_id = ?
        """, (record_id,))
        posm_rows = await cursor.fetchall()
        posm_cols = ["department", "item_name", "uom", "quantity", "unit_cost", "total_cost",
                     "b_code", "b_name", "b_dept", "b_principal", "b_desc", "s_dept", "s_principal"]
        record['posm_products'] = [dict(zip(posm_cols, r)) for r in posm_rows]

        await conn.close()
        return record
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/history")
async def get_history(user: dict = Depends(require_permission("view_history"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        permissions = user.get('permissions', [])
        username = user.get('username')

        base_query = """
            SELECT ch.*, 
                   u1.username as created_by_name, 
                   u2.username as submitted_by_name, 
                   u3.username as claimed_by_name 
            FROM Calculation_History ch
            LEFT JOIN Users u1 ON ch.created_by = u1.id
            LEFT JOIN Users u2 ON ch.submitted_by = u2.id
            LEFT JOIN Users u3 ON ch.claimed_by = u3.id
        """

        if 'view_all_history' in permissions:
            await cursor.execute(base_query + " ORDER BY ch.created_at DESC")
        elif 'claim_calculation' in permissions:
            await cursor.execute(base_query + """
                WHERE ch.created_by = ? OR ch.status IN ('submitted', 'claimed') 
                ORDER BY ch.created_at DESC
            """, (user['id'],))
        else:
            await cursor.execute(base_query + " WHERE ch.created_by = ? OR ch.created_by IS NULL ORDER BY ch.created_at DESC", (user['id'],))
            
        rows = await cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        history = []
        for row in rows:
            record_dict = dict(zip(columns, row))
            record_dict["doc_nums"] = json.loads(record_dict["doc_nums"]) if record_dict["doc_nums"] else []
            
            # Swap the integer IDs for the dynamic string usernames for the frontend
            record_dict["created_by"] = record_dict.pop("created_by_name", None)
            record_dict["submitted_by"] = record_dict.pop("submitted_by_name", None)
            record_dict["claimed_by"] = record_dict.pop("claimed_by_name", None)
            
            history.append(record_dict)
        await conn.close()
        return {"history": history}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error loading history: {str(e)}")

@app.delete("/history/{record_id}")
async def delete_history_item(record_id: int, user: dict = Depends(require_permission("delete_history"))):
    conn = None
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id FROM Calculation_History WHERE id = ?", (record_id,))
        if not await cursor.fetchone():
            raise HTTPException(status_code=404, detail="Record not found")
        await cursor.execute("DELETE FROM Calculation_Products WHERE calc_id = ?", (record_id,))
        await cursor.execute("DELETE FROM Calculation_POSM_Products WHERE calc_id = ?", (record_id,))
        await cursor.execute("DELETE FROM Calculation_History WHERE id = ?", (record_id,))
        await conn.commit()
        await log_user_activity(user['id'], "DELETE_CALCULATION", f"Deleted calculation ID: {record_id}")
        return {"message": "Record deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error deleting record: {str(e)}")
    finally:
        if conn: await conn.close()

@app.get("/history/{record_id}/download")
async def download_history_excel(record_id: int, user: dict = Depends(require_permission("view_history"))):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT * FROM Calculation_History WHERE id = ?", (record_id,))
        row = await cursor.fetchone()
        
        if not row:
            await conn.close()
            raise HTTPException(status_code=404, detail="History record not found")
            
        columns = [desc[0] for desc in cursor.description]
        record = dict(zip(columns, row))
        record['doc_nums'] = json.loads(record['doc_nums']) if record['doc_nums'] else []
        
        await cursor.execute("""
            SELECT code, name, 
                   COALESCE(edited_weight, weight) AS weight, 
                   doc_date, sin_no, principal, brand, 
                   COALESCE(edited_ctns, ctns) AS ctns, 
                   bu, b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal,
                   calculation_type, system_rate, unit_cost, total_cost, standard_unit_cost,
                   FromWhsCode, ToWhsCode, 
                   CASE WHEN edited_ctns IS NOT NULL THEN ctns ELSE NULL END AS original_ctns, 
                   CASE WHEN edited_weight IS NOT NULL THEN weight ELSE NULL END AS original_weight
            FROM Calculation_Products WHERE calc_id = ?
        """, (record_id,))
        
        prod_rows = await cursor.fetchall()
        
        if not prod_rows:
            await conn.close() # Close connection before raising an error
            raise HTTPException(status_code=404, detail="Historical product data not found for this saved record.")
            
        prod_cols = ["code","name","weight","doc_date","sin_no","principal","brand","ctns","bu",
                     "b_code","b_name","b_dept","b_principal","b_desc","s_dept","s_principal",
                     "calculation_type","system_rate","unit_cost","total_cost","standard_unit_cost",
                     "FromWhsCode", "ToWhsCode", "original_ctns", "original_weight"]
        products = [dict(zip(prod_cols, r)) for r in prod_rows]
        
        # Filter out items where cartons have been reduced to 0
        products = [p for p in products if Decimal(str(p.get('ctns', 0))) > 0]
        
        # Location -> Branch code mapping (e.g. "Yangon" -> "YGN"), used for description fields below
        await cursor.execute("SELECT to_location, branch_code FROM Location_Mapping")
        loc_mapping_rows = await cursor.fetchall()

        await cursor.execute("SELECT log_pric, code, name FROM SD_Code")
        sd_code_excel_map = {str(row[0]).strip().lower(): {"code": row[1], "name": row[2]} for row in await cursor.fetchall() if row[0]}

        # --- POSM Calculation sheet (optional) ---
        await cursor.execute("""
            SELECT department, item_name, uom, quantity, unit_cost, total_cost,
                   b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal
            FROM Calculation_POSM_Products WHERE calc_id = ?
        """, (record_id,))
        posm_rows = await cursor.fetchall()

        await conn.close()

        channel_name = record.get('channel', '')
        is_sd_channel = channel_name in ["SD", "Telecom SD"]
        is_branch_channel = channel_name in ["Branch", "Outlet", "Telecom Branch"]

        LOCATION_CODE_MAP = {}
        for loc_row in loc_mapping_rows:
            if loc_row[0] and loc_row[1]:
                LOCATION_CODE_MAP[str(loc_row[0]).strip().upper()] = str(loc_row[1]).strip()

        def _to_loc_code(loc_name):
            if not loc_name:
                return ""
            return LOCATION_CODE_MAP.get(str(loc_name).strip().upper(), loc_name)

        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cost Details"

            headers = [
                "No", "Claim Date", "Delivery Date", "SIN No", "Area", "Code", "Name", "Principal", "Brand", "Item Code", "Item", "Ctns", 
                "Price", "Total Amount", "Weight", "UOM", "Gate", "Channel", "Month", "Year", "Description for Account", 
                "Description with cnts and price", "Branch"
            ]
            
            if not is_sd_channel:
                headers.extend(["B-Dept", "B-Principal"])
            if not is_branch_channel:
                headers.extend(["S-Dept", "S-Principal"])
                
            headers.extend(["BU", "Calculation ID"])
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            border_style = Side(border_style="thin", color="000000")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            now = datetime.datetime.now()
            claim_month = now.strftime("%B") 
            claim_year = now.year

            claimed_at_val = record.get('claimed_at')
            if record.get('status') == 'claimed' and claimed_at_val:
                if isinstance(claimed_at_val, datetime.datetime):
                    claim_date_str = claimed_at_val.strftime("%d/%m/%Y")
                else:
                    try: claim_date_str = datetime.datetime.strptime(str(claimed_at_val)[:19], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
                    except ValueError: claim_date_str = str(claimed_at_val)
            else:
                claim_date_str = ""
            
            cost_details_delivery_date = ""
            cost_details_sin_no = ""

            for idx, item in enumerate(products, 1):
                row_num = idx + 1
                doc_date_val = item.get('doc_date')
                if isinstance(doc_date_val, datetime.datetime): doc_date_str = doc_date_val.strftime("%d/%m/%Y")
                elif isinstance(doc_date_val, str) and len(doc_date_val) >= 10:
                    try: doc_date_str = datetime.datetime.strptime(doc_date_val[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except ValueError: doc_date_str = doc_date_val
                else: doc_date_str = str(doc_date_val) if doc_date_val else ""

                b_code_val = item.get('b_code', '')
                b_name_val = item.get('b_name', '')
                b_desc = item.get('b_desc', '')
                
                display_code = b_code_val
                display_name = b_name_val
                base_desc = b_desc.strip() if b_desc else ""

                principal_val = str(item.get('principal', '')).strip().lower()
                if is_sd_channel:
                    sd_info = sd_code_excel_map.get(principal_val, {})
                    display_code = sd_info.get('code', display_code)
                    display_name = sd_info.get('name', display_name)
                    base_desc = sd_info.get('name', base_desc)

                from_loc_code = _to_loc_code(record['from_loc'])
                to_loc_code = _to_loc_code(record['to_loc'])
                desc_for_account = f"{base_desc}-{from_loc_code} to {to_loc_code}"

                ctns_val = Decimal(str(item.get('ctns', 0)))
                total_cost_val = Decimal(str(item.get('total_cost', 0)))
                ctns_formatted = int(ctns_val) if float(ctns_val).is_integer() else ctns_val
                price_per_ctn = total_cost_val / ctns_val if ctns_val > Decimal("0") else Decimal("0.0")
                price_formatted = int(price_per_ctn) if float(price_per_ctn).is_integer() else round(price_per_ctn, 2)

                concat_desc = f"{desc_for_account} - {ctns_formatted} ctns @{price_formatted} kyats"

                raw_sin_no = str(item.get('sin_no', ''))
                clean_sin_no = raw_sin_no.replace('PDG - ', '').replace('PDG-', '').replace('PG - ', '').replace('PG-', '').strip()

                if idx == 1:
                    cost_details_delivery_date = doc_date_str
                    cost_details_sin_no = clean_sin_no

                b_dept_val = item.get('b_dept', '')
                b_principal_val = item.get('b_principal', '')
                s_dept_val = item.get('s_dept', '')
                s_principal_val = item.get('s_principal', '')

                row_vals = [
                    idx, claim_date_str, doc_date_str, clean_sin_no, record['to_loc'],
                    display_code, display_name, item.get('principal', ''), item.get('brand', ''),
                    item.get('code', ''), item.get('name', ''), ctns_formatted,
                    float(price_per_ctn), float(total_cost_val), float(item.get('weight', 0)),
                    "Kg", record['gate_name'], record.get('channel', ''), claim_month, claim_year,
                    desc_for_account, concat_desc, record['to_loc']
                ]

                if not is_sd_channel:
                    row_vals.extend([b_dept_val, b_principal_val])
                if not is_branch_channel:
                    row_vals.extend([s_dept_val, s_principal_val])

                row_vals.extend([item.get('bu', ''), record['id']])

                for col_num, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=val)
                    cell.border = border
                    if isinstance(val, (int, float, Decimal)) and col_num in (13, 14, 15):
                        cell.number_format = '#,##0.00'
            
            for col in ws.columns:    
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = max_length + 2

            if posm_rows:
                ws2 = wb.create_sheet("POSM Details")
                posm_headers = [
                    "No", "Claim Date", "Delivery Date", "SIN No", "Area", "Name", "Department", "Principal", "Brand",
                    "Item", "Quantity", "Unit Cost", "Total cost", "Uom", "Gate", "Channel", "Month", "Year",
                    "Description for Account", "Description with pcs and price", "Branch"
                ]
                
                if not is_sd_channel:
                    posm_headers.extend(["B-Dept", "B-Principal"])
                if not is_branch_channel:
                    posm_headers.extend(["S-Dept", "S-Principal"])
                    
                posm_headers.append("Calculation ID")

                for col_num, header in enumerate(posm_headers, 1):
                    cell = ws2.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border

                from_loc_code = _to_loc_code(record['from_loc'])
                to_loc_code = _to_loc_code(record['to_loc'])

                for idx, r in enumerate(posm_rows, 1):
                    row_num = idx + 1
                    department, item_name, uom, quantity, unit_cost, total_cost, b_code, b_name, b_dept, b_principal, b_desc, s_dept, s_principal = r

                    qty_val = Decimal(str(quantity or 0))
                    qty_formatted = int(qty_val) if float(qty_val).is_integer() else qty_val
                    unit_cost_val = Decimal(str(unit_cost or 0))
                    unit_cost_formatted = int(unit_cost_val) if float(unit_cost_val).is_integer() else round(unit_cost_val, 2)

                    display_name = b_name or "Transport Charges-POSM"
                    b_desc_clean = b_desc.strip() if b_desc else "POSM-Transport Charges"

                    if is_sd_channel:
                        sd_info = sd_code_excel_map.get("posm", {})
                        display_name = sd_info.get('name', display_name)
                        b_desc_clean = sd_info.get('name', b_desc_clean)

                    posm_desc_account = f"{b_desc_clean}-{from_loc_code} to {to_loc_code}"
                    posm_desc_with_price = f"{posm_desc_account} - {qty_formatted} pcs @{unit_cost_formatted} kyats"

                    b_dept_val = b_dept or "Logistics"
                    b_principal_val = b_principal or "POSM"
                    s_dept_val = s_dept or "Logistics"
                    s_principal_val = s_principal or "POSM"

                    row_vals = [
                        idx, claim_date_str, cost_details_delivery_date, cost_details_sin_no, record['to_loc'],
                        display_name, department or "", b_principal or "POSM", b_principal or "POSM", item_name or "", float(qty_val),
                        float(unit_cost_val), float(total_cost or 0), uom or "", record['gate_name'],
                        record.get('channel', ''), claim_month, claim_year, posm_desc_account, posm_desc_with_price,
                        record['to_loc']
                    ]
                    
                    if not is_sd_channel:
                        row_vals.extend([b_dept_val, b_principal_val])
                    if not is_branch_channel:
                        row_vals.extend([s_dept_val, s_principal_val])
                    
                    row_vals.append(record_id)

                    for col_num, val in enumerate(row_vals, 1):
                        cell = ws2.cell(row=row_num, column=col_num, value=val)
                        cell.border = border
                        if isinstance(val, (int, float, Decimal)) and col_num in (11, 12, 13):
                            cell.number_format = '#,##0.00'
                for col in ws2.columns:    
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                    ws2.column_dimensions[col_letter].width = max_length + 2

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        filename = f"Calculation_{record_id}_{record['gate_name']}.xlsx"
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error generating download: {str(e)}")
    
# --- Item Pricing Excel Export/Import ---

@app.get("/account/item-pricing/export/{gate_id}")
async def export_item_pricing_excel(gate_id: int):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT gate_name, from_loc, to_loc FROM Gate WHERE id = ?", (gate_id,))
        gate_row = await cursor.fetchone()
        if not gate_row:
            await conn.close()
            raise HTTPException(status_code=404, detail="Gate not found")
        gate_name, from_loc, to_loc = gate_row[0], gate_row[1] or "", gate_row[2] or ""
        
        await cursor.execute("SELECT bu, item_id, item_name, principal, brand, transportation_cost FROM Item_Pricing WHERE gate_id = ? ORDER BY item_id", (gate_id,))
        rows = await cursor.fetchall()
        await conn.close()
        
        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Item Pricing"
            ws['A1'] = f"Gate: {gate_name} ({from_loc} -> {to_loc})"
            ws['A1'].font = Font(bold=True, size=14)
            
            headers = ['BU', 'Item Code', 'Item Name', 'Principal', 'Brand', 'Transportation Cost']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            for row_num, row_data in enumerate(rows, 4):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        filename = f"item_pricing_{gate_name.replace(' ', '_')}.xlsx"
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")

@app.post("/account/item-pricing/import/{gate_id}")
async def import_item_pricing_excel(gate_id: int, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    if "add_item" not in perms or "edit_item" not in perms:
        raise HTTPException(status_code=403, detail="Requires both 'add_item' and 'edit_item' permissions for bulk import")

    try:
        contents = await file.read()
        file_bytes = io.BytesIO(contents)
        wb = await run_in_threadpool(openpyxl.load_workbook, file_bytes)
        ws = wb.active
        
        excel_rows = []
        item_codes_to_check = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if not row[1]: continue
            item_code = str(row[1]).strip()
            excel_rows.append({
                "row_num": row_idx, 
                "bu": str(row[0]).strip() if row[0] else "", 
                "code": item_code, 
                "name": str(row[2]).strip() if row[2] else "",
                "principal": str(row[3]).strip() if row[3] else "", 
                "brand": str(row[4]).strip() if row[4] else "",
                "cost": str(row[5]).strip() if len(row) > 5 and row[5] else ""
            })
            item_codes_to_check.add(item_code)
            
        if not excel_rows: raise HTTPException(status_code=400, detail="No data found in Excel file")

        conn_dwbi = await get_dwbi_connection()
        cursor_dwbi = await conn_dwbi.cursor()
        dwbi_data = {}
        unique_codes_list = list(item_codes_to_check)
        batch_size = 1000
        
        for i in range(0, len(unique_codes_list), batch_size):
            batch = unique_codes_list[i:i + batch_size]
            placeholders = ','.join('?' * len(batch))
            await cursor_dwbi.execute(f"SELECT ItemCode, ItemName, ItmsGrpNam, U_BrandName, Sector FROM Itemmasterallpp WHERE ItemCode IN ({placeholders})", batch)
            rows = await cursor_dwbi.fetchall()
            for r in rows:
                dwbi_data[str(r[0]).strip()] = {
                    "name": str(r[1]).strip() if r[1] else "", 
                    "principal": str(r[2]).strip() if r[2] else "", 
                    "brand": str(r[3]).strip() if r[3] else "",
                    "bu": str(r[4]).strip() if r[4] else ""
                }
        await conn_dwbi.close()
        
        errors = []
        for row in excel_rows:
            code = row["code"]
            if code not in dwbi_data:
                errors.append(f"Row {row['row_num']}: Item Code '{code}' not found in DWBI.")
                continue
            db_item = dwbi_data[code]
            if row["name"].lower() != db_item["name"].lower(): errors.append(f"Row {row['row_num']}: Item Name mismatch. Excel: '{row['name']}', System: '{db_item['name']}'")
            if row["principal"].lower() != db_item["principal"].lower(): errors.append(f"Row {row['row_num']}: Principal mismatch.")
            if row["brand"].lower() != db_item["brand"].lower(): errors.append(f"Row {row['row_num']}: Brand mismatch.")
            if row["bu"].lower() != db_item["bu"].lower(): errors.append(f"Row {row['row_num']}: BU mismatch.")

        if errors:
            error_msg = errors[:10]
            if len(errors) > 10: error_msg.append(f"... and {len(errors) - 10} more errors.")
            raise HTTPException(status_code=400, detail=error_msg)

        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        await cursor.execute("SELECT item_id, id, transportation_cost FROM Item_Pricing WHERE gate_id = ?", (gate_id,))
        existing_items_data = {row[0]: {'pricing_id': row[1], 'cost': row[2]} for row in await cursor.fetchall()}
        existing_items = set(existing_items_data.keys())
        
        await cursor.execute("SELECT id FROM Item_Pricing")
        used_ids = {row[0] for row in await cursor.fetchall()}
        
        updated_items_set = set()
        updates_made, inserts_made, deletes_made = 0, 0, 0
        
        change_logs = []
        change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        username = user['username']
        
        for row in excel_rows:
            item_code = row["code"]
            updated_items_set.add(item_code)
            
            if item_code in existing_items:
                pricing_id = existing_items_data[item_code]['pricing_id']
                old_cost = existing_items_data[item_code]['cost']
                old_cost_str = str(old_cost).strip() if old_cost else ""
                new_cost_str = str(row["cost"]).strip() if row["cost"] else ""
                
                if old_cost_str != new_cost_str:
                    change_logs.append((pricing_id, user['id'], change_date, 'Transportation Cost', old_cost_str, new_cost_str))

                # ADDED: edited_at and edited_by
                await cursor.execute("""
                    UPDATE Item_Pricing
                    SET bu = ?, item_name = ?, principal = ?, brand = ?, transportation_cost = ?,
                        edited_at = ?, edited_by = ?
                    WHERE id = ?
                """, (row["bu"], row["name"], row["principal"], row["brand"], row["cost"], change_date, user['id'], pricing_id))
                updates_made += 1
            else:
                while True:
                    new_id = random.randint(10000000, 99999999)
                    if new_id not in used_ids:
                        used_ids.add(new_id)
                        break
                # ADDED: created_at and created_by
                await cursor.execute("""
                    INSERT INTO Item_Pricing (id, gate_id, bu, item_id, item_name, principal, brand, transportation_cost, created_at, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (new_id, gate_id, row["bu"], item_code, row["name"], row["principal"], row["brand"], row["cost"], change_date, user['id']))
                inserts_made += 1
                
        if change_logs:
            await cursor.executemany("INSERT INTO Item_Change_Log (pricing_id, changed_by, change_date, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)", change_logs)
        
        items_to_delete = existing_items - updated_items_set
        for item_code in items_to_delete:
            pricing_id = existing_items_data[item_code]['pricing_id']
            await cursor.execute("DELETE FROM Item_Change_Log WHERE pricing_id = ?", (pricing_id,))
            await cursor.execute("DELETE FROM Item_Pricing WHERE id = ?", (pricing_id,))
            deletes_made += 1
            
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "BULK_IMPORT_ITEMS", f"Imported items to gate ID {gate_id} (Updates: {updates_made}, Inserts: {inserts_made}, Deletes: {deletes_made})")
        return {"message": "Import completed successfully", "updates": updates_made, "inserts": inserts_made, "deletes": deletes_made}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error importing: {str(e)}")

# --- Gate Management Endpoints (SQLite) ---

@app.get("/account/gates")
async def get_all_gates(
    gate_name: Optional[str] = Query(None, description="Filter by gate name"),
    from_loc: Optional[str] = Query(None, description="Filter by origin location"),
    to_loc: Optional[str] = Query(None, description="Filter by destination location"),
    uom: Optional[str] = Query(None, description="Filter by UOM"),
    cost: Optional[str] = Query(None, description="Filter by cost"),
    user: dict = Depends(get_current_user)
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        query = "SELECT id, gate_name, from_loc, to_loc, destination, uom, unit, cost FROM Gate WHERE 1=1"
        params = []
        
        if gate_name:
            query += " AND gate_name LIKE ?"
            params.append(f"%{gate_name}%")
        if from_loc:
            query += " AND from_loc LIKE ?"
            params.append(f"%{from_loc}%")
        if to_loc:
            query += " AND to_loc LIKE ?"
            params.append(f"%{to_loc}%")
        if uom:
            query += " AND uom LIKE ?"
            params.append(f"%{uom}%")
        if cost:
            query += " AND CAST(cost AS TEXT) LIKE ?"
            params.append(f"%{cost}%")

        await cursor.execute(query, params)
        rows = await cursor.fetchall()
        
        gates = []
        for row in rows:
            gate_id = row[0]
            calc_type = await determine_calculation_type_sql(gate_id)
            gates.append({
                "gate_id": gate_id, 
                "gate_name": row[1], 
                "from_loc": row[2], 
                "to_loc": row[3],
                "destination": row[4],
                "uom": row[5],         
                "unit": row[6], 
                "cost": float(row[7]) if row[7] is not None else None, 
                "calculation_type": calc_type
            })
        await conn.close()
        return {"gates": gates}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Error loading gates: {str(e)}")

@app.post("/account/gates")
async def save_gate(gate_data: GateData, user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        await cursor.execute("""
            SELECT id FROM Gate 
            WHERE gate_name = ? AND from_loc = ? AND to_loc = ?
        """, (gate_data.gate_name, gate_data.from_loc, gate_data.to_loc))
        existing_gate = await cursor.fetchone()

        if existing_gate:
            if gate_data.gate_id is None:
                await conn.close()
                raise HTTPException(status_code=400, detail="A gate with this name, origin, and destination already exists.")
            elif existing_gate[0] != gate_data.gate_id:
                await conn.close()
                raise HTTPException(status_code=400, detail="Another gate with this name, origin, and destination already exists.")

        if gate_data.gate_id is not None:
            if "edit_gate" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'edit_gate' permission")
            
            await cursor.execute("SELECT id, uom, unit, cost FROM Gate WHERE id = ?", (gate_data.gate_id,))
            current_row = await cursor.fetchone()
            
            if current_row:
                gate_id, old_uom, old_unit, old_cost = current_row
                old_uom = old_uom if old_uom else None
                new_uom = gate_data.uom if gate_data.uom else None
                old_unit = old_unit if old_unit is not None else None
                new_unit = gate_data.unit if gate_data.unit is not None else None
                old_cost = Decimal(str(old_cost)) if old_cost is not None else None
                new_cost = gate_data.cost if gate_data.cost is not None else None

                change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                username = user['id']
                changes = []

                if old_uom != new_uom: changes.append((gate_id, user['id'], change_date, 'UOM', str(old_uom or ''), str(new_uom or '')))
                if old_unit != new_unit: changes.append((gate_id, user['id'], change_date, 'Unit', str(old_unit) if old_unit is not None else '', str(new_unit) if new_unit is not None else ''))
                if old_cost != new_cost: changes.append((gate_id, user['id'], change_date, 'Cost', str(old_cost) if old_cost is not None else '', str(new_cost) if new_cost is not None else ''))
                
                if changes:
                    await cursor.executemany("INSERT INTO Gate_Change_Log (gate_id, changed_by, change_date, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)", changes)

            await cursor.execute("""
                UPDATE Gate SET gate_name = ?, from_loc = ?, to_loc = ?, destination = ?, uom = ?, unit = ?, cost = ?, edited_at = ?, edited_by = ? WHERE id = ?
            """, (gate_data.gate_name, gate_data.from_loc, gate_data.to_loc, gate_data.destination, gate_data.uom, gate_data.unit, gate_data.cost, change_date, username, gate_data.gate_id))
            await log_user_activity(user['id'], "UPDATE_GATE", f"Updated gate ID {gate_data.gate_id}: {gate_data.gate_name}")
        else:
            if "add_gate" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'add_gate' permission")
                
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']

            await cursor.execute("INSERT INTO Gate (gate_name, from_loc, to_loc, destination, uom, unit, cost, created_at, created_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                  (gate_data.gate_name, gate_data.from_loc, gate_data.to_loc, gate_data.destination, gate_data.uom, gate_data.unit, gate_data.cost, now_str, username))
            await log_user_activity(user['id'], "CREATE_GATE", f"Created gate: {gate_data.gate_name}")
        
        await conn.commit()
        await conn.close()
        return {"message": "Gate saved successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error saving gate: {str(e)}")

@app.get("/account/gates/{gate_id}/logs", response_model=List[GateLogItem])
async def get_gate_logs(gate_id: int, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.gate_id, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM Gate_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.gate_id = ? ORDER BY l.change_date DESC
        """, (gate_id,))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "gate_id": r[1], "changed_by": r[2], "change_date": r[3], "field_name": r[4], "old_value": r[5], "new_value": r[6]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching logs: {str(e)}")

@app.delete("/account/gates/{gate_id}")
async def delete_gate(gate_id: int, user: dict = Depends(require_permission("delete_gate"))): 
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("DELETE FROM Gate_Change_Log WHERE gate_id = ?", (gate_id,))
        await cursor.execute("DELETE FROM Item_Pricing WHERE gate_id = ?", (gate_id,))
        await cursor.execute("DELETE FROM Gate WHERE id = ?", (gate_id,))
        if cursor.rowcount == 0:
             await conn.close()
             raise HTTPException(status_code=404, detail="Gate not found")
        await conn.commit()
        await conn.close()
        
        await log_user_activity(user['id'], "DELETE_GATE", f"Deleted gate ID: {gate_id}")
        return {"message": f"Gate {gate_id} deleted successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error deleting gate: {str(e)}")

@app.get("/account/item-pricing/{gate_id}")
async def get_item_pricing(gate_id: int, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id, bu, item_id, item_name, principal, brand, transportation_cost FROM Item_Pricing WHERE gate_id = ?", (gate_id,))
        rows = await cursor.fetchall()
        items = [{"pricing_id": r[0], "bu": r[1], "item_code": r[2], "item_name": r[3], "principal": r[4], "brand": r[5], "transportation_cost": r[6]} for r in rows]
        await conn.close()
        return {"items": items, "gate_id": gate_id}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error loading items: {str(e)}")

@app.post("/account/item-pricing")
async def save_item_pricing(item_data: ItemPricingData, user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        is_edit = item_data.pricing_id is not None

        if is_edit:
            if "edit_item" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'edit_item' permission")

            await cursor.execute("SELECT id, transportation_cost FROM Item_Pricing WHERE id = ?", (item_data.pricing_id,))
            existing = await cursor.fetchone()
            if not existing:
                await conn.close()
                raise HTTPException(status_code=404, detail="Item not found")
            pricing_id, old_cost = existing

            # If the item code is being changed, make sure it doesn't collide with a different row on this gate
            await cursor.execute(
                "SELECT id FROM Item_Pricing WHERE gate_id = ? AND item_id = ? AND id != ?",
                (item_data.gate_id, item_data.item_code, pricing_id)
            )
            collision = await cursor.fetchone()
            if collision:
                await conn.close()
                raise HTTPException(status_code=400, detail=f"Item '{item_data.item_code}' already exists on this gate. Duplicate items are not allowed.")

            old_cost_str = str(old_cost).strip() if old_cost else ""
            new_cost_str = str(item_data.transportation_cost).strip() if item_data.transportation_cost else ""

            change_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']
            changes = []

            if old_cost_str != new_cost_str:
                 changes.append((pricing_id, user['id'], change_date, 'Transportation Cost', old_cost_str, new_cost_str))

            if changes:
                 await cursor.executemany("INSERT INTO Item_Change_Log (pricing_id, changed_by, change_date, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?)", changes)

            await cursor.execute("UPDATE Item_Pricing SET item_id = ?, bu = ?, item_name = ?, principal = ?, brand = ?, transportation_cost = ?, edited_at = ?, edited_by = ? WHERE id = ?", 
                (item_data.item_code, item_data.bu, item_data.item_name, item_data.principal, item_data.brand, item_data.transportation_cost, change_date, username, pricing_id))
            await log_user_activity(user['id'], "UPDATE_ITEM_PRICING", f"Updated pricing for item {item_data.item_code} on gate ID {item_data.gate_id}")
        else:
            if "add_item" not in perms:
                await conn.close()
                raise HTTPException(status_code=403, detail="Requires 'add_item' permission")

            await cursor.execute("SELECT id FROM Item_Pricing WHERE gate_id = ? AND item_id = ?", (item_data.gate_id, item_data.item_code))
            existing = await cursor.fetchone()
            if existing:
                await conn.close()
                raise HTTPException(status_code=400, detail=f"Item '{item_data.item_code}' already exists on this gate. Duplicate items are not allowed.")

            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            username = user['id']

            while True:
                new_id = random.randint(10000000, 99999999)
                await cursor.execute("SELECT 1 FROM Item_Pricing WHERE id = ?", (new_id,))
                if not await cursor.fetchone(): break
            await cursor.execute("INSERT INTO Item_Pricing (id, gate_id, bu, item_id, item_name, principal, brand, transportation_cost, created_at, created_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                (new_id, item_data.gate_id, item_data.bu, item_data.item_code, item_data.item_name, item_data.principal, item_data.brand, item_data.transportation_cost, now_str, username))

            await log_user_activity(user['id'], "CREATE_ITEM_PRICING", f"Added pricing for item {item_data.item_code} to gate ID {item_data.gate_id}")

        await conn.commit()
        await conn.close()
        return {"message": "Item saved successfully"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error saving item: {str(e)}")

@app.get("/account/items/{pricing_id}/logs", response_model=List[ItemLogItem])
async def get_item_logs(pricing_id: int, user: dict = Depends(get_current_user)):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.id, l.pricing_id, u.username, l.change_date, l.field_name, l.old_value, l.new_value 
            FROM Item_Change_Log l
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.pricing_id = ? ORDER BY l.change_date DESC
        """, (pricing_id,))
        rows = await cursor.fetchall()
        logs = [{"id": r[0], "pricing_id": r[1], "changed_by": r[2], "change_date": r[3], "field_name": r[4], "old_value": r[5], "new_value": r[6]} for r in rows]
        await conn.close()
        return logs
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching logs: {str(e)}")

@app.get("/account/reports/cost-changes")
async def get_cost_change_report(user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    if "view_gates" not in perms and "view_items" not in perms:
        raise HTTPException(status_code=403, detail="Requires 'view_gates' or 'view_items' permission")
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        gate_cost_changes = []
        if "view_gates" in perms:
            await cursor.execute("""
                SELECT l.id, l.gate_id, g.gate_name, g.from_loc, g.to_loc, u.username, l.change_date, l.old_value, l.new_value
                FROM Gate_Change_Log l
                LEFT JOIN Gate g ON l.gate_id = g.id
                LEFT JOIN Users u ON l.changed_by = u.id
                WHERE l.field_name = 'Cost'
                ORDER BY l.change_date DESC
            """)
            rows = await cursor.fetchall()
            gate_cost_changes = [{
                "id": r[0], "gate_id": r[1], "gate_name": r[2] or 'Deleted Gate',
                "from_loc": r[3], "to_loc": r[4], "changed_by": r[5],
                "change_date": r[6], "old_value": r[7], "new_value": r[8]
            } for r in rows]

        item_cost_changes = []
        if "view_items" in perms:
            await cursor.execute("""
                SELECT l.id, l.pricing_id, ip.item_id, ip.item_name, ip.gate_id, g.gate_name, u.username, l.change_date, l.old_value, l.new_value, ip.principal, ip.brand, g.from_loc, g.to_loc
                FROM Item_Change_Log l
                LEFT JOIN Item_Pricing ip ON l.pricing_id = ip.id
                LEFT JOIN Gate g ON ip.gate_id = g.id
                LEFT JOIN Users u ON l.changed_by = u.id
                WHERE l.field_name = 'Transportation Cost'
                ORDER BY l.change_date DESC
            """)
            rows = await cursor.fetchall()
            item_cost_changes = [{
                "id": r[0], "pricing_id": r[1], "item_code": r[2] or 'Deleted Item',
                "item_name": r[3], "gate_id": r[4], "gate_name": r[5],
                "changed_by": r[6], "change_date": r[7], "old_value": r[8], "new_value": r[9],
                "principal": r[10], "brand": r[11], "from_loc": r[12], "to_loc": r[13]
            } for r in rows]

        await conn.close()
        return {"gate_cost_changes": gate_cost_changes, "item_cost_changes": item_cost_changes}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error fetching cost change report: {str(e)}")

@app.get("/account/reports/cost-changes/export/gates")
async def export_gate_cost_report(user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    if "view_gates" not in perms:
        raise HTTPException(status_code=403, detail="Requires 'view_gates' permission")
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.change_date, g.gate_name, g.from_loc, g.to_loc, u.username, l.old_value, l.new_value
            FROM Gate_Change_Log l
            LEFT JOIN Gate g ON l.gate_id = g.id
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.field_name = 'Cost'
            ORDER BY l.change_date DESC
        """)
        rows = await cursor.fetchall()
        await conn.close()

        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gate Cost Changes"
            headers = ['Date', 'Gate Name', 'From', 'To', 'Changed By', 'Old Cost', 'New Cost']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            for row_num, row_data in enumerate(rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=gate_cost_changes.xlsx"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")

@app.get("/account/reports/cost-changes/export/items")
async def export_item_cost_report(user: dict = Depends(get_current_user)):
    perms = user.get("permissions", [])
    if "view_items" not in perms:
        raise HTTPException(status_code=403, detail="Requires 'view_items' permission")
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("""
            SELECT l.change_date, ip.item_id, ip.item_name, ip.principal, ip.brand, g.gate_name, g.from_loc, g.to_loc, u.username, l.old_value, l.new_value
            FROM Item_Change_Log l
            LEFT JOIN Item_Pricing ip ON l.pricing_id = ip.id
            LEFT JOIN Gate g ON ip.gate_id = g.id
            LEFT JOIN Users u ON l.changed_by = u.id
            WHERE l.field_name = 'Transportation Cost'
            ORDER BY l.change_date DESC
        """)
        rows = await cursor.fetchall()
        await conn.close()

        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Item Transport Cost Changes"
            headers = ['Date', 'Item Code', 'Item Name', 'Principal', 'Brand', 'Gate', 'From', 'To', 'Changed By', 'Old Cost', 'New Cost']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            for row_num, row_data in enumerate(rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=item_transport_cost_changes.xlsx"}
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")

@app.delete("/account/item-pricing/{gate_id}/{item_code}")
async def delete_item_pricing(gate_id: int, item_code: str, user: dict = Depends(require_permission("delete_item"))): 
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT id FROM Item_Pricing WHERE gate_id = ? AND item_id = ?", (gate_id, item_code))
        row = await cursor.fetchone()
        if not row:
            await conn.close()
            raise HTTPException(status_code=404, detail="Item not found")

        pricing_id = row[0]
        await cursor.execute("DELETE FROM Item_Change_Log WHERE pricing_id = ?", (pricing_id,))

        await cursor.execute("DELETE FROM Item_Pricing WHERE id = ?", (pricing_id,))
        if cursor.rowcount == 0:
            await conn.close()
            raise HTTPException(status_code=404, detail="Item not found")
        await conn.commit()
        await conn.close()
        await log_user_activity(user['id'], "DELETE_ITEM_PRICING", f"Deleted item {item_code} from gate {gate_id}")
        return {"message": "Item deleted successfully"}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error deleting item: {str(e)}")

@app.get("/dwbi/items/search")
async def search_dwbi_items(q: str = Query(..., min_length=2)):
    try:
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        search_term = f"%{q}%"
        await cursor.execute("SELECT TOP 50 ItemCode, ItemName, ItmsGrpNam, U_BrandName, Sector FROM Itemmasterallpp WHERE ItemCode LIKE ? OR ItemName LIKE ?", (search_term, search_term))
        rows = await cursor.fetchall()
        items = [{"item_code": r[0], "item_name": r[1], "principal": r[2], "brand": r[3], "bu": r[4]} for r in rows]
        await conn.close()
        return {"items": items}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Error searching items: {str(e)}")

@app.get("/dwbi/items/validate")
async def validate_dwbi_item(code: str = Query(...)):
    try:
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT ItemCode, ItemName, ItmsGrpNam, U_BrandName, Sector FROM Itemmasterallpp WHERE ItemCode = ?", (code,))
        row = await cursor.fetchone()
        await conn.close()
        
        if row: 
            return {"valid": True, "item": {"item_code": row[0], "item_name": row[1], "principal": row[2], "brand": row[3], "bu": row[4]}}
        return {"valid": False}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Validation error: {str(e)}")
    
@app.get("/dwbi/principals/search")
async def search_dwbi_principals(q: str = Query(..., min_length=2)):
    try:
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        search_term = f"%{q}%"
        await cursor.execute("SELECT DISTINCT TOP 50 ItmsGrpNam FROM Itemmasterallpp WHERE ItmsGrpNam LIKE ?", (search_term,))
        rows = await cursor.fetchall()
        principals = [r[0] for r in rows if r[0]]
        await conn.close()
        return {"principals": principals}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Error searching principals: {str(e)}")

@app.get("/dwbi/principals/validate")
async def validate_dwbi_principal(name: str = Query(...)):
    try:
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT TOP 1 ItmsGrpNam FROM Itemmasterallpp WHERE ItmsGrpNam = ?", (name,))
        row = await cursor.fetchone()
        await conn.close()
        
        if row: 
            return {"valid": True, "principal": row[0]}
        return {"valid": False}
    except Exception as e: 
        raise HTTPException(status_code=500, detail=f"Validation error: {str(e)}")

@app.get("/locations/from")
async def get_from_locations():
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        await cursor.execute("SELECT DISTINCT from_loc FROM Gate WHERE from_loc IS NOT NULL ORDER BY from_loc")
        rows = await cursor.fetchall()
        await conn.close()
        return {"locations": [r[0] for r in rows if r[0]]}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error loading from locations: {str(e)}")

@app.get("/locations/to")
async def get_to_locations(from_loc: Optional[str] = None):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        if from_loc: await cursor.execute("SELECT DISTINCT to_loc FROM Gate WHERE from_loc = ? AND to_loc IS NOT NULL ORDER BY to_loc", (from_loc,))
        else: await cursor.execute("SELECT DISTINCT to_loc FROM Gate WHERE to_loc IS NOT NULL ORDER BY to_loc")
        rows = await cursor.fetchall()
        await conn.close()
        return {"locations": [r[0] for r in rows if r[0]]}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Error loading to locations: {str(e)}")

@app.post("/calculate-with-gate")
async def calculate_with_gate(
    gate_name: str, 
    from_loc: Optional[str] = Query(None),
    to_loc: Optional[str] = Query(None),
    doc_nums: List[str] = Query(...),
    manual_total_cost: Optional[Decimal] = None, 
    additional_charges: Optional[Decimal] = Decimal("0.0"),
    posm_items_json: Optional[str] = Query(None),
    posm_total_cost: Optional[Decimal] = None,
    edited_items_json: Optional[str] = Query(None), # <-- NEW
    user: dict = Depends(require_permission("view_calculator"))
):
    try:
        if not doc_nums: raise HTTPException(status_code=400, detail="No Doc Nums provided")

        posm_items = None
        if posm_items_json:
            try:
                posm_items = json.loads(posm_items_json)
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid POSM items payload")

        return await _perform_calculation_logic(
            gate_name=gate_name, 
            doc_nums=doc_nums, 
            from_loc=from_loc,
            to_loc=to_loc,
            manual_total_cost=manual_total_cost, 
            additional_charges=additional_charges,
            posm_items=posm_items,
            posm_total_cost=posm_total_cost,
            edited_items_json=edited_items_json # <-- NEW
        )
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=f"Calculation error: {str(e)}")

@app.get("/doc-nums")
async def get_doc_nums():
    try:
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        query = """
            SELECT source, DocNum, DocDate FROM (
                SELECT 'PG' as source, DocNum, MAX(DocDate) as DocDate FROM PG_Transfer_Details WHERE DocNum IS NOT NULL GROUP BY DocNum
                UNION ALL
                SELECT 'PDG' as source, DocNum, MAX(DocDate) as DocDate FROM PDG_Transfer_Details WHERE DocNum IS NOT NULL GROUP BY DocNum
            ) t
            ORDER BY DocDate DESC, DocNum DESC
        """
        await cursor.execute(query)
        rows = await cursor.fetchall()
        doc_nums = []
        for row in rows:
            doc_date_val = row[2]
            if isinstance(doc_date_val, datetime.datetime): doc_date_str = doc_date_val.strftime("%Y-%m-%d")
            elif isinstance(doc_date_val, str) and len(doc_date_val) >= 10: doc_date_str = doc_date_val[:10]
            else: doc_date_str = str(doc_date_val) if doc_date_val else ""
            
            doc_nums.append({"doc_num": f"{row[0]} - {row[1]}", "doc_date": doc_date_str})
        await conn.close()
        return {"doc_nums": doc_nums}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/products-by-doc-nums")
async def get_products_by_doc_nums(doc_nums: List[str] = Query(..., alias="doc_nums")):
    try:
        if not doc_nums: return {"products": [], "total_weight": Decimal("0.0")}
        
        pg_nums = [str(d).replace("PG - ", "").replace("PG-", "") for d in doc_nums if not str(d).startswith("PDG")]
        pdg_nums = [str(d).replace("PDG - ", "").replace("PDG-", "") for d in doc_nums if str(d).startswith("PDG")]
        
        conn = await get_dwbi_connection()
        cursor = await conn.cursor()
        products, total_weight = [], Decimal("0.0")
        
        if pg_nums:
            placeholders = ','.join('?' * len(pg_nums))
            await cursor.execute(f"SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(ItemWeight), DocNum, SUM(BatchQtyByCtn), MAX(Brand), 'PG', MAX(FromWhsCod), MAX(WhsCode) FROM PG_Transfer_Details WHERE DocNum IN ({placeholders}) GROUP BY DocNum, ItemCode", pg_nums)
            rows = await cursor.fetchall()
            for row in rows:
                weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
                products.append({"code": row[0] or "", "name": row[1] or "", "uom": row[2] or "", "weight": weight, "ctns": get_rounded_ctns(row[5] if len(row) > 5 else 0), "brand": row[6] or "", "bu": row[7], "sin_no": f"{row[7]} - {row[4]}", "FromWhsCode": row[8] or "", "ToWhsCode": row[9] or ""})
                
        if pdg_nums:
            placeholders = ','.join('?' * len(pdg_nums))
            await cursor.execute(f"SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(LineTotalWeight), DocNum, SUM(QtyCtn), MAX(Brand), 'PDG', MAX(FromWhsCod), MAX(WhsCode) FROM PDG_Transfer_Details WHERE DocNum IN ({placeholders}) GROUP BY DocNum, ItemCode", pdg_nums)
            rows = await cursor.fetchall()
            for row in rows:
                weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
                products.append({"code": row[0] or "", "name": row[1] or "", "uom": row[2] or "", "weight": weight, "ctns": get_rounded_ctns(row[5] if len(row) > 5 else 0), "brand": row[6] or "", "bu": row[7], "sin_no": f"{row[7]} - {row[4]}", "FromWhsCode": row[8] or "", "ToWhsCode": row[9] or ""})
        
        await conn.close()
        
        # --- NEW LOGIC: Calculate remaining amounts ---
        conn_log = await get_logistic_connection()
        cursor_log = await conn_log.cursor()
        
        sin_nos = list(set([p["sin_no"] for p in products]))
        if sin_nos:
            placeholders = ','.join('?' * len(sin_nos))
            await cursor_log.execute(f"""
                SELECT cp.sin_no, cp.code, SUM(COALESCE(cp.edited_ctns, cp.ctns)), SUM(COALESCE(cp.edited_weight, cp.weight))
                FROM Calculation_Products cp
                JOIN Calculation_History ch ON cp.calc_id = ch.id
                WHERE ch.status IN ('submitted', 'claimed')
                AND cp.sin_no IN ({placeholders})
                GROUP BY cp.sin_no, cp.code
            """, sin_nos)
            used_amounts_rows = await cursor_log.fetchall()
            
            used_map = {}
            for r in used_amounts_rows:
                used_map[(r[0], r[1])] = {
                    "used_ctns": Decimal(str(r[2])) if r[2] is not None else Decimal("0.0"),
                    "used_weight": Decimal(str(r[3])) if r[3] is not None else Decimal("0.0")
                }
                
            filtered_products = []
            for p in products:
                key = (p["sin_no"], p["code"])
                used = used_map.get(key, {"used_ctns": Decimal("0.0"), "used_weight": Decimal("0.0")})
                
                remaining_ctns = p["ctns"] - used["used_ctns"]
                remaining_weight = p["weight"] - used["used_weight"]
                
                if remaining_ctns < Decimal("0.01"): remaining_ctns = Decimal("0.0")
                if remaining_weight < Decimal("0.01"): remaining_weight = Decimal("0.0")
                
                p["original_total_ctns"] = p["ctns"]
                p["original_total_weight"] = p["weight"]
                p["used_ctns"] = used["used_ctns"]
                p["used_weight"] = used["used_weight"]
                p["ctns"] = remaining_ctns
                p["weight"] = remaining_weight
                filtered_products.append(p)
                    
            products = filtered_products
            
        await conn_log.close()
        
        # Recalculate total_weight 
        total_weight = sum(p["weight"] for p in products)
        # ----------------------------------------------
        
        return {"products": products, "total_weight": total_weight}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/products/{doc_num}")
async def get_products_by_doc_num(doc_num: str):
    try:
        is_pdg = doc_num.startswith("PDG")
        actual_num = doc_num.replace("PDG - ", "").replace("PDG-", "").replace("PG - ", "").replace("PG-", "")

        conn = await get_dwbi_connection()
        cursor = await conn.cursor()

        if is_pdg:
            await cursor.execute("SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(LineTotalWeight), DocNum, SUM(QtyCtn), MAX(Brand), 'PDG', MAX(FromWhsCod), MAX(WhsCode) FROM PDG_Transfer_Details WHERE DocNum = ? GROUP BY DocNum, ItemCode ORDER BY DocNum, ItemCode", (actual_num,))
        else:
            await cursor.execute("SELECT ItemCode, MAX(Dscription), MAX(UoM), SUM(ItemWeight), DocNum, SUM(BatchQtyByCtn), MAX(Brand), 'PG', MAX(FromWhsCod), MAX(WhsCode) FROM PG_Transfer_Details WHERE DocNum = ? GROUP BY DocNum, ItemCode ORDER BY DocNum, ItemCode", (actual_num,))
            
        rows = await cursor.fetchall()
        if not rows:
            await conn.close()
            raise HTTPException(status_code=404, detail="No products found")
            
        products, total_weight = [], Decimal("0.0")
        for row in rows:
            weight = Decimal(str(row[3])) if row[3] else Decimal("0.0")
            total_weight += weight
            products.append({"item_code": row[0] or "", "description": row[1] or "", "uom": row[2] or "", "item_weight": weight, "ctns": get_rounded_ctns(row[5] if len(row) > 5 else 0), "brand": row[6] or "", "bu": row[7], "FromWhsCode": row[8] or "", "ToWhsCode": row[9] or ""})
            
        await conn.close()
        return {"products": products, "total_weight": total_weight}
    except Exception as e: raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    
# --- Submitted Calculation Allocation Report Endpoints ---

@app.get("/account/submitted-allocation-report")
async def get_submitted_allocation_report(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user: dict = Depends(require_permission("view_daily_report"))
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()

        query = """
            SELECT
                ch.id, ch.submitted_at, ch.gate_name, ch.from_loc, ch.to_loc, ch.channel,
                cp.doc_date, cp.sin_no, cp.bu, cp.code, cp.name, cp.principal, cp.brand,
                cp.b_code, cp.b_name, cp.b_desc, cp.s_dept, cp.s_principal,
                COALESCE(cp.edited_ctns, cp.ctns) AS ctns, 
                COALESCE(cp.edited_weight, cp.weight) AS weight, 
                cp.total_cost, cp.unit_cost, cp.calculation_type
            FROM Calculation_History ch
            JOIN Calculation_Products cp ON cp.calc_id = ch.id
            WHERE ch.status IN ('submitted', 'claimed')
        """
        params = []

        if start_date and end_date:
            try:
                datetime.datetime.strptime(start_date, "%Y-%m-%d")
                datetime.datetime.strptime(end_date, "%Y-%m-%d")
                query += " AND SUBSTRING(ch.submitted_at, 1, 10) BETWEEN ? AND ?"
                params.extend([start_date, end_date])
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

        query += " ORDER BY ch.submitted_at DESC"

        await cursor.execute(query, params)
        rows = await cursor.fetchall()
        await conn.close()

        allocation_data = []
        for row in rows:
            ctns = Decimal(str(row[18] or 0))
            cost = Decimal(str(row[20] or 0))
            if ctns <= Decimal("0") and cost <= Decimal("0"):
                continue
            allocation_data.append({
                "calc_id": row[0],
                "submitted_at": row[1],
                "gate_name": row[2],
                "from_loc": row[3],
                "to_loc": row[4],
                "channel": row[5],
                "doc_date": row[6] or "",
                "sin_no": row[7] or "",
                "bu": row[8] or "",
                "item_code": row[9] or "",
                "item_name": row[10] or "",
                "principal": row[11] or "",
                "brand": row[12] or "",
                "b_code": row[13] or "",
                "b_name": row[14] or "",
                "b_desc": row[15] or "",
                "s_dept": row[16] or "",
                "s_principal": row[17] or "",
                "ctns": ctns,
                "weight": Decimal(str(row[19] or 0)),
                "total_cost": cost,
                "unit_cost": Decimal(str(row[21] or 0)),
                "calculation_type": row[22] or ""
            })

        return {
            "status": "success", 
            "total_records": len(allocation_data),
            "data": allocation_data
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating submitted allocation report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating submitted allocation report: {str(e)}")
    
# --- Dashboard Helper Functions ---

async def _generate_allocation_data(target_month: str, target_branch: Optional[str] = None):
    try:
        year, month = map(int, target_month.split('-'))
    except ValueError:
        raise ValueError("Invalid target_month format. Use YYYY-MM")

    months_list = []
    temp_year, temp_month = year, month
    for _ in range(12):
        months_list.append(f"{temp_year:04d}-{temp_month:02d}")
        temp_month -= 1
        if temp_month == 0:
            temp_month = 12
            temp_year -= 1

    valid_months = set(months_list)
    month0_label = months_list[0]

    start_date = f"{months_list[-1]}-01"
    _, last_day = calendar.monthrange(year, month)
    end_date = f"{year:04d}-{month:02d}-{last_day:02d}"

    conn = await get_logistic_connection()
    cursor = await conn.cursor()
    await cursor.execute("""
        SELECT target_date, bu, branch, driver_name, item_code, item_name, principal, brand,
               ctns, allocated_cost, cost_per_carton, driver_total_ctns, rate_cart_cost, sales_amount
        FROM Daily_Item_Report
        WHERE target_date >= ? AND target_date <= ?
    """, (start_date, end_date))
    rows = await cursor.fetchall()
    await conn.close()

    brands_data = {}
    available_branches = set()

    for row in rows:
        target_date_str = row[0]
        task_month = target_date_str[:7]

        if task_month not in valid_months:
            continue

        report_items = [{
            "bu": row[1], "branch": row[2], "driver_name": row[3], "item_code": row[4],
            "item_name": row[5], "principal": row[6], "brand": row[7],
            "ctns": row[8], "allocated_cost": row[9], "cost_per_carton": row[10],
            "driver_total_ctns": row[11], "rate_cart_cost": row[12], "sales_amount": row[13]
        }]

        for item in report_items:
            brand = item.get("brand", "").strip() or "UNKNOWN"
            branch = item.get("branch", "").strip() or "UNKNOWN"
            
            available_branches.add(branch)
            
            if target_branch and target_branch.strip().lower() != branch.lower():
                continue
            
            try:
                ctns = Decimal(str(item.get("ctns", "0.0") or "0.0"))
                allocated_cost = Decimal(str(item.get("allocated_cost", "0.0") or "0.0"))
            except ValueError:
                continue

            if ctns <= Decimal("0"):
                continue
                
            if brand not in brands_data:
                brands_data[brand] = {m: {"cost": Decimal("0.0"), "ctns": Decimal("0.0")} for m in months_list}

            brands_data[brand][task_month]["cost"] += allocated_cost
            brands_data[brand][task_month]["ctns"] += ctns

    dashboard_results = []
    for brand, data in brands_data.items():
        result = {"brand": brand}
        
        trend_data = []
        for m_label in reversed(months_list):
            t_cost = data[m_label]["cost"]
            t_ctns = data[m_label]["ctns"]
            avg_cost = (t_cost / t_ctns) if t_ctns > Decimal("0") else Decimal("0.0")
            
            trend_data.append({
                "month": m_label,
                "avg_cost": round(float(avg_cost), 2),
                "total_ctns": round(float(t_ctns), 2),
                "total_cost": round(float(t_cost), 2)
            })
        
        result["trend"] = trend_data
        dashboard_results.append(result)

    dashboard_results.sort(key=lambda x: x["trend"][-1]["total_ctns"], reverse=True)
    return dashboard_results, month0_label, sorted(list(available_branches))


async def _generate_calculated_data(target_month: str, target_to_loc: Optional[str] = None):
    try:
        year, month = map(int, target_month.split('-'))
    except ValueError:
        raise ValueError("Invalid target_month format. Use YYYY-MM")

    months_list = []
    temp_year, temp_month = year, month
    for _ in range(12):
        months_list.append(f"{temp_year:04d}-{temp_month:02d}")
        temp_month -= 1
        if temp_month == 0:
            temp_month = 12
            temp_year -= 1

    valid_months = set(months_list)
    month0_label = months_list[0]

    conn = await get_logistic_connection()
    cursor = await conn.cursor()

    await cursor.execute("""
        SELECT cp.brand, cp.ctns, cp.total_cost, cp.doc_date,
               COALESCE(ch.submitted_at, ch.created_at) AS fallback_date, ch.to_loc
        FROM Calculation_Products cp
        JOIN Calculation_History ch ON cp.calc_id = ch.id
        WHERE ch.status IN ('submitted', 'claimed')
    """)
    rows = await cursor.fetchall()
    await conn.close()

    brands_data = {}
    available_to_locs = set()

    for row in rows:
        brand = (row[0] or "").strip() or "UNKNOWN"
        try:
            ctns = Decimal(str(row[1] or 0))
            cost = Decimal(str(row[2] or 0))
        except ValueError:
            continue
        if ctns <= Decimal("0"):
            continue

        doc_date_str = str(row[3] or "")
        fallback_date = row[4][:10] if row[4] else ""
        to_loc = (row[5] or "UNKNOWN").strip()

        available_to_locs.add(to_loc)

        if target_to_loc and target_to_loc.strip().lower() != to_loc.lower():
            continue

        if len(doc_date_str) >= 7:
            task_month = doc_date_str[:7]
        elif len(fallback_date) >= 7:
            task_month = fallback_date[:7]
        else:
            continue

        if task_month in valid_months:
            if brand not in brands_data:
                brands_data[brand] = {m: {"cost": Decimal("0.0"), "ctns": Decimal("0.0")} for m in months_list}

            brands_data[brand][task_month]["cost"] += cost
            brands_data[brand][task_month]["ctns"] += ctns

    dashboard_results = []
    for brand, data in brands_data.items():
        result = {"brand": brand}
        
        trend_data = []
        for m_label in reversed(months_list):
            t_cost = data[m_label]["cost"]
            t_ctns = data[m_label]["ctns"]
            avg_cost = (t_cost / t_ctns) if t_ctns > Decimal("0") else Decimal("0.0")
            
            trend_data.append({
                "month": m_label,
                "avg_cost": round(float(avg_cost), 2),
                "total_ctns": round(float(t_ctns), 2),
                "total_cost": round(float(t_cost), 2)
            })
        
        result["trend"] = trend_data
        dashboard_results.append(result)

    dashboard_results.sort(key=lambda x: x["trend"][-1]["total_ctns"], reverse=True)
    return dashboard_results, month0_label, sorted(list(available_to_locs))


# --- Dashboard Endpoints ---

@app.get("/dashboard/brand-allocation-cost")
async def get_brand_allocation_cost_dashboard(
    target_month: Optional[str] = None, 
    branch: Optional[str] = Query(None, description="Filter by Branch"),
    user: dict = Depends(get_current_user)
):
    if not target_month:
        target_month = datetime.datetime.now().strftime("%Y-%m")
    try:
        results, month, available_branches = await _generate_allocation_data(target_month, target_branch=branch)
        return {
            "status": "success", 
            "target_month": month, 
            "data": results,
            "available_branches": available_branches
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating dashboard data: {str(e)}")
    
@app.get("/dashboard/calculated-brand-cost")
async def get_calculated_brand_cost_dashboard(
    target_month: Optional[str] = None, 
    to_loc: Optional[str] = Query(None, description="Filter by To Location"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    if not target_month:
        target_month = datetime.datetime.now().strftime("%Y-%m")
    try:
        results, month, available_to_locs = await _generate_calculated_data(target_month, target_to_loc=to_loc)
        return {
            "status": "success", 
            "target_month": month, 
            "data": results,
            "available_to_locs": available_to_locs
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating calculated dashboard data: {str(e)}")

@app.get("/dashboard/combined")
async def get_combined_dashboard(
    target_month: Optional[str] = None, 
    branch: Optional[str] = Query(None, description="Filter by Branch"),
    to_loc: Optional[str] = Query(None, description="Filter by To Location"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    if not target_month:
        target_month = datetime.datetime.now().strftime("%Y-%m")
        
    try:
        alloc_data, month, available_branches = await _generate_allocation_data(target_month, target_branch=branch)
        calc_data, _, available_to_locs = await _generate_calculated_data(target_month, target_to_loc=to_loc)
        
        return {
            "status": "success",
            "target_month": month,
            "allocation_data": alloc_data,
            "calculated_data": calc_data,
            "available_branches": available_branches,
            "available_to_locs": available_to_locs
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating combined dashboard data: {str(e)}")


@app.get("/dashboard/principal-brand-allocation")
async def get_principal_brand_allocation(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        query = """
            SELECT target_date, bu, branch, item_name, principal, brand, ctns, allocated_cost
            FROM Daily_Item_Report
        """
        params = []

        if start_date and end_date:
            query += " WHERE target_date BETWEEN ? AND ?"
            params.extend([start_date, end_date])
        elif start_date:
            query += " WHERE target_date >= ?"
            params.append(start_date)
        elif end_date:
            query += " WHERE target_date <= ?"
            params.append(end_date)

        await cursor.execute(query, params)
        rows = await cursor.fetchall()
        await conn.close()

        hierarchy = {}

        for row in rows:
            target_date = row[0]
            if not target_date: continue

            bu = str(row[1] or "").strip() or "UNKNOWN"
            branch = str(row[2] or "").strip() or "UNKNOWN"
            item_name = str(row[3] or "").strip() or "UNKNOWN"
            principal = str(row[4] or "").strip() or "UNKNOWN"
            brand = str(row[5] or "").strip() or "UNKNOWN"
            try:
                ctns = Decimal(str(row[6] or 0))
                cost = Decimal(str(row[7] or 0))
            except ValueError:
                continue

            if ctns <= Decimal("0"): continue

            for item in [{"bu": bu, "branch": branch, "principal": principal, "brand": brand,
                          "item_name": item_name, "ctns": ctns, "allocated_cost": cost}]:
                bu = str(item.get("bu", "")).strip() or "UNKNOWN"
                branch = str(item.get("branch", "")).strip() or "UNKNOWN"
                principal = str(item.get("principal", "")).strip() or "UNKNOWN"
                brand = str(item.get("brand", "")).strip() or "UNKNOWN"
                item_name = str(item.get("item_name", "")).strip() or "UNKNOWN"

                try:
                    ctns = Decimal(str(item.get("ctns", "0.0") or "0.0"))
                    cost = Decimal(str(item.get("allocated_cost", "0.0") or "0.0"))
                except ValueError:
                    continue

                if ctns <= Decimal("0"): continue

                if bu not in hierarchy: hierarchy[bu] = {}
                if branch not in hierarchy[bu]: hierarchy[bu][branch] = {}
                if principal not in hierarchy[bu][branch]: hierarchy[bu][branch][principal] = {}
                if brand not in hierarchy[bu][branch][principal]: hierarchy[bu][branch][principal][brand] = {}
                if item_name not in hierarchy[bu][branch][principal][brand]:
                    hierarchy[bu][branch][principal][brand][item_name] = {"ctns": Decimal("0.0"), "cost": Decimal("0.0"), "dates": {}}
                
                hierarchy[bu][branch][principal][brand][item_name]["ctns"] += ctns
                hierarchy[bu][branch][principal][brand][item_name]["cost"] += cost

                if target_date not in hierarchy[bu][branch][principal][brand][item_name]["dates"]:
                    hierarchy[bu][branch][principal][brand][item_name]["dates"][target_date] = {"ctns": Decimal("0.0"), "cost": Decimal("0.0")}
                
                hierarchy[bu][branch][principal][brand][item_name]["dates"][target_date]["ctns"] += ctns
                hierarchy[bu][branch][principal][brand][item_name]["dates"][target_date]["cost"] += cost

        result = []
        for bu, branch_data in hierarchy.items():
            bu_total_cost = Decimal("0.0"); bu_total_ctns = Decimal("0.0"); bu_branches = []
            
            for branch, princ_data in branch_data.items():
                b_total_cost = Decimal("0.0"); b_total_ctns = Decimal("0.0"); branch_principals = []
                
                for princ, brand_data in princ_data.items():
                    p_total_cost = Decimal("0.0"); p_total_ctns = Decimal("0.0"); princ_brands = []
                    
                    for brnd, item_data in brand_data.items():
                        br_total_cost = Decimal("0.0"); br_total_ctns = Decimal("0.0"); br_items = []
                        
                        for itm, d_data in item_data.items():
                            i_total_cost = d_data["cost"]; i_total_ctns = d_data["ctns"]
                            
                            i_dates = []
                            for date, d_vals in d_data["dates"].items():
                                d_avg_cost = d_vals["cost"] / d_vals["ctns"] if d_vals["ctns"] > Decimal("0") else Decimal("0.0")
                                i_dates.append({
                                    "date": date,
                                    "avg_cost": round(float(d_avg_cost), 2),
                                    "total_cost": round(float(d_vals["cost"]), 2),
                                    "total_ctns": round(float(d_vals["ctns"]), 2)
                                })
                            
                            i_dates.sort(key=lambda x: x["date"], reverse=True)
                            i_avg_cost = i_total_cost / i_total_ctns if i_total_ctns > Decimal("0") else Decimal("0.0")
                            
                            br_items.append({
                                "item_name": itm,
                                "avg_cost": round(float(i_avg_cost), 2),
                                "total_cost": round(float(i_total_cost), 2),
                                "total_ctns": round(float(i_total_ctns), 2),
                                "dates": i_dates
                            })
                            
                            br_total_cost += i_total_cost
                            br_total_ctns += i_total_ctns
                        
                        br_items.sort(key=lambda x: x["item_name"])
                        br_avg_cost = br_total_cost / br_total_ctns if br_total_ctns > Decimal("0") else Decimal("0.0")
                        
                        princ_brands.append({
                            "brand": brnd,
                            "avg_cost": round(float(br_avg_cost), 2),
                            "total_cost": round(float(br_total_cost), 2),
                            "total_ctns": round(float(br_total_ctns), 2),
                            "items": br_items
                        })
                        
                        p_total_cost += br_total_cost
                        p_total_ctns += br_total_ctns
                        
                    princ_brands.sort(key=lambda x: x["brand"])
                    p_avg_cost = p_total_cost / p_total_ctns if p_total_ctns > Decimal("0") else Decimal("0.0")
                    
                    branch_principals.append({
                        "principal": princ,
                        "avg_cost": round(float(p_avg_cost), 2),
                        "total_cost": round(float(p_total_cost), 2),
                        "total_ctns": round(float(p_total_ctns), 2),
                        "brands": princ_brands
                    })
                    
                    b_total_cost += p_total_cost
                    b_total_ctns += p_total_ctns
                
                branch_principals.sort(key=lambda x: x["principal"])
                b_avg_cost = b_total_cost / b_total_ctns if b_total_ctns > Decimal("0") else Decimal("0.0")
                
                bu_branches.append({
                    "branch": branch,
                    "avg_cost": round(float(b_avg_cost), 2),
                    "total_cost": round(float(b_total_cost), 2),
                    "total_ctns": round(float(b_total_ctns), 2),
                    "principals": branch_principals
                })
                
                bu_total_cost += b_total_cost
                bu_total_ctns += b_total_ctns
            
            bu_branches.sort(key=lambda x: x["branch"])
            bu_avg_cost = bu_total_cost / bu_total_ctns if bu_total_ctns > Decimal("0") else Decimal("0.0")
            
            result.append({
                "bu": bu,
                "avg_cost": round(float(bu_avg_cost), 2),
                "total_cost": round(float(bu_total_cost), 2),
                "total_ctns": round(float(bu_total_ctns), 2),
                "branches": bu_branches
            })
        
        result.sort(key=lambda x: x["bu"])
        return {"status": "success", "data": result}

    except Exception as e:
        logger.error(f"Error generating dynamic allocation hierarchy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating hierarchy data: {str(e)}")
    
@app.get("/dashboard/third-party-allocation")
async def get_third_party_allocation(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        query = """
            SELECT ch.submitted_at, ch.to_loc, cp.bu, cp.principal, cp.brand, cp.name,
                   COALESCE(cp.edited_ctns, cp.ctns) AS ctns, cp.total_cost
            FROM Calculation_History ch
            JOIN Calculation_Products cp ON cp.calc_id = ch.id
            WHERE ch.status IN ('submitted', 'claimed')
        """
        params = []

        if start_date and end_date:
            query += " AND SUBSTRING(ch.submitted_at, 1, 10) BETWEEN ? AND ?"
            params.extend([start_date, end_date])
        elif start_date:
            query += " AND SUBSTRING(ch.submitted_at, 1, 10) >= ?"
            params.append(start_date)
        elif end_date:
            query += " AND SUBSTRING(ch.submitted_at, 1, 10) <= ?"
            params.append(end_date)

        await cursor.execute(query, params)
        rows = await cursor.fetchall()
        await conn.close()

        hierarchy = {}

        for row in rows:
            submitted_at = row[0]
            target_date = submitted_at[:10] if submitted_at else ""
            if not target_date: continue

            to_loc = str(row[1] or "UNKNOWN").strip()
            bu = str(row[2] or "").strip() or "UNKNOWN"
            principal = str(row[3] or "").strip() or "UNKNOWN"
            brand = str(row[4] or "").strip() or "UNKNOWN"
            item_name = str(row[5] or "").strip() or "UNKNOWN"

            try:
                ctns = Decimal(str(row[6] or 0))
                cost = Decimal(str(row[7] or 0))
            except ValueError:
                continue

            if ctns <= Decimal("0"): continue

            for item in [{"bu": bu, "principal": principal, "brand": brand, "name": item_name, "ctns": ctns, "total_cost": cost}]:
                bu = str(item.get("bu", "")).strip() or "UNKNOWN"
                principal = str(item.get("principal", "")).strip() or "UNKNOWN"
                brand = str(item.get("brand", "")).strip() or "UNKNOWN"
                item_name = str(item.get("name", "")).strip() or "UNKNOWN"

                try:
                    ctns = Decimal(str(item.get("ctns", "0.0") or "0.0"))
                    cost = Decimal(str(item.get("total_cost", "0.0") or "0.0"))
                except ValueError:
                    continue

                if ctns <= Decimal("0"): continue

                if bu not in hierarchy: hierarchy[bu] = {}
                if to_loc not in hierarchy[bu]: hierarchy[bu][to_loc] = {}
                if principal not in hierarchy[bu][to_loc]: hierarchy[bu][to_loc][principal] = {}
                if brand not in hierarchy[bu][to_loc][principal]: hierarchy[bu][to_loc][principal][brand] = {}
                if item_name not in hierarchy[bu][to_loc][principal][brand]:
                    hierarchy[bu][to_loc][principal][brand][item_name] = {"ctns": Decimal("0.0"), "cost": Decimal("0.0"), "dates": {}}
                
                hierarchy[bu][to_loc][principal][brand][item_name]["ctns"] += ctns
                hierarchy[bu][to_loc][principal][brand][item_name]["cost"] += cost

                if target_date not in hierarchy[bu][to_loc][principal][brand][item_name]["dates"]:
                    hierarchy[bu][to_loc][principal][brand][item_name]["dates"][target_date] = {"ctns": Decimal("0.0"), "cost": Decimal("0.0")}
                
                hierarchy[bu][to_loc][principal][brand][item_name]["dates"][target_date]["ctns"] += ctns
                hierarchy[bu][to_loc][principal][brand][item_name]["dates"][target_date]["cost"] += cost

        result = []
        for bu, branch_data in hierarchy.items():
            bu_total_cost = Decimal("0.0"); bu_total_ctns = Decimal("0.0"); bu_branches = []
            
            for branch, princ_data in branch_data.items():
                b_total_cost = Decimal("0.0"); b_total_ctns = Decimal("0.0"); branch_principals = []
                
                for princ, brand_data in princ_data.items():
                    p_total_cost = Decimal("0.0"); p_total_ctns = Decimal("0.0"); princ_brands = []
                    
                    for brnd, item_data in brand_data.items():
                        br_total_cost = Decimal("0.0"); br_total_ctns = Decimal("0.0"); br_items = []
                        
                        for itm, d_data in item_data.items():
                            i_total_cost = d_data["cost"]; i_total_ctns = d_data["ctns"]
                            
                            i_dates = []
                            for date, d_vals in d_data["dates"].items():
                                d_avg_cost = d_vals["cost"] / d_vals["ctns"] if d_vals["ctns"] > Decimal("0") else Decimal("0.0")
                                i_dates.append({
                                    "date": date,
                                    "avg_cost": round(float(d_avg_cost), 2),
                                    "total_cost": round(float(d_vals["cost"]), 2),
                                    "total_ctns": round(float(d_vals["ctns"]), 2)
                                })
                            
                            i_dates.sort(key=lambda x: x["date"], reverse=True)
                            i_avg_cost = i_total_cost / i_total_ctns if i_total_ctns > Decimal("0") else Decimal("0.0")
                            
                            br_items.append({
                                "item_name": itm,
                                "avg_cost": round(float(i_avg_cost), 2),
                                "total_cost": round(float(i_total_cost), 2),
                                "total_ctns": round(float(i_total_ctns), 2),
                                "dates": i_dates
                            })
                            
                            br_total_cost += i_total_cost
                            br_total_ctns += i_total_ctns
                        
                        br_items.sort(key=lambda x: x["item_name"])
                        br_avg_cost = br_total_cost / br_total_ctns if br_total_ctns > Decimal("0") else Decimal("0.0")
                        
                        princ_brands.append({
                            "brand": brnd,
                            "avg_cost": round(float(br_avg_cost), 2),
                            "total_cost": round(float(br_total_cost), 2),
                            "total_ctns": round(float(br_total_ctns), 2),
                            "items": br_items
                        })
                        
                        p_total_cost += br_total_cost
                        p_total_ctns += br_total_ctns
                        
                    princ_brands.sort(key=lambda x: x["brand"])
                    p_avg_cost = p_total_cost / p_total_ctns if p_total_ctns > Decimal("0") else Decimal("0.0")
                    
                    branch_principals.append({
                        "principal": princ,
                        "avg_cost": round(float(p_avg_cost), 2),
                        "total_cost": round(float(p_total_cost), 2),
                        "total_ctns": round(float(p_total_ctns), 2),
                        "brands": princ_brands
                    })
                    
                    b_total_cost += p_total_cost
                    b_total_ctns += p_total_ctns
                
                branch_principals.sort(key=lambda x: x["principal"])
                b_avg_cost = b_total_cost / b_total_ctns if b_total_ctns > Decimal("0") else Decimal("0.0")
                
                bu_branches.append({
                    "branch": branch,
                    "avg_cost": round(float(b_avg_cost), 2),
                    "total_cost": round(float(b_total_cost), 2),
                    "total_ctns": round(float(b_total_ctns), 2),
                    "principals": branch_principals
                })
                
                bu_total_cost += b_total_cost
                bu_total_ctns += b_total_ctns
            
            bu_branches.sort(key=lambda x: x["branch"])
            bu_avg_cost = bu_total_cost / bu_total_ctns if bu_total_ctns > Decimal("0") else Decimal("0.0")
            
            result.append({
                "bu": bu,
                "avg_cost": round(float(bu_avg_cost), 2),
                "total_cost": round(float(bu_total_cost), 2),
                "total_ctns": round(float(bu_total_ctns), 2),
                "branches": bu_branches
            })
        
        result.sort(key=lambda x: x["bu"])
        return {"status": "success", "data": result}

    except Exception as e:
        logger.error(f"Error generating third party allocation hierarchy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating hierarchy data: {str(e)}")
    
# --- Cost Comparison Report Endpoint ---

@app.get("/dashboard/cost-comparison")
async def get_cost_comparison(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        conn = await get_logistic_connection()
        cursor = await conn.cursor()
        
        rc_query = """
            SELECT target_date, bu, branch, principal, brand, item_code, item_name, ctns, allocated_cost
            FROM Daily_Item_Report
            WHERE 1=1
        """
        rc_params = []
        if start_date and end_date:
            rc_query += " AND target_date BETWEEN ? AND ?"
            rc_params.extend([start_date, end_date])
        elif start_date:
            rc_query += " AND target_date >= ?"
            rc_params.append(start_date)
        elif end_date:
            rc_query += " AND target_date <= ?"
            rc_params.append(end_date)

        await cursor.execute(rc_query, rc_params)
        rc_rows = await cursor.fetchall()

        cc_query = """
            SELECT ch.submitted_at, ch.to_loc, cp.doc_date, cp.bu, cp.principal, cp.brand,
                   cp.code, cp.name, COALESCE(cp.edited_ctns, cp.ctns) AS ctns, cp.total_cost
            FROM Calculation_History ch
            JOIN Calculation_Products cp ON cp.calc_id = ch.id
            WHERE ch.status IN ('submitted', 'claimed')
        """
        await cursor.execute(cc_query)
        cc_rows = await cursor.fetchall()
        
        await cursor.execute("SELECT to_location, branch_code FROM Location_Mapping")
        mapping_rows = await cursor.fetchall()
        
        LOCATION_MAP = {}
        DISPLAY_MAP = {} 
        
        for row in mapping_rows:
            if row[0] and row[1]:
                raw_to_loc = str(row[0]).strip()
                raw_branch = str(row[1]).strip()
                LOCATION_MAP[raw_to_loc.upper()] = raw_branch.upper()
                DISPLAY_MAP[raw_branch.upper()] = raw_to_loc
        
        await conn.close()
        
        comparison_dict = {}
        
        for row in rc_rows:
            date = row[0]
            if not date: continue

            bu = str(row[1] or "").strip() or "UNKNOWN"
            branch = str(row[2] or "").strip().upper() or "UNKNOWN"
            principal = str(row[3] or "").strip() or "UNKNOWN"
            brand = str(row[4] or "").strip() or "UNKNOWN"
            item_code = str(row[5] or "").strip() or "UNKNOWN"
            item_name = str(row[6] or "").strip() or "UNKNOWN"
            try:
                ctns = Decimal(str(row[7] or 0))
                cost = Decimal(str(row[8] or 0))
            except ValueError:
                continue

            if ctns <= Decimal("0"): continue

            key = (date, bu, branch, principal, brand, item_code, item_name)
            if key not in comparison_dict:
                comparison_dict[key] = {"rc_ctns": Decimal("0.0"), "rc_cost": Decimal("0.0"), "cc_ctns": Decimal("0.0"), "cc_cost": Decimal("0.0")}

            comparison_dict[key]["rc_ctns"] += ctns
            comparison_dict[key]["rc_cost"] += cost
                
        for row in cc_rows:
            submitted_at = row[0]
            fallback_date = submitted_at[:10] if submitted_at else ""
            raw_to_loc = str(row[1] or "UNKNOWN").strip().upper()

            doc_date_raw = str(row[2] or "").strip()
            item_date = doc_date_raw[:10] if len(doc_date_raw) >= 10 else fallback_date
            if not item_date: continue

            if start_date and item_date < start_date: continue
            if end_date and item_date > end_date: continue

            bu = str(row[3] or "").strip() or "UNKNOWN"
            branch = LOCATION_MAP.get(raw_to_loc, raw_to_loc)
            principal = str(row[4] or "").strip() or "UNKNOWN"
            brand = str(row[5] or "").strip() or "UNKNOWN"
            item_code = str(row[6] or "").strip() or "UNKNOWN"
            item_name = str(row[7] or "").strip() or "UNKNOWN"

            try:
                ctns = Decimal(str(row[8] or 0))
                cost = Decimal(str(row[9] or 0))
            except ValueError:
                continue

            if ctns <= Decimal("0"): continue

            key = (item_date, bu, branch, principal, brand, item_code, item_name)
            if key not in comparison_dict:
                comparison_dict[key] = {"rc_ctns": Decimal("0.0"), "rc_cost": Decimal("0.0"), "cc_ctns": Decimal("0.0"), "cc_cost": Decimal("0.0")}

            comparison_dict[key]["cc_ctns"] += ctns
            comparison_dict[key]["cc_cost"] += cost
                
        result = []
        for key, val in comparison_dict.items():
            date, bu, branch, principal, brand, item_code, item_name = key
            
            rc_avg = val["rc_cost"] / val["rc_ctns"] if val["rc_ctns"] > Decimal("0") else None
            cc_avg = val["cc_cost"] / val["cc_ctns"] if val["cc_ctns"] > Decimal("0") else None
            
            if rc_avg is None and cc_avg is None:
                continue
                
            total_avg = Decimal("0.0")
            if rc_avg is not None: total_avg += rc_avg
            if cc_avg is not None: total_avg += cc_avg
            
            display_branch = branch
            if branch in DISPLAY_MAP:
                display_branch = f"{branch} / {DISPLAY_MAP[branch]}"
                
            result.append({
                "date": date,
                "bu": bu,
                "branch": display_branch, 
                "principal": principal,
                "brand": brand,
                "item_code": item_code,
                "item_name": item_name,
                "avg_cost_rate_cart": round(float(rc_avg), 2) if rc_avg is not None else None,
                "avg_cost_calculated": round(float(cc_avg), 2) if cc_avg is not None else None,
                "total_avg_cost": round(float(total_avg), 2)
            })
            
        result.sort(key=lambda x: (x["date"], x["branch"], x["item_code"]), reverse=True)
        
        return {
            "status": "success",
            "data": result
        }

    except Exception as e:
        logger.error(f"Error generating cost comparison report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating comparison data: {str(e)}")


# --- Shared Excel Builder Helper for Dashboard Exports ---

def _build_dashboard_excel(sheet_title: str, headers: list, rows: list, currency_cols: set):
    """Builds a styled .xlsx workbook (same style as the daily report export)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row_idx, row_values in enumerate(rows, 2):
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if (col_idx - 1) in currency_cols and isinstance(val, (int, float, Decimal)):
                cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _flatten_allocation_hierarchy(hierarchy_result: list, branch_label_key: str = "branch"):
    """Flattens bu -> branch -> principal -> brand -> item -> dates into flat leaf rows,
    mirroring the frontend's flattenHierarchy()."""
    flat_rows = []
    for bu_data in hierarchy_result:
        bu = bu_data.get("bu")
        for branch_data in bu_data.get("branches", []):
            branch = branch_data.get("branch")
            for princ_data in branch_data.get("principals", []):
                principal = princ_data.get("principal")
                for brand_data in princ_data.get("brands", []):
                    brand = brand_data.get("brand")
                    for item_data in brand_data.get("items", []):
                        item_name = item_data.get("item_name")
                        dates = item_data.get("dates", [])
                        if not dates:
                            flat_rows.append({
                                "bu": bu, branch_label_key: branch, "principal": principal, "brand": brand,
                                "item_name": item_name, "date": None,
                                "avg_cost": item_data.get("avg_cost", 0),
                                "total_ctns": item_data.get("total_ctns", 0),
                                "total_cost": item_data.get("total_cost", 0),
                            })
                        else:
                            for d in dates:
                                flat_rows.append({
                                    "bu": bu, branch_label_key: branch, "principal": principal, "brand": brand,
                                    "item_name": item_name, "date": d.get("date"),
                                    "avg_cost": d.get("avg_cost", 0),
                                    "total_ctns": d.get("total_ctns", 0),
                                    "total_cost": d.get("total_cost", 0),
                                })
    flat_rows.sort(key=lambda r: (str(r["bu"]), str(r[branch_label_key]), str(r["principal"]),
                                   str(r["brand"]), str(r["item_name"]), str(r["date"] or "")))
    return flat_rows


ALLOWED_DASHBOARD_COLUMNS = ['bu', 'branch', 'principal', 'brand', 'item', 'date']


def _parse_columns_param(columns: Optional[str]) -> list:
    """Parses the `columns` query param (comma-separated, in the exact order the
    frontend has them) and falls back to the full default set if missing/invalid."""
    if not columns:
        return list(ALLOWED_DASHBOARD_COLUMNS)
    requested = [c.strip() for c in columns.split(',') if c.strip()]
    valid = [c for c in requested if c in ALLOWED_DASHBOARD_COLUMNS]
    return valid if valid else list(ALLOWED_DASHBOARD_COLUMNS)


def _aggregate_flat_rows_by_columns(flat_rows: list, columns: list, branch_label_key: str):
    """
    Re-aggregates the finest-grain flattened rows down to only the requested column
    dimensions (in the given order), summing cartons/cost and recomputing avg cost.
    This mirrors exactly what the frontend's dynamic tree table does when a column
    is removed from columnOrder - its dimension gets folded into the totals instead
    of being broken out as its own row/column.
    """
    key_field_map = {
        'bu': 'bu',
        'branch': branch_label_key,
        'principal': 'principal',
        'brand': 'brand',
        'item': 'item_name',
        'date': 'date',
    }
    fields = [key_field_map[c] for c in columns if c in key_field_map]

    groups = {}
    for row in flat_rows:
        key = tuple(row.get(f) for f in fields)
        if key not in groups:
            groups[key] = {"ctns": Decimal("0"), "cost": Decimal("0")}
        groups[key]["ctns"] += Decimal(str(row.get("total_ctns") or 0))
        groups[key]["cost"] += Decimal(str(row.get("total_cost") or 0))

    result = []
    for key, vals in groups.items():
        avg = vals["cost"] / vals["ctns"] if vals["ctns"] > Decimal("0") else Decimal("0")
        entry = {f: (v if v is not None else "-") for f, v in zip(fields, key)}
        entry["avg_cost"] = float(avg)
        entry["total_ctns"] = float(vals["ctns"])
        entry["total_cost"] = float(vals["cost"])
        result.append(entry)

    result.sort(key=lambda r: tuple(str(r[f]) for f in fields))
    return result, fields


@app.get("/dashboard/principal-brand-allocation/export")
async def export_principal_brand_allocation(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    columns: Optional[str] = Query(None, description="Comma-separated column order, e.g. 'bu,principal'"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        payload = await get_principal_brand_allocation(start_date=start_date, end_date=end_date, user=user)
        flat_rows = _flatten_allocation_hierarchy(payload["data"], branch_label_key="branch")

        selected_columns = _parse_columns_param(columns)
        agg_rows, fields = _aggregate_flat_rows_by_columns(flat_rows, selected_columns, branch_label_key="branch")

        label_map = {'bu': 'BU', 'branch': 'Branch', 'principal': 'Principal', 'brand': 'Brand', 'item': 'Item Name', 'date': 'Date'}
        headers = [label_map[c] for c in selected_columns] + ["Avg Cost", "Total Cartons", "Total Allocated Cost"]

        rows = [
            [r[f] for f in fields] + [r["avg_cost"], r["total_ctns"], r["total_cost"]]
            for r in agg_rows
        ]

        currency_start = len(fields)
        output = await run_in_threadpool(
            _build_dashboard_excel,
            "Rate Cart Allocation", headers, rows,
            currency_cols={currency_start, currency_start + 1, currency_start + 2}
        )
        date_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"rate_cart_branch_allocation_{date_str}.xlsx"

        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting rate cart allocation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting rate cart allocation: {str(e)}")


@app.get("/dashboard/third-party-allocation/export")
async def export_third_party_allocation(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    columns: Optional[str] = Query(None, description="Comma-separated column order, e.g. 'bu,principal'"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        payload = await get_third_party_allocation(start_date=start_date, end_date=end_date, user=user)
        flat_rows = _flatten_allocation_hierarchy(payload["data"], branch_label_key="to_loc")

        selected_columns = _parse_columns_param(columns)
        agg_rows, fields = _aggregate_flat_rows_by_columns(flat_rows, selected_columns, branch_label_key="to_loc")

        label_map = {'bu': 'BU', 'branch': 'To Location', 'principal': 'Principal', 'brand': 'Brand', 'item': 'Item Name', 'date': 'Date'}
        headers = [label_map[c] for c in selected_columns] + ["Avg Cost", "Total Cartons", "Total Calculated Cost"]

        rows = [
            [r[f] for f in fields] + [r["avg_cost"], r["total_ctns"], r["total_cost"]]
            for r in agg_rows
        ]

        currency_start = len(fields)
        output = await run_in_threadpool(
            _build_dashboard_excel,
            "Third Party Allocation", headers, rows,
            currency_cols={currency_start, currency_start + 1, currency_start + 2}
        )
        date_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"third_party_calculated_cost_{date_str}.xlsx"

        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting third party allocation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting third party allocation: {str(e)}")


@app.get("/dashboard/cost-comparison/export")
async def export_cost_comparison(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    user: dict = Depends(require_permission("view_dashboard"))
):
    try:
        payload = await get_cost_comparison(start_date=start_date, end_date=end_date, user=user)
        data = payload["data"]

        headers = ["Date", "BU", "Branch/To Loc", "Principal", "Brand", "Item Code", "Item Name",
                   "Avg Cost (Rate Cart)", "Avg Cost (Calculated)", "Total Avg Cost"]
        rows = [
            [r["date"], r["bu"], r["branch"], r["principal"], r["brand"], r["item_code"], r["item_name"],
             float(r["avg_cost_rate_cart"]) if r["avg_cost_rate_cart"] is not None else "-",
             float(r["avg_cost_calculated"]) if r["avg_cost_calculated"] is not None else "-",
             float(r["total_avg_cost"] or 0)]
            for r in data
        ]

        output = await run_in_threadpool(
            _build_dashboard_excel, 
            "Cost Comparison", headers, rows, 
            currency_cols={7, 8, 9}
        )
        date_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"cost_comparison_{date_str}.xlsx"

        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting cost comparison: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting cost comparison: {str(e)}")


# --- Daily Report Excel Export Endpoints ---

@app.get("/account/daily-rate-cut-report/export")
async def export_daily_rate_cut_report(
    request: Request,
    report_type: str = Query(..., description="'item' or 'township'"),
    target_date: Optional[str] = None, 
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_permission("view_daily_report"))
):
    try:
        if start_date and end_date:
            start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            if start_dt > end_dt: 
                raise HTTPException(status_code=400, detail="start_date cannot be after end_date")
            
            daily_datas = []
            current_dt = start_dt
            while current_dt <= end_dt:
                daily_datas.append(await get_or_generate_daily_report(current_dt.strftime("%Y-%m-%d")))
                current_dt += datetime.timedelta(days=1)
            
            aggregated = _aggregate_reports(daily_datas)
            report_data = aggregated["item_report"] if report_type == 'item' else aggregated["township_report"]
            date_str = f"{start_date}_to_{end_date}"
        else:
            if not target_date: 
                target_date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            
            data = await get_or_generate_daily_report(target_date)
            report_data = data["item_report"] if report_type == 'item' else data["township_report"]
            date_str = target_date

        filters = dict(request.query_params)
        for k in ['report_type', 'target_date', 'start_date', 'end_date']:
            filters.pop(k, None) 
            
        if filters:
            filtered_data = []
            for row in report_data:
                match = True
                for key, val in filters.items():
                    actual_key = 'target_date' if key == 'date_filter' else key
                    row_val = str(row.get(actual_key, '')).lower()
                    if str(val).lower() not in row_val:
                        match = False
                        break
                if match:
                    filtered_data.append(row)
            report_data = filtered_data

        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Item Allocation" if report_type == 'item' else "Township Allocation"
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            border_style = Side(border_style="thin", color="000000")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            if report_type == 'item':
                headers = [
                    "BU", "Date", "Branch", "Driver Name", "Principal", "Brand", "Item Code", "Item Name", 
                    "Cartons", "Weight", "Volumetric Weight", "Driver Total (Ctns)", "Driver Total Weight", "Driver Total Volumetric Weight", "Original Driver Total (Ctns)", "Override Ctns", "Total Rate Cart Cost", "Original Rate Cart Cost", "Override Amount", "Cost per Carton",
                    "Allocated Cost", "Sales Amount"
                ]
            else:
                headers = [
                    "Branch", "Date", "Driver Name", "Township", "Customer Code", "Contact Person", 
                    "Customer Total (Ctns)", "Customer Total Weight", "Customer Total Volumetric Weight", "Driver Total (Ctns)", "Driver Total Weight", "Driver Total Volumetric Weight", "Original Driver Total (Ctns)", "Override Ctns", "Total Rate Cart Cost", "Original Rate Cart Cost", "Override Amount", "Total Drop Points",
                    "Cost per Drop Point", "Cost per Carton", "Allocated Cost", "Sales Amount"
                ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            for idx, row in enumerate(report_data, 2):
                driver_total = float(row.get("override_driver_total_ctns")) if row.get("override_driver_total_ctns") is not None else float(row.get("driver_total_ctns", 0))
                branch_cost = float(row.get("override_rate_cart_cost")) if row.get("override_rate_cart_cost") is not None else float(row.get("rate_cart_cost", 0))
                
                original_driver_total = float(row.get("driver_total_ctns", 0))
                original_branch_cost = float(row.get("rate_cart_cost", 0))
                
                if report_type == 'item':
                    row_data = [
                        row.get("bu", "-"), row.get("target_date", ""), row.get("branch", ""), row.get("driver_name", ""),
                        row.get("principal", ""), row.get("brand", ""), row.get("item_code", ""),
                        row.get("item_name", ""), float(row.get("ctns", 0)), float(row.get("weight", 0)), float(row.get("volumetric_weight", 0)),
                        driver_total, float(row.get("driver_total_weight", 0)), float(row.get("driver_total_volumetric_weight", 0)), original_driver_total,
                        float(row.get("override_ctns")) if row.get("override_ctns") is not None else None,
                        branch_cost, original_branch_cost,
                        float(row.get("override_amount")) if row.get("override_amount") is not None else None, 
                        float(row.get("cost_per_carton", 0)), float(row.get("allocated_cost", 0)),
                        float(row.get("sales_amount", 0))
                    ]
                else:
                    row_data = [
                        row.get("branch", ""), row.get("target_date", ""), row.get("driver_name", ""), row.get("township", ""),
                        row.get("customer_code", ""), row.get("contact_person", ""), float(row.get("ctns", 0)), float(row.get("weight", 0)), float(row.get("volumetric_weight", 0)),
                        driver_total, float(row.get("driver_total_weight", 0)), float(row.get("driver_total_volumetric_weight", 0)), original_driver_total,
                        float(row.get("override_ctns")) if row.get("override_ctns") is not None else None, 
                        branch_cost, original_branch_cost,
                        float(row.get("override_amount")) if row.get("override_amount") is not None else None, 
                        float(row.get("total_drop_points", 0)),
                        float(row.get("cost_per_drop_point", 0)), float(row.get("cost_per_carton", 0)), float(row.get("allocated_cost", 0)),
                        float(row.get("sales_amount", 0))
                    ]
                
                for col_num, val in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col_num, value=val)
                    cell.border = border
                    if isinstance(val, (int, float, Decimal)):
                        cell.number_format = '#,##0.00'

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: 
                            max_length = len(str(cell.value))
                    except: 
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        filename = f"{report_type}_allocation_{date_str}.xlsx"
        
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting daily report: {str(e)}")

@app.get("/account/submitted-allocation-report/export")
async def export_submitted_allocation_report(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    user: dict = Depends(require_permission("view_daily_report"))
):
    try:
        report_response = await get_submitted_allocation_report(start_date, end_date, user)
        allocation_data = report_response["data"]

        filters = dict(request.query_params)
        for k in ['start_date', 'end_date']:
            filters.pop(k, None)
            
        if filters:
            filtered_data = []
            for row in allocation_data:
                match = True
                for key, val in filters.items():
                    actual_key = 'submitted_at' if key == 'date_filter' else key
                    
                    if actual_key == 'route':
                        route_str = f"{row.get('from_loc', '')} {row.get('to_loc', '')}".lower()
                        if val.lower() not in route_str:
                            match = False
                            break
                    else:
                        row_val = str(row.get(actual_key, '')).lower()
                        if actual_key == 'submitted_at':
                            row_val = row_val[:10]
                            
                        if val.lower() not in row_val:
                            match = False
                            break
                if match:
                    filtered_data.append(row)
            allocation_data = filtered_data

        def build_excel():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Submitted Allocations"
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            border_style = Side(border_style="thin", color="000000")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            headers = [
                "Calc ID", "Date", "Doc Date", "SIN No", "Gate Name", "From Loc", "To Loc", "Channel",
                "BU", "Item Code", "Item Name", "Principal", "Brand", "B-Code", "B-Name", "B-Desc", "S-Dept", 
                "S-Principal", "Cartons", "Weight", "Unit Cost", "Total Cost", "Calculation Type"
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            for idx, row in enumerate(allocation_data, 2):
                raw_date = row.get("submitted_at", "")
                formatted_date = raw_date[:10] if raw_date else ""

                row_data = [
                    row.get("calc_id", ""), formatted_date, row.get("doc_date", ""),
                    row.get("sin_no", ""), row.get("gate_name", ""), row.get("from_loc", ""),
                    row.get("to_loc", ""), row.get("channel", ""), row.get("bu", ""),
                    row.get("item_code", ""), row.get("item_name", ""), row.get("principal", ""),
                    row.get("brand", ""), row.get("b_code", ""), row.get("b_name", ""),
                    row.get("b_desc", ""), row.get("s_dept", ""), row.get("s_principal", ""),
                    float(row.get("ctns", 0)), float(row.get("weight", 0)), float(row.get("unit_cost", 0)),
                    float(row.get("total_cost", 0)), row.get("calculation_type", "")
                ]
                
                for col_num, val in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col_num, value=val)
                    cell.border = border
                    if isinstance(val, (int, float, Decimal)):
                        cell.number_format = '#,##0.00'

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: 
                            max_length = len(str(cell.value))
                    except: 
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        output = await run_in_threadpool(build_excel)
        date_str = f"{start_date}_to_{end_date}" if start_date and end_date else "all_time"
        filename = f"submitted_allocation_report_{date_str}.xlsx"
        
        return StreamingResponse(
            output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting submitted report: {str(e)}")

if __name__ == "__main__":
    import uvicorn

    # Attach the same file handlers to uvicorn's loggers so server-level
    # logs (startup, access, errors) land in the same files as app logs.
    for uv_logger_name in ("uvicorn", "uvicorn.error", "uvicorn.access"):
        uv_logger = logging.getLogger(uv_logger_name)
        uv_logger.handlers = []
        uv_logger.propagate = True  # bubble up to root_logger's handlers

    uvicorn.run(app, host="0.0.0.0", port=8000, log_config=None)